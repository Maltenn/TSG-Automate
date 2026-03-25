#!/usr/bin/env python3
# TSG Automate Control App (PySide6)
# ------------------------------------------------------------------
# Buttons:
#   - Extract From PDF         → PDFExtract.py
#   - Place in Broberry Shop   → BroberryShop.py
#   - Download XML             → ShoptoPM.py
#   - Place in PM              → Launch PAD flow, wait for PM_Done.txt, then Add_PM_Nums.py
#   - Place Orders with Vendor        → PMtoWRG.py + PMtoARIAT.py
#   - Get Order IDs            → GetOrderId.py
#   - Next Wrangler Order ↩︎   → Sends Enter to PMtoWRG stdin
#   - Run All Steps ▶          → Extract → Broberry → Download XML → Place in PM → Wrangler
#   - Run to PM (Manual) ▶     → Extract → Broberry → Download XML → (Prompt PMs) → Add_PM_Nums.py
#   - Clear PDF Folder         → Deletes *.pdf/*.csv in profile's pdfs folder
#   - Open PDFs Folder, Open Processed Orders.xlsx, Kill Current Task ⛔
#
# Drag/drop: drop PDFs anywhere on the app to copy them to the profile's:
#   <Workspace Folder>\pdfs
# ------------------------------------------------------------------

import os
import sys
import shutil
import time
import threading
import subprocess
import json
import re
from dataclasses import dataclass

from PySide6 import QtCore, QtWidgets
from PySide6.QtCore import Qt, Signal, QDir

# =============================================================================
#  Visual Theme
# =============================================================================
DARK_STYLESHEET = """
/* ── Base ────────────────────────────────────────────────────────────────── */
QMainWindow, QDialog, QWidget {
    background-color: #0f1117;
    color: #cdd6f4;
    font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
    font-size: 10pt;
}

/* ── Group boxes ─────────────────────────────────────────────────────────── */
QGroupBox {
    background-color: #181825;
    border: 1px solid #313244;
    border-radius: 8px;
    margin-top: 14px;
    padding: 10px 8px 8px 8px;
    font-size: 8pt;
    font-weight: 600;
    letter-spacing: 0.08em;
    color: #6c7086;
    text-transform: uppercase;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    padding: 0 8px;
    left: 10px;
    color: #585b70;
}

/* ── Buttons (default) ───────────────────────────────────────────────────── */
QPushButton {
    background-color: #1e1e2e;
    color: #cdd6f4;
    border: 1px solid #313244;
    border-radius: 6px;
    padding: 7px 14px;
    min-height: 28px;
    text-align: left;
    font-size: 10pt;
}
QPushButton:hover {
    background-color: #252535;
    border-color: #45475a;
    color: #ffffff;
}
QPushButton:pressed { background-color: #16161e; }
QPushButton:disabled {
    background-color: #181825;
    color: #45475a;
    border-color: #1e1e2e;
}

/* ── Step buttons ────────────────────────────────────────────────────────── */
QPushButton[btnstyle="step"] {
    background-color: #1c2140;
    color: #89b4fa;
    border: 1px solid #2a3a6a;
}
QPushButton[btnstyle="step"]:hover {
    background-color: #243060;
    border-color: #4a6aaa;
    color: #b8d0ff;
}
QPushButton[btnstyle="step"]:pressed { background-color: #161830; }

/* ── Pipeline (go) buttons ───────────────────────────────────────────────── */
QPushButton[btnstyle="go"] {
    background-color: #162a20;
    color: #a6e3a1;
    border: 1px solid #254a35;
    font-weight: 600;
}
QPushButton[btnstyle="go"]:hover {
    background-color: #1e3a2a;
    border-color: #3a7a55;
    color: #c8f5c3;
}
QPushButton[btnstyle="go"]:pressed { background-color: #0f1e16; }

/* ── Danger button ───────────────────────────────────────────────────────── */
QPushButton[btnstyle="danger"] {
    background-color: #2a1518;
    color: #f38ba8;
    border: 1px solid #5a2530;
    font-weight: 600;
}
QPushButton[btnstyle="danger"]:hover {
    background-color: #3a1e22;
    border-color: #8a3545;
    color: #ffb3c6;
}
QPushButton[btnstyle="danger"]:pressed { background-color: #1a0c10; }

/* ── Verify (amber) button ───────────────────────────────────────────────── */
QPushButton[btnstyle="verify"] {
    background-color: #2a2210;
    color: #f9e2af;
    border: 1px solid #5a4820;
    font-weight: 600;
}
QPushButton[btnstyle="verify"]:hover {
    background-color: #3a3018;
    border-color: #8a6a30;
    color: #ffe8c0;
}
QPushButton[btnstyle="verify"]:disabled {
    background-color: #181610;
    color: #454030;
    border-color: #252010;
}
QPushButton[btnstyle="verify"]:pressed { background-color: #1a1508; }

/* ── Utility buttons ─────────────────────────────────────────────────────── */
QPushButton[btnstyle="util"] {
    background-color: #181825;
    color: #7f849c;
    border: 1px solid #28283a;
    font-size: 9pt;
    padding: 5px 14px;
    min-height: 24px;
}
QPushButton[btnstyle="util"]:hover {
    background-color: #202030;
    border-color: #3a3a55;
    color: #a0a8c0;
}

/* ── Manage button ───────────────────────────────────────────────────────── */
QPushButton[btnstyle="manage"] {
    background-color: #201e35;
    color: #cba6f7;
    border: 1px solid #3a3260;
    padding: 5px 12px;
    min-height: 28px;
    font-size: 9pt;
}
QPushButton[btnstyle="manage"]:hover {
    background-color: #2c2845;
    border-color: #5a4a90;
    color: #e0c8ff;
}

/* ── ComboBox ────────────────────────────────────────────────────────────── */
QComboBox {
    background-color: #1e1e2e;
    color: #cdd6f4;
    border: 1px solid #313244;
    border-radius: 6px;
    padding: 5px 10px;
    min-height: 28px;
}
QComboBox::drop-down {
    border: none;
    width: 22px;
    subcontrol-origin: padding;
    subcontrol-position: center right;
}
QComboBox::down-arrow {
    width: 10px;
    height: 10px;
}
QComboBox QAbstractItemView {
    background-color: #1e1e2e;
    color: #cdd6f4;
    border: 1px solid #313244;
    border-radius: 4px;
    selection-background-color: #313244;
    outline: none;
}

/* ── Line edits ──────────────────────────────────────────────────────────── */
QLineEdit {
    background-color: #13131f;
    color: #cdd6f4;
    border: 1px solid #313244;
    border-radius: 6px;
    padding: 5px 10px;
    selection-background-color: #4a4a8a;
}
QLineEdit:focus { border-color: #89b4fa; }
QLineEdit::placeholder { color: #45475a; }

/* ── List widget ─────────────────────────────────────────────────────────── */
QListWidget {
    background-color: #13131f;
    color: #cdd6f4;
    border: 1px solid #313244;
    border-radius: 6px;
    outline: none;
}
QListWidget::item { padding: 4px 8px; }
QListWidget::item:selected { background-color: #313244; color: #ffffff; border-radius: 3px; }
QListWidget::item:hover { background-color: #1e1e2e; }

/* ── Log (plain text) ────────────────────────────────────────────────────── */
QPlainTextEdit {
    background-color: #0b0b11;
    color: #a6e3a1;
    border: 1px solid #1e1e2e;
    border-radius: 8px;
    font-family: 'Cascadia Code', 'Consolas', 'Fira Code', monospace;
    font-size: 9pt;
    padding: 6px;
    selection-background-color: #313244;
}

/* ── Scrollbars ──────────────────────────────────────────────────────────── */
QScrollBar:vertical {
    background: #13131f;
    width: 8px;
    border-radius: 4px;
    margin: 0;
}
QScrollBar::handle:vertical {
    background: #313244;
    border-radius: 4px;
    min-height: 24px;
}
QScrollBar::handle:vertical:hover { background: #45475a; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: none; }

QScrollBar:horizontal {
    background: #13131f;
    height: 8px;
    border-radius: 4px;
}
QScrollBar::handle:horizontal {
    background: #313244;
    border-radius: 4px;
    min-width: 24px;
}
QScrollBar::handle:horizontal:hover { background: #45475a; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

/* ── Checkbox ────────────────────────────────────────────────────────────── */
QCheckBox { color: #7f849c; spacing: 6px; }
QCheckBox::indicator {
    width: 15px; height: 15px;
    border: 1px solid #313244;
    border-radius: 3px;
    background: #13131f;
}
QCheckBox::indicator:checked { background: #89b4fa; border-color: #89b4fa; }

/* ── Dialogs / message boxes ─────────────────────────────────────────────── */
QMessageBox { background-color: #181825; }
QMessageBox QLabel { color: #cdd6f4; font-size: 10pt; }
QDialog { background-color: #0f1117; }

/* ── Form layout labels ──────────────────────────────────────────────────── */
QFormLayout QLabel { color: #7f849c; font-size: 9pt; }

/* ── Splitter ────────────────────────────────────────────────────────────── */
QSplitter::handle { background-color: #313244; }
"""

# --- App base dir -------------------------------------------------------------
APP_DIR = os.path.dirname(os.path.abspath(__file__))
PYTHON_EXE = sys.executable

# --- Auto-updater -------------------------------------------------------------
# Set this to wherever you host the manifest.
#   GitHub:        "https://raw.githubusercontent.com/ORG/REPO/main/update_manifest.json"
#   Network share: r"\\server\share\TSG_Automate\update_manifest.json"
MANIFEST_URL = "https://raw.githubusercontent.com/YOUR_ORG/YOUR_REPO/main/update_manifest.json"

try:
    import app_updater as _app_updater
    _UPDATER_AVAILABLE = True
except ImportError:
    _app_updater = None          # type: ignore
    _UPDATER_AVAILABLE = False

# Power Automate flow (Place in PM)
FLOW_URL = r"ms-powerautomate:/console/flow/run?environmentid=Default-52a1c361-41d1-4e3f-bb5a-b47cd9e802d6&workflowid=90caa939-cf29-4681-94e4-e8cde8f340e3&source=Other"
PAD_EXE_CANDIDATES = [
    r"C:\Program Files (x86)\Power Automate Desktop\PAD.Console.Host.exe",
    r"C:\Program Files (x86)\Power Automate Desktop\dotnet\PAD.Console.Host.exe",
]
PM_DONE_TIMEOUT_SECS = 7200
PM_DONE_POLL_SECS = 2.0

# Profiles storage
PROFILES_JSON = os.path.join(APP_DIR, 'profiles.json')
DEFAULT_PROFILES = {
    "Default": {
        "admin_email": "",
        "admin_password": "",
        "initials": "MY",
        "wrg_email": "",
        "wrg_password": "",
        "ariat_email": "",
        "ariat_password": "",
        "propper_email": "",
        "propper_password": "",
        "workspace_dir": APP_DIR,
        "download_dir": ""
    }
}

def load_profiles():
    try:
        with open(PROFILES_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, dict) and data:
                # Backfill new keys so older JSONs still work
                for v in data.values():
                    v.setdefault("workspace_dir", APP_DIR)
                    v.setdefault("download_dir", "")
                    v.setdefault("wrg_email", "")
                    v.setdefault("wrg_password", "")
                    v.setdefault("initials", "MY")
                    v.setdefault("admin_email", "")
                    v.setdefault("admin_password", "")
                    v.setdefault("ariat_email", "")
                    v.setdefault("ariat_password", "")
                    v.setdefault("propper_email", "")
                    v.setdefault("propper_password", "")
                return data
    except Exception:
        pass
    return DEFAULT_PROFILES.copy()

def save_profiles(data: dict):
    try:
        with open(PROFILES_JSON, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass

# --- Optional: pywin32 (Excel save/close) ------------------------------------
try:
    import win32com.client  # type: ignore
except Exception:  # pragma: no cover
    win32com = None  # type: ignore

def ensure_processed_orders_closed_path(xlsx_path: str, log_fn):
    """If the given Processed_orders.xlsx path is open in Excel, save and close it."""
    if not xlsx_path or not os.path.isfile(xlsx_path):
        return
    if 'win32com' not in globals() or win32com is None:
        log_fn("[INFO] pywin32 not installed → skipping Excel auto-close. (pip install pywin32)")
        return
    try:
        try:
            xl = win32com.client.GetActiveObject('Excel.Application')
        except Exception:
            return
        for wb in list(xl.Workbooks):
            try:
                full = os.path.abspath(wb.FullName)
            except Exception:
                continue
            if os.path.abspath(xlsx_path).lower() == full.lower():
                try:
                    wb.Save()
                except Exception:
                    pass
                try:
                    wb.Close(SaveChanges=False)
                    log_fn("💾 Closed Processed_orders.xlsx before running step.")
                    try:
                        if xl.Workbooks.Count == 0:
                            xl.Quit()
                            log_fn("✅ Closed Excel application.")
                    except Exception:
                        pass
                except Exception as e:
                    log_fn(f"[WARN] Could not close Processed_orders.xlsx: {e}")
                break
    except Exception as e:
        log_fn(f"[WARN] Excel auto-close failed: {e}")

# Purge helper for PDFs/CSVs
def purge_pdfs_and_csvs(target_dir: str, log_fn):
    removed = 0
    if not target_dir:
        return
    # Safety: only purge inside a folder literally named "pdfs"
    if os.path.basename(os.path.normpath(target_dir)).lower() != "pdfs":
        log_fn(f"[WARN] Refusing to purge non-pdfs folder: {target_dir}")
        return
    try:
        for name in os.listdir(target_dir):
            if name.lower().endswith(('.pdf', '.csv')):
                p = os.path.join(target_dir, name)
                try:
                    os.remove(p)
                    removed += 1
                except Exception as e:
                    log_fn(f"[WARN] Could not delete {p}: {e}")
        if removed:
            log_fn(f"🧹 Purged {removed} file(s) from {target_dir}.")
    except FileNotFoundError:
        pass

# --- General helpers ----------------------------------------------------------
def file_signature(path: str):
    try:
        st = os.stat(path)
        with open(path, 'rb') as f:
            head = f.read(256)
        return (st.st_mtime, st.st_size, head)
    except FileNotFoundError:
        return None

def native_path(p: str) -> str:
    """Return a Windows-native path with backslashes; tolerant of empty/odd input."""
    if not p:
        return ""
    try:
        return QDir.toNativeSeparators(os.path.normpath(p))
    except Exception:
        # Fallback: replace slashes if Qt/QDir isn't available for some reason
        return p.replace("/", "\\")


# --- PM number entry dialog ---------------------------------------------------
class PMNumberEntryDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Enter PM Numbers")
        self.resize(520, 380)
        v = QtWidgets.QVBoxLayout(self)
        v.setSpacing(8)
        v.setContentsMargins(14, 14, 14, 14)

        lbl = QtWidgets.QLabel("Type a PM number and press Enter or Save:")
        lbl.setStyleSheet("color: #7f849c; font-size: 9pt;")
        v.addWidget(lbl)

        input_row = QtWidgets.QHBoxLayout()
        self.input = QtWidgets.QLineEdit()
        self.input.setPlaceholderText("PM number…")
        self.input.setFixedHeight(32)
        self.btn_save = QtWidgets.QPushButton("Save")
        self.btn_save.setFixedHeight(32)
        self.btn_save.setProperty("btnstyle", "step")
        input_row.addWidget(self.input, 1)
        input_row.addWidget(self.btn_save)
        v.addLayout(input_row)

        self.list = QtWidgets.QListWidget()
        v.addWidget(self.list, 1)

        btns = QtWidgets.QHBoxLayout()
        self.btn_clear = QtWidgets.QPushButton("Clear All")
        self.btn_done = QtWidgets.QPushButton("Done ✓")
        self.btn_cancel = QtWidgets.QPushButton("Cancel")
        self.btn_done.setProperty("btnstyle", "go")
        self.btn_clear.setProperty("btnstyle", "util")
        btns.addWidget(self.btn_clear)
        btns.addStretch(1)
        btns.addWidget(self.btn_cancel)
        btns.addWidget(self.btn_done)
        v.addLayout(btns)

        self.btn_save.clicked.connect(self.add_current)
        self.btn_clear.clicked.connect(self.clear_all)
        self.btn_done.clicked.connect(self.accept)
        self.btn_cancel.clicked.connect(self.reject)
        self.input.returnPressed.connect(self.add_current)

    def add_current(self):
        t = self.input.text().strip()
        if not t:
            return
        self.list.addItem(t)
        self.input.clear()
        self.input.setFocus()

    def clear_all(self):
        self.list.clear()

    def values(self):
        return [self.list.item(i).text() for i in range(self.list.count())]

# --- Manage Profiles Dialog ---------------------------------------------------
class ManageProfilesDialog(QtWidgets.QDialog):
    def __init__(self, profiles: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Manage Profiles")
        self.resize(860, 460)
        self.profiles = {k: dict(v) for k, v in profiles.items()}

        layout = QtWidgets.QHBoxLayout(self)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(14)

        # ── Left: profile list ──────────────────────────────────────────────
        left_widget = QtWidgets.QWidget()
        left_widget.setFixedWidth(180)
        left = QtWidgets.QVBoxLayout(left_widget)
        left.setContentsMargins(0, 0, 0, 0)
        left.setSpacing(6)

        list_label = QtWidgets.QLabel("PROFILES")
        list_label.setStyleSheet("color: #585b70; font-size: 8pt; font-weight: 600; letter-spacing: 0.08em;")
        left.addWidget(list_label)

        self.list = QtWidgets.QListWidget()
        self.list.addItems(self.profiles.keys())
        self.list.currentItemChanged.connect(self.on_select)
        left.addWidget(self.list, 1)

        btnsL = QtWidgets.QHBoxLayout()
        self.btn_add = QtWidgets.QPushButton("＋ Add")
        self.btn_del = QtWidgets.QPushButton("✕ Delete")
        self.btn_add.setProperty("btnstyle", "step")
        self.btn_del.setProperty("btnstyle", "danger")
        self.btn_add.clicked.connect(self.add_profile)
        self.btn_del.clicked.connect(self.delete_profile)
        btnsL.addWidget(self.btn_add)
        btnsL.addWidget(self.btn_del)
        left.addLayout(btnsL)
        layout.addWidget(left_widget)

        # ── Right: fields ───────────────────────────────────────────────────
        right_widget = QtWidgets.QWidget()
        right = QtWidgets.QFormLayout(right_widget)
        right.setFieldGrowthPolicy(QtWidgets.QFormLayout.AllNonFixedFieldsGrow)
        right.setSpacing(8)
        right.setContentsMargins(4, 0, 0, 0)

        self.name = QtWidgets.QLineEdit()
        self.admin_email = QtWidgets.QLineEdit()
        self.admin_password = QtWidgets.QLineEdit(); self.admin_password.setEchoMode(QtWidgets.QLineEdit.Password)
        self.initials = QtWidgets.QLineEdit()
        self.wrg_email = QtWidgets.QLineEdit()
        self.wrg_password = QtWidgets.QLineEdit(); self.wrg_password.setEchoMode(QtWidgets.QLineEdit.Password)
        self.ariat_email = QtWidgets.QLineEdit()
        self.ariat_password = QtWidgets.QLineEdit(); self.ariat_password.setEchoMode(QtWidgets.QLineEdit.Password)
        self.propper_email = QtWidgets.QLineEdit()
        self.propper_password = QtWidgets.QLineEdit(); self.propper_password.setEchoMode(QtWidgets.QLineEdit.Password)

        # Workspace
        self.workspace_dir = QtWidgets.QLineEdit()
        self.workspace_dir.setPlaceholderText("Path to your TSG_Automate folder")
        browse_btn = QtWidgets.QPushButton("Browse…")
        def pick_dir():
            d = QtWidgets.QFileDialog.getExistingDirectory(
            self, "Choose Workspace Folder", self.workspace_dir.text() or APP_DIR
            )
            if d:
                self.workspace_dir.setText(native_path(d))
        browse_btn.clicked.connect(pick_dir)
        ws_row = QtWidgets.QHBoxLayout()
        ws_row.addWidget(self.workspace_dir, 1)
        ws_row.addWidget(browse_btn, 0)

        # Download folder
        self.download_dir = QtWidgets.QLineEdit()
        self.download_dir.setPlaceholderText("Folder where browser downloads go")
        dl_browse = QtWidgets.QPushButton("Browse…")
        def pick_dl():
            d = QtWidgets.QFileDialog.getExistingDirectory(
                self, "Choose Download Folder", self.download_dir.text() or os.path.expanduser("~\\Downloads")
            )
            if d:
                self.download_dir.setText(native_path(d))
        dl_browse.clicked.connect(pick_dl)
        dl_row = QtWidgets.QHBoxLayout()
        dl_row.addWidget(self.download_dir, 1)
        dl_row.addWidget(dl_browse, 0)

        for le in (self.name, self.admin_email, self.admin_password,
                   self.initials, self.wrg_email, self.wrg_password,
                   self.ariat_email, self.ariat_password,
                   self.propper_email, self.propper_password,
                   self.workspace_dir, self.download_dir):
            le.setMinimumWidth(420)
            le.setFixedHeight(30)
            le.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)

        # Show toggles for passwords
        self.admin_show = QtWidgets.QCheckBox("Show")
        self.admin_show.toggled.connect(
            lambda checked: self.admin_password.setEchoMode(
                QtWidgets.QLineEdit.Normal if checked else QtWidgets.QLineEdit.Password
            )
        )
        self.wrg_show = QtWidgets.QCheckBox("Show")
        self.wrg_show.toggled.connect(
            lambda checked: self.wrg_password.setEchoMode(
                QtWidgets.QLineEdit.Normal if checked else QtWidgets.QLineEdit.Password
            )
        )
        self.ariat_show = QtWidgets.QCheckBox("Show")
        self.ariat_show.toggled.connect(
            lambda checked: self.ariat_password.setEchoMode(
                QtWidgets.QLineEdit.Normal if checked else QtWidgets.QLineEdit.Password
            )
        )
        self.propper_show = QtWidgets.QCheckBox("Show")
        self.propper_show.toggled.connect(
            lambda checked: self.propper_password.setEchoMode(
                QtWidgets.QLineEdit.Normal if checked else QtWidgets.QLineEdit.Password
            )
        )

        right.addRow("Name", self.name)
        right.addRow("Admin Email", self.admin_email)
        right.addRow("Admin Password", self.admin_password)
        right.addRow("", self.admin_show)
        right.addRow("Initials", self.initials)
        right.addRow("Wrangler Email", self.wrg_email)
        right.addRow("Wrangler Password", self.wrg_password)
        right.addRow("", self.wrg_show)
        right.addRow("Ariat Email", self.ariat_email)
        right.addRow("Ariat Password", self.ariat_password)
        right.addRow("", self.ariat_show)
        right.addRow("Propper Email", self.propper_email)
        right.addRow("Propper Password", self.propper_password)
        right.addRow("", self.propper_show)
        right.addRow("Workspace Folder", ws_row)
        right.addRow("Download Folder", dl_row)
        layout.addWidget(right_widget, 2)

        # ── Bottom: save / close ────────────────────────────────────────────
        outer_v = QtWidgets.QVBoxLayout()
        outer_v.addStretch(1)
        bottom = QtWidgets.QHBoxLayout()
        bottom.addStretch(1)
        self.btn_save = QtWidgets.QPushButton("💾  Save")
        self.btn_close = QtWidgets.QPushButton("Close")
        self.btn_save.setProperty("btnstyle", "go")
        self.btn_save.clicked.connect(self.save)
        self.btn_close.clicked.connect(self.accept)
        bottom.addWidget(self.btn_save)
        bottom.addWidget(self.btn_close)
        outer_v.addLayout(bottom)
        container = QtWidgets.QWidget()
        container.setLayout(outer_v)
        layout.addWidget(container, 0)

        if self.list.count():
            self.list.setCurrentRow(0)

    def on_select(self, cur, prev=None):
        if not cur:
            self.name.clear()
            self.admin_email.clear()
            self.admin_password.clear()
            self.initials.clear()
            self.wrg_email.clear()
            self.wrg_password.clear()
            self.workspace_dir.setText(APP_DIR)
            self.download_dir.setText(native_path(data.get('download_dir', "")))
            return

        n = cur.text()
        data = self.profiles.get(n, {})
        self.name.setText(n)
        self.admin_email.setText(data.get('admin_email', ''))
        self.admin_password.setText(data.get('admin_password', ''))
        self.initials.setText(data.get('initials', ''))
        self.wrg_email.setText(data.get('wrg_email', ''))
        self.wrg_password.setText(data.get('wrg_password', ''))
        self.ariat_email.setText(data.get('ariat_email', ''))
        self.ariat_password.setText(data.get('ariat_password', ''))
        self.propper_email.setText(data.get('propper_email', ''))
        self.propper_password.setText(data.get('propper_password', ''))
        self.workspace_dir.setText(data.get('workspace_dir', APP_DIR))
        self.download_dir.setText(data.get('download_dir', ""))

    def add_profile(self):
        base = "New Profile"
        n = base
        i = 1
        while n in self.profiles:
            i += 1
            n = f"{base} {i}"
        self.profiles[n] = {
            "admin_email": "", "admin_password": "", "initials": "MY",
            "wrg_email": "", "wrg_password": "",
            "ariat_email": "", "ariat_password": "",
            "propper_email": "", "propper_password": "",
            "workspace_dir": APP_DIR,
            "download_dir": ""
        }
        self.list.addItem(n)
        self.list.setCurrentRow(self.list.count()-1)

    def delete_profile(self):
        it = self.list.currentItem()
        if not it:
            return
        n = it.text()
        self.profiles.pop(n, None)
        row = self.list.row(it)
        self.list.takeItem(row)
        if self.list.count():
            self.list.setCurrentRow(min(row, self.list.count()-1))

    def save(self):
        it = self.list.currentItem()
        if it:
            old = it.text()
            new = self.name.text().strip() or old
            data = {
                "admin_email": self.admin_email.text().strip(),
                "admin_password": self.admin_password.text(),
                "initials": self.initials.text().strip() or "MY",
                "wrg_email": self.wrg_email.text().strip(),
                "wrg_password": self.wrg_password.text(),
                "ariat_email": self.ariat_email.text().strip(),
                "ariat_password": self.ariat_password.text(),
                "propper_email": self.propper_email.text().strip(),
                "propper_password": self.propper_password.text(),
                "workspace_dir": native_path(self.workspace_dir.text().strip() or APP_DIR),
                "download_dir": native_path(self.download_dir.text().strip()),
            }
            if new != old:
                self.profiles.pop(old, None)
                self.profiles[new] = data
                it.setText(new)
            else:
                self.profiles[old] = data
        save_profiles(self.profiles)
        QtWidgets.QMessageBox.information(self, "Saved", "Profiles saved.")

# --- Process Runner -----------------------------------------------------------
class ProcWorker(QtCore.QThread):
    line = Signal(str)
    finished = Signal(int)

    def __init__(self, cmd, cwd=None, env=None, stdin_pipe=False, parent=None):
        super().__init__(parent)
        self.cmd = cmd
        self.cwd = cwd
        self.env = env or os.environ.copy()
        self.env["PYTHONIOENCODING"] = "utf-8"
        self.env.setdefault("PYTHONUTF8", "1")
        self.stdin_pipe = stdin_pipe
        self.proc: subprocess.Popen | None = None

    def run(self):
        try:
            if isinstance(self.cmd, list) and self.cmd and self.cmd[0] == PYTHON_EXE:
                if '-u' not in self.cmd:
                    self.cmd.insert(1, '-u')
            self.proc = subprocess.Popen(
                self.cmd,
                cwd=self.cwd,
                env=self.env,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                stdin=(subprocess.PIPE if self.stdin_pipe else None),
                bufsize=1,
                encoding='utf-8',
                errors='replace',
            )
            assert self.proc.stdout is not None
            for line in self.proc.stdout:
                self.line.emit(line.rstrip('\n'))
            self.proc.wait()
            self.finished.emit(self.proc.returncode)
        except Exception as e:
            self.line.emit(f"[ERROR] {e}")
            self.finished.emit(1)

    def send_enter(self):
        if self.proc and self.proc.stdin:
            try:
                self.proc.stdin.write('\n')
                self.proc.stdin.flush()
            except Exception as e:
                self.line.emit(f"[ERROR] Failed to send Enter: {e}")

    def is_running(self) -> bool:
        try:
            return self.proc is not None and self.proc.poll() is None
        except Exception:
            return False

    def terminate_now(self):
        if not self.proc:
            return
        try:
            self.line.emit("⛔ Terminating current task...")
            self.proc.terminate()
        except Exception:
            pass
        try:
            self.proc.wait(timeout=2.0)
        except Exception:
            try:
                self.line.emit("⛔ Forcing kill...")
                self.proc.kill()
            except Exception:
                pass

# --- Drag & Drop Widget -------------------------------------------------------
class DropArea(QtWidgets.QFrame):
    def __init__(self, log_fn, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self._idle_style = (
            "QFrame { background-color:#0f1117; border: 2px dashed #313244; border-radius: 10px; }"
        )
        self._hover_style = (
            "QFrame { background-color:#131328; border: 2px dashed #89b4fa; border-radius: 10px; }"
        )
        self.setStyleSheet(self._idle_style)
        self._target_dir = ""
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)

        icon = QtWidgets.QLabel("📂")
        icon.setAlignment(Qt.AlignCenter)
        icon.setStyleSheet("font-size: 22px; border: none; background: transparent;")
        layout.addWidget(icon)

        self.title = QtWidgets.QLabel("")
        self.title.setAlignment(Qt.AlignCenter)
        self.title.setWordWrap(True)
        self.title.setStyleSheet(
            "font-size: 9pt; color: #585b70; border: none; background: transparent; padding: 2px;"
        )
        layout.addWidget(self.title)
        self.log_fn = log_fn

    def setTargetDir(self, path: str):
        self._target_dir = path
        self.title.setText("Drop PDFs here — copied to:\n" + path)

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            self.setStyleSheet(self._hover_style)
            e.acceptProposedAction()

    def dragLeaveEvent(self, e):
        self.setStyleSheet(self._idle_style)

    def dropEvent(self, e):
        self.setStyleSheet(self._idle_style)
        if not self._target_dir:
            self.log_fn("[ERROR] PDF target directory is not set.")
            return
        urls = e.mimeData().urls()
        copied = 0
        for u in urls:
            src = u.toLocalFile()
            if not src:
                continue
            if not src.lower().endswith('.pdf'):
                self.log_fn(f"Skipped non-PDF: {src}")
                continue
            dest = os.path.join(self._target_dir, os.path.basename(src))
            try:
                shutil.copy2(src, dest)
                self.log_fn(f"📄 Copied → {dest}")
                copied += 1
            except Exception as ex:
                self.log_fn(f"[ERROR] Copy failed for {src}: {ex}")
        if copied:
            self.log_fn(f"✅ {copied} file(s) copied to {self._target_dir}")

# --- Main Window --------------------------------------------------------------
@dataclass
class RunningState:
    pm_to_wrg: ProcWorker | None = None
    pm_to_ariat: ProcWorker | None = None
    pm_to_propper: ProcWorker | None = None
    interactive: ProcWorker | None = None

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("TSG Automate")
        self.setMinimumSize(760, 520)
        self.showMaximized()   # open full-screen by default

        self.running = RunningState()
        self.pipeline_active = False
        self.active_workers = []

        # ── Central widget ──────────────────────────────────────────────────
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        root = QtWidgets.QHBoxLayout(central)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        # ═══════════════════════════════════════════════════════════════════
        #  LEFT SIDEBAR  (scrollable so nothing clips at small heights)
        # ═══════════════════════════════════════════════════════════════════

        # Inner widget that holds all sidebar content
        sidebar_inner = QtWidgets.QWidget()
        sidebar_layout = QtWidgets.QVBoxLayout(sidebar_inner)
        sidebar_layout.setContentsMargins(0, 0, 6, 0)   # 6px right margin for scrollbar gap
        sidebar_layout.setSpacing(8)

        # ── App title ───────────────────────────────────────────────────────
        title_label = QtWidgets.QLabel("TSG AUTOMATE")
        title_label.setStyleSheet(
            "font-size: 13pt; font-weight: 700; color: #89b4fa;"
            " letter-spacing: 0.1em; padding: 4px 2px 8px 2px;"
        )
        sidebar_layout.addWidget(title_label)

        # ── Profile group ───────────────────────────────────────────────────
        prof_group = QtWidgets.QGroupBox("Active Profile")
        prof_layout = QtWidgets.QVBoxLayout(prof_group)
        prof_layout.setSpacing(6)

        self.profile_combo = QtWidgets.QComboBox()
        self.profiles = load_profiles()
        self.profile_combo.addItems(sorted(self.profiles.keys()))
        prof_layout.addWidget(self.profile_combo)

        self.btn_manage = self._btn("⚙  Manage Profiles", "manage")
        self.btn_manage.clicked.connect(self.manage_profiles)
        prof_layout.addWidget(self.btn_manage)

        sidebar_layout.addWidget(prof_group)
        self.profile_combo.currentIndexChanged.connect(
            lambda _: (self.log(f"👤 Profile: {self.profile_combo.currentText()}"),
                       self.refresh_paths_and_ui())
        )

        # ── Workflow steps group ────────────────────────────────────────────
        steps_group = QtWidgets.QGroupBox("Workflow Steps")
        steps_layout = QtWidgets.QVBoxLayout(steps_group)
        steps_layout.setSpacing(4)

        self.btn_extract  = self._btn("📄  Extract From PDF",         "step")
        self.btn_broberry = self._btn("🛒  Place in Broberry Shop",   "step")
        self.btn_download = self._btn("⬇  Download XML",              "step")
        self.btn_pad      = self._btn("📋  Place in PM",              "step")
        self.btn_wrg      = self._btn("📦  Place Orders with Vendor", "step")
        self.btn_wrg_only   = self._btn("🤠  Place with Wrangler",    "step")
        self.btn_ariat_only = self._btn("👢  Place with Ariat",       "step")
        self.btn_submit_ariat = self._btn("🚀  Submit Ariat Order",   "verify")
        self.btn_submit_ariat.setEnabled(False)
        self.btn_submit_ariat.setToolTip("Send Enter to PMtoARIAT.py — confirms the cart and submits the current order")
        self.btn_propper_only = self._btn("🪖  Place with Propper",   "step")
        self.btn_wrg_only.setToolTip("Run PMtoWRG.py directly")
        self.btn_ariat_only.setToolTip("Run PMtoARIAT.py directly")
        self.btn_propper_only.setToolTip("Run PMtoPropper.py directly")
        self.btn_get_order_ids = self._btn("🔍  Get Order IDs",       "step")

        self.btn_backorders = self._btn("♻  Place Back-Orders",        "step")
        self.btn_backorders.setToolTip("Place previously skipped (back-order) orders, then enter PM numbers")
        self.btn_verify = self._btn("✅  Verification Complete", "verify")
        self.btn_verify.setEnabled(False)
        self.btn_verify.setToolTip("Click after completing browser verification")

        for btn in (self.btn_extract, self.btn_broberry, self.btn_download,
                    self.btn_pad, self.btn_wrg,
                    self.btn_wrg_only, self.btn_ariat_only, self.btn_submit_ariat,
                    self.btn_propper_only,
                    self.btn_get_order_ids, self.btn_backorders, self.btn_verify):
            steps_layout.addWidget(btn)

        sidebar_layout.addWidget(steps_group)

        # ── Pipelines group ─────────────────────────────────────────────────
        pipe_group = QtWidgets.QGroupBox("Pipelines")
        pipe_layout = QtWidgets.QVBoxLayout(pipe_group)
        pipe_layout.setSpacing(4)

        self.btn_run_all   = self._btn("▶  Run All Steps",       "go")
        self.btn_run_to_pm = self._btn("▶  Run to PM (Manual)",  "go")
        self.btn_run_all.setToolTip("Extract → Broberry → Download → PM → Vendor")
        self.btn_run_to_pm.setToolTip("Extract → Broberry → Download → enter PM numbers manually → merge")

        pipe_layout.addWidget(self.btn_run_all)
        pipe_layout.addWidget(self.btn_run_to_pm)
        sidebar_layout.addWidget(pipe_group)

        # ── Utilities group ─────────────────────────────────────────────────
        util_group = QtWidgets.QGroupBox("Utilities")
        util_layout = QtWidgets.QVBoxLayout(util_group)
        util_layout.setSpacing(3)

        self.btn_clear_pdf       = self._btn("🧹  Clear PDF Folder",          "util")
        self.btn_open_pdfs       = self._btn("📁  Open PDFs Folder",          "util")
        self.btn_open_processed  = self._btn("📊  Processed Orders.xlsx",     "util")
        self.btn_open_skipped    = self._btn("⚠   Skipped Orders.xlsx",       "util")

        for btn in (self.btn_clear_pdf, self.btn_open_pdfs,
                    self.btn_open_processed, self.btn_open_skipped):
            util_layout.addWidget(btn)

        self.btn_update_app = self._btn("🔄  Update App", "manage")
        self.btn_update_app.setToolTip("Check for and download the latest scripts from the update server")
        if not _UPDATER_AVAILABLE:
            self.btn_update_app.setEnabled(False)
            self.btn_update_app.setToolTip("app_updater.py not found in app folder")
        util_layout.addWidget(self.btn_update_app)

        self.btn_kill = self._btn("⛔  Kill Current Task", "danger")
        util_layout.addWidget(self.btn_kill)
        sidebar_layout.addWidget(util_group)

        sidebar_layout.addStretch(1)

        # Wrap inner widget in a scroll area — no clipping even on small screens
        sidebar_scroll = QtWidgets.QScrollArea()
        sidebar_scroll.setWidgetResizable(True)
        sidebar_scroll.setWidget(sidebar_inner)
        sidebar_scroll.setFixedWidth(248)          # 230 content + scrollbar room
        sidebar_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        sidebar_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        sidebar_scroll.setFrameShape(QtWidgets.QFrame.NoFrame)
        sidebar_scroll.setStyleSheet(
            "QScrollArea { background: transparent; border: none; }"
            "QScrollArea > QWidget > QWidget { background: transparent; }"
        )
        root.addWidget(sidebar_scroll)

        # ═══════════════════════════════════════════════════════════════════
        #  RIGHT PANEL
        # ═══════════════════════════════════════════════════════════════════
        right = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(8)

        # ── Drop area ───────────────────────────────────────────────────────
        self.drop = DropArea(self.log)
        self.drop.setFixedHeight(90)
        right_layout.addWidget(self.drop)

        # ── Activity log ────────────────────────────────────────────────────
        log_group = QtWidgets.QGroupBox("Activity Log")
        log_gl = QtWidgets.QVBoxLayout(log_group)
        log_gl.setContentsMargins(6, 6, 6, 6)

        self.log_box = QtWidgets.QPlainTextEdit()
        self.log_box.setReadOnly(True)
        log_gl.addWidget(self.log_box)
        right_layout.addWidget(log_group, 1)

        root.addWidget(right, 1)

        # ── Wire up buttons ─────────────────────────────────────────────────
        self.btn_extract.clicked.connect(lambda: self.run_script("PDFExtract.py"))
        self.btn_broberry.clicked.connect(lambda: self.run_script("BroberryShop.py"))
        self.btn_download.clicked.connect(lambda: self.run_script("ShoptoPM.py"))
        self.btn_pad.clicked.connect(self.run_pad_flow_sequence)
        self.btn_wrg.clicked.connect(self.run_orders_with_vendor)
        self.btn_wrg_only.clicked.connect(self.run_pm_to_wrg)
        self.btn_ariat_only.clicked.connect(self.run_pm_to_ariat)
        self.btn_submit_ariat.clicked.connect(self.submit_ariat_order)
        self.btn_propper_only.clicked.connect(self.run_pm_to_propper)
        self.btn_get_order_ids.clicked.connect(self.run_get_order_ids)
        self.btn_backorders.clicked.connect(self.run_backorders_then_pm)
        self.btn_verify.clicked.connect(self.verification_complete)
        self.btn_run_all.clicked.connect(self.run_all_steps)
        self.btn_run_to_pm.clicked.connect(self.run_to_pm_manual)
        self.btn_clear_pdf.clicked.connect(self.clear_pdf_folder)
        self.btn_open_pdfs.clicked.connect(self.open_pdfs_folder)
        self.btn_open_processed.clicked.connect(self.open_processed_orders)
        self.btn_open_skipped.clicked.connect(self.open_skipped_orders)
        self.btn_kill.clicked.connect(self.kill_current_task)
        self.btn_update_app.clicked.connect(self.update_app)

        # ── Initialize paths/UI for selected profile ────────────────────────
        self.refresh_paths_and_ui()

    # ── Button factory ───────────────────────────────────────────────────────
    def _btn(self, text: str, style: str = "") -> QtWidgets.QPushButton:
        b = QtWidgets.QPushButton(text)
        if style:
            b.setProperty("btnstyle", style)
        return b


    # --- Kill current task ---
    def kill_current_task(self):
        cand = None
        if self.running.pm_to_wrg and self.running.pm_to_wrg.is_running():
            cand = self.running.pm_to_wrg
        if cand is None:
            for w in reversed(self.active_workers):
                if getattr(w, "is_running", None) and w.is_running():
                    cand = w
                    break
        if cand is None:
            self.log("[INFO] No running task to kill.")
            return
        try:
            cand.terminate_now()
            self.log("⛔ Current task terminated by user.")
        except Exception as e:
            self.log(f"[WARN] Could not terminate task: {e}")
        if cand is self.running.pm_to_wrg:
            self.running.pm_to_wrg = None
        if getattr(self, 'pipeline_active', False):
            self.pipeline_active = False
            self.log("◆ Pipeline cancelled after manual kill. No further steps will run.")

    # --- App self-updater ---
    def update_app(self):
        if not _UPDATER_AVAILABLE:
            self.log("[ERROR] app_updater.py is missing from the app folder.")
            return
        if getattr(self, 'pipeline_active', False):
            self.log("[INFO] Cannot update while a pipeline is running.")
            QtWidgets.QMessageBox.warning(
                self, "Update Blocked",
                "Please wait for the current pipeline to finish before updating."
            )
            return

        self.btn_update_app.setEnabled(False)
        self.log("─" * 40)
        self.log("🔄 Starting app update…")

        def _do_update():
            main_app_updated = _app_updater.check_and_update(
                manifest_url=MANIFEST_URL,
                app_dir=APP_DIR,
                log=self.log,
                parent_window=self,
            )
            if main_app_updated:
                reply = QtWidgets.QMessageBox.question(
                    self,
                    "Restart Required",
                    "The main application file was updated.\n\n"
                    "Restart TSG Automate now to apply changes?",
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                )
                if reply == QtWidgets.QMessageBox.Yes:
                    self.log("♻ Restarting…")
                    QtCore.QTimer.singleShot(
                        300,
                        lambda: (
                            subprocess.Popen([PYTHON_EXE, os.path.join(APP_DIR, "tsg_automate_app.py")]),
                            QtWidgets.QApplication.quit(),
                        )
                    )
            self.btn_update_app.setEnabled(True)

        threading.Thread(target=_do_update, daemon=True).start()

    # --- Path helpers (per-profile) ---
    def current_workspace(self) -> str:
        name = self.profile_combo.currentText().strip()
        data = self.profiles.get(name, {}) if hasattr(self, 'profiles') else {}
        ws = native_path((data.get("workspace_dir") or APP_DIR).strip())
        return os.path.abspath(ws)

    def paths(self) -> dict:
        ws = self.current_workspace()
        return {
            "ws": ws,
            "pdfs": os.path.join(ws, "pdfs"),
            "pm_done": os.path.join(ws, "PM_Done.txt"),
            "processed": os.path.join(ws, "Processed_orders.xlsx"),
            "skipped": os.path.join(ws, "skipped_orders.xlsx"),
            "pmnum": os.path.join(ws, "PMNum.xlsx"),
            "scripts": {
                "PDFExtract.py": os.path.join(ws, "PDFExtract.py"),
                "BroberryShop.py": os.path.join(ws, "BroberryShop.py"),
                "ShoptoPM.py": os.path.join(ws, "ShoptoPM.py"),
                "Add_PM_Nums.py": os.path.join(ws, "Add_PM_Nums.py"),
                "PMtoWRG.py": os.path.join(ws, "PMtoWRG.py"),
                "PMtoARIAT.py": os.path.join(ws, "PMtoARIAT.py"),
                "PMtoPropper.py": os.path.join(ws, "PMtoPropper.py"),
                "GetOrderId.py": os.path.join(ws, "GetOrderId.py"),
                "BroberryShop_Backorders.py": os.path.join(ws, "BroberryShop_Backorders.py"),
            }
        }

    def refresh_paths_and_ui(self):
        p = self.paths()
        os.makedirs(p["pdfs"], exist_ok=True)
        self.drop.setTargetDir(native_path(p["pdfs"]))   # title text shows backslashes
        self.log(f"📁 Workspace: {native_path(p['ws'])}")

    # --- Logging helpers ---
    def log(self, msg: str):
        self.log_box.appendPlainText(msg)
        self.log_box.verticalScrollBar().setValue(self.log_box.verticalScrollBar().maximum())
        print(msg)

    # --- File ops ---
    def open_pdfs_folder(self):
        try:
            os.startfile(self.paths()["pdfs"])
        except Exception as e:
            self.log(f"[ERROR] Cannot open folder: {e}")

    def open_processed_orders(self):
        xlsx = self.paths()["processed"]
        if os.path.isfile(xlsx):
            try:
                os.startfile(xlsx)
            except Exception as e:
                self.log(f"[ERROR] Cannot open Processed_orders.xlsx: {e}")
        else:
            self.log(f"[ERROR] Not found: {xlsx}")

    def open_skipped_orders(self):
        xlsx = self.paths()["skipped"]
        if os.path.isfile(xlsx):
            try:
                os.startfile(xlsx)
            except Exception as e:
                self.log(f"[ERROR] Cannot open skipped_orders.xlsx: {e}")
        else:
            self.log("[INFO] No skipped_orders.xlsx found — no orders were skipped this run.")

    # --- Profiles ---
    def manage_profiles(self):
        dlg = ManageProfilesDialog(self.profiles, self)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            cur = self.profile_combo.currentText()
            self.profiles = load_profiles()
            self.profile_combo.blockSignals(True)
            self.profile_combo.clear()
            self.profile_combo.addItems(sorted(self.profiles.keys()))
            if cur in self.profiles:
                idx = self.profile_combo.findText(cur)
                self.profile_combo.setCurrentIndex(idx)
            self.profile_combo.blockSignals(False)
            self.log("✅ Profiles updated.")
            self.refresh_paths_and_ui()

    def profile_env(self) -> dict:
        name = self.profile_combo.currentText().strip() if hasattr(self, 'profile_combo') else ''
        data = self.profiles.get(name, {}) if hasattr(self, 'profiles') else {}
        env = {}
        if data.get('admin_email') and data.get('admin_password'):
            env['BROBERRY_ADMIN_EMAIL'] = data['admin_email']
            env['BROBERRY_ADMIN_PASSWORD'] = data['admin_password']
        if data.get('initials'):
            env['ORDER_USER_INITIALS'] = data['initials']
        if data.get('wrg_email') and data.get('wrg_password'):
            env['WRANGLER_EMAIL']    = data['wrg_email']
            env['WRANGLER_PASSWORD'] = data['wrg_password']
        if data.get('ariat_email') and data.get('ariat_password'):
            env['ARIAT_USERNAME'] = data['ariat_email']
            env['ARIAT_PASSWORD'] = data['ariat_password']
        if data.get('propper_email') and data.get('propper_password'):
            env['PROPPER_USERNAME'] = data['propper_email']
            env['PROPPER_PASSWORD'] = data['propper_password']
        env['TSG_WORKSPACE_DIR'] = native_path(self.current_workspace())
        dl = (data.get('download_dir') or "").strip()
        if dl:
            dl = native_path(dl)
            os.makedirs(dl, exist_ok=True)
            env['TSG_DOWNLOAD_DIR'] = dl

        return env

    # --- Process execution helpers ---
    def run_script(self, script_name: str, stdin_pipe: bool=False, label: str | None=None):
        p = self.paths()
        ensure_processed_orders_closed_path(p["processed"], self.log)
        script_path = p["scripts"].get(script_name, script_name)
        if not os.path.isfile(script_path):
            self.log(f"[ERROR] Script not found: {script_path}")
            return None
        tag = label or os.path.basename(script_path)
        self.log(f"▶︎ Running {tag}…")
        env = os.environ.copy()
        env.update(self.profile_env())
        worker = ProcWorker([PYTHON_EXE, script_path], cwd=p["ws"], stdin_pipe=stdin_pipe, env=env)
        worker.line.connect(self.log)
        def _done(rc):
            self.log(f"◆ {tag} exited with code {rc}")
            if worker is self.running.pm_to_wrg:
                self.btn_verify.setEnabled(False)
                self.running.pm_to_wrg = None
                purge_pdfs_and_csvs(self.paths()["pdfs"], self.log)
            try:
                self.active_workers.remove(worker)
            except ValueError:
                pass
        worker.finished.connect(_done)
        self.active_workers.append(worker)
        worker.start()
        return worker

    # --- PAD flow → wait on PM_Done.txt → Add_PM_Nums.py ---
    def run_pad_flow_sequence(self, on_complete=None):
        p = self.paths()
        ensure_processed_orders_closed_path(p["processed"], self.log)
        baseline = file_signature(p["pm_done"])
        self.log(f"⌛ Waiting for PM update signal: {p['pm_done']}")

        started = False
        for exe in PAD_EXE_CANDIDATES:
            if os.path.isfile(exe):
                try:
                    subprocess.Popen([exe, FLOW_URL], env={**os.environ, **self.profile_env()})
                    started = True
                    break
                except Exception as e:
                    self.log(f"[WARN] PAD console host failed: {e}")
        if not started:
            try:
                subprocess.Popen(["cmd", "/c", "start", "", FLOW_URL], close_fds=True)
                started = True
            except Exception as e:
                self.log(f"[ERROR] Could not start PAD flow: {e}")
                if on_complete:
                    on_complete(1)
                return
        self.log("…Place in PM launched.")

        def _wait_then_merge():
            start = time.time()
            while True:
                cur = file_signature(p["pm_done"])
                if (baseline is None and cur is not None) or (baseline is not None and cur != baseline):
                    self.log("✅ Detected PM_Done.txt update.")
                    break
                if time.time() - start > PM_DONE_TIMEOUT_SECS:
                    self.log("[ERROR] Timed out waiting for PM_Done.txt to update.")
                    if on_complete:
                        on_complete(1)
                    return
                time.sleep(PM_DONE_POLL_SECS)
            self.log("▶︎ Running Add_PM_Nums.py…")
            worker = self.run_script("Add_PM_Nums.py", label='Add_PM_Nums.py')
            if worker:
                def _after_merge(rc):
                    if on_complete:
                        on_complete(rc)
                worker.finished.connect(_after_merge)
            elif on_complete:
                on_complete(1)

        threading.Thread(target=_wait_then_merge, daemon=True).start()

    # --- Wrangler flow ---
    def run_pm_to_wrg(self):
        if self.running.pm_to_wrg is not None:
            self.log("[INFO] PMtoWRG is already running.")
            return None
        worker = self.run_script("PMtoWRG.py", stdin_pipe=True, label='PMtoWRG.py')
        if worker is not None:
            self.running.pm_to_wrg = worker
            self.running.interactive = worker
            self.btn_verify.setEnabled(True)

            # Guidance message so users know the next step
            self.log("🧩 Complete Wrangler login verification in the browser, then press “Verification Complete ✅”.")
        return worker

    def run_pm_to_ariat(self):
        if getattr(self.running, "pm_to_ariat", None) is not None and self.running.pm_to_ariat.is_running():
            self.log("[INFO] PMtoARIAT is already running.")
            return None
        worker = self.run_script("PMtoARIAT.py", stdin_pipe=True, label="PMtoARIAT.py")
        if worker is not None:
            self.running.pm_to_ariat = worker
            self.running.interactive = worker
            self.btn_verify.setEnabled(True)
            self.btn_submit_ariat.setEnabled(True)
            self.log("👢 Ariat running — review the cart in the browser, then click Submit Ariat Order to place it.")

            def _ariat_done(rc):
                self.btn_submit_ariat.setEnabled(False)
            worker.finished.connect(_ariat_done)
        return worker

    def submit_ariat_order(self):
        """Send Enter to the running PMtoARIAT.py process to confirm and submit the current order."""
        w = getattr(self.running, "pm_to_ariat", None)
        if w is None or not w.is_running():
            self.log("[INFO] No Ariat script is currently waiting.")
            return
        w.send_enter()
        self.log("🚀 Sent submit signal to Ariat — order submission in progress...")


    def run_pm_to_propper(self):
        if getattr(self.running, "pm_to_propper", None) is not None and self.running.pm_to_propper.is_running():
            self.log("[INFO] PMtoPropper is already running.")
            return None
        worker = self.run_script("PMtoPropper.py", stdin_pipe=True, label='PMtoPropper.py')
        if worker is not None:
            self.running.pm_to_propper = worker
            self.running.interactive = worker
            self.btn_verify.setEnabled(True)
            self.log('\U0001f9e9 Propper will pause for manual submit; press "Verification Complete \u2705" to continue when prompted.')
        return worker

    def run_orders_with_vendor(self, on_complete=None):
        """
        Reads Processed_orders.xlsx and routes orders based on Column K (vendor).
        - Wrangler rows -> PMtoWRG.py
        - Ariat/Carhartt rows -> PMtoARIAT.py

        Implementation detail:
        We temporarily overwrite Processed_orders.xlsx with a filtered subset for each vendor script,
        then restore the full workbook afterwards.
        """
        try:
            from openpyxl import load_workbook, Workbook
        except Exception as e:
            self.log(f"[ERROR] openpyxl not installed: {e}")
            if on_complete:
                on_complete(1)
            return None

        # Prevent overlapping interactive runs
        if getattr(self.running, "interactive", None) is not None and self.running.interactive.is_running():
            self.log("[INFO] A vendor script is already running. Finish it or Kill Current Task ⛔.")
            if on_complete:
                on_complete(1)
            return None

        p = self.paths()
        processed_path = p["processed"]
        ensure_processed_orders_closed_path(processed_path, self.log)

        if not os.path.isfile(processed_path):
            self.log(f"[ERROR] Missing Processed Orders.xlsx at: {processed_path}")
            if on_complete:
                on_complete(1)
            return None

        wb = load_workbook(processed_path, data_only=True)
        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column

        header = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]

        VENDOR_COL = 11  # Column K
        if max_col < VENDOR_COL:
            self.log("[ERROR] Processed_orders.xlsx does not have Column K (vendor).")
            if on_complete:
                on_complete(1)
            return None

        def classify_vendor(raw: str) -> str:
            s = (raw or "").strip().lower()
            primary = re.split(r"[,/;|]+", s)[0].strip() if s else ""
            targets = [primary, s]

            def is_wrg(x):
                return ("wrangler" in x) or (x == "wrg") or (" wrg" in x) or ("wrg " in x)

            def is_ariat(x):
                return ("ariat" in x) or ("carhartt" in x)

            for x in targets:
                if is_wrg(x):
                    return "wrangler"
                if is_ariat(x):
                    return "ariat"
            return ""

        wrg_rows = []
        ariat_rows = []

        for r in range(2, max_row + 1):
            vend_val = ws.cell(row=r, column=VENDOR_COL).value
            vend = classify_vendor(str(vend_val) if vend_val is not None else "")
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if vend == "wrangler":
                wrg_rows.append(row_vals)
            elif vend == "ariat":
                ariat_rows.append(row_vals)

        if not wrg_rows and not ariat_rows:
            self.log("[INFO] No Wrangler or Ariat rows found in Column K. Nothing to run.")
            if on_complete:
                on_complete(0)
            return None

        steps = []
        if wrg_rows:
            steps.append(("Wrangler", "PMtoWRG.py", wrg_rows))
        if ariat_rows:
            steps.append(("Ariat", "PMtoARIAT.py", ariat_rows))

        backup_path = processed_path + ".FULL.bak"
        try:
            import shutil
            shutil.copy2(processed_path, backup_path)
        except Exception as e:
            self.log(f"[ERROR] Could not backup Processed_orders.xlsx: {e}")
            if on_complete:
                on_complete(1)
            return None

        def write_subset(rows):
            sub = Workbook()
            sws = sub.active
            sws.title = ws.title if ws.title else "Sheet1"
            for c, val in enumerate(header, start=1):
                sws.cell(row=1, column=c, value=val)
            for rr, row_vals in enumerate(rows, start=2):
                for c, val in enumerate(row_vals, start=1):
                    sws.cell(row=rr, column=c, value=val)
            sub.save(processed_path)

        def restore_full():
            try:
                import shutil
                shutil.copy2(backup_path, processed_path)
                os.remove(backup_path)
            except Exception as e:
                self.log(f"[WARN] Could not restore full Processed_orders.xlsx automatically: {e}")

        def run_step(i: int):
            if i >= len(steps):
                restore_full()
                self.btn_verify.setEnabled(False)
                self.running.interactive = None
                self.log("◆ Place Orders with Vendor complete.")
                if on_complete:
                    on_complete(0)
                return

            vendor_name, script_name, rows = steps[i]
            self.log(f"▶ Starting {vendor_name}… ({len(rows)} row(s))")

            try:
                write_subset(rows)
            except Exception as e:
                self.log(f"[ERROR] Failed writing vendor subset for {vendor_name}: {e}")
                restore_full()
                if on_complete:
                    on_complete(1)
                return

            worker = self.run_script(script_name, stdin_pipe=True, label=f"{script_name} ({vendor_name})")
            if worker is None:
                self.log(f"[ERROR] {script_name} did not start. Skipping {vendor_name}.")
                run_step(i + 1)
                return

            self.running.interactive = worker
            self.btn_verify.setEnabled(True)

            if script_name == "PMtoWRG.py":
                self.running.pm_to_wrg = worker
            elif script_name == "PMtoARIAT.py":
                self.running.pm_to_ariat = worker

            def _after(rc):
                self.running.interactive = None
                self.btn_verify.setEnabled(False)
                self.log(f"◆ {vendor_name} exited with code {rc}")
                run_step(i + 1)

            worker.finished.connect(_after)

        run_step(0)
        return None


    def run_get_order_ids(self):
        """Run GetOrderId.py with stdin piping for CAPTCHA verification."""
        ensure_processed_orders_closed_path(self.paths()["processed"], self.log)
        self.log("▶ Starting Get Order IDs...")
        
        worker = self.run_script("GetOrderId.py", stdin_pipe=True, label="GetOrderId.py")
        if worker is None:
            self.log("[ERROR] GetOrderId.py did not start.")
            return
        
        # Track as interactive worker and enable verification button
        self.running.interactive = worker
        self.btn_verify.setEnabled(True)
        
        def _after(rc):
            self.running.interactive = None
            self.btn_verify.setEnabled(False)
            if rc == 0:
                self.log("◆ Get Order IDs completed successfully")
            else:
                self.log(f"◆ Get Order IDs exited with code {rc}")
        
        worker.finished.connect(_after)


    def verification_complete(self):
        # Sends Enter to whichever script is currently running & waiting (GetOrderId, Wrangler, or Ariat).
        w = getattr(self.running, "interactive", None)

        if w is None or not w.is_running():
            w = getattr(self.running, "pm_to_wrg", None)
        if w is None or not w.is_running():
            w = getattr(self.running, "pm_to_ariat", None)

        if w is None or not w.is_running():
            self.log("[INFO] No interactive script is currently waiting.")
            return

        w.send_enter()
        self.log('✅ Sent Enter to the running script ("Verification Complete").')

        # Keep enabled; many scripts pause multiple times per run.

    # --- Back-Orders + PM numbers pipeline ---
    def run_backorders_then_pm(self):
        if getattr(self, 'pipeline_active', False):
            self.log("[INFO] Another pipeline is already in progress.")
            return
        self.pipeline_active = True
        ensure_processed_orders_closed_path(self.paths()["processed"], self.log)
        self.log("▶ Place Back-Orders starting…")

        def halt(step_name: str, rc: int):
            self.pipeline_active = False
            self.log(f"[ERROR] {step_name} failed with exit code {rc}.")
            self.log("⏸ Pipeline halted.")

        def end():
            self.pipeline_active = False
            self.log("◆ Place Back-Orders + PM numbers complete.")

        def after_backorders():
            # Show the same PM numbers dialog as the manual pipeline
            nums = self.prompt_pm_numbers()
            if not nums:
                self.log("[INFO] No PM numbers entered — skipping Add_PM_Nums step.")
                end()
                return
            if not self.write_pm_numbers_excel(nums):
                end()
                return
            self.log("▶︎ Running Add_PM_Nums.py…")
            w = self.run_script("Add_PM_Nums.py", label="Add_PM_Nums.py")
            if w:
                w.finished.connect(lambda rc: end() if rc == 0 else halt("Add_PM_Nums.py", rc))
            else:
                halt("Add_PM_Nums.py", 1)

        def start_backorders():
            w = self.run_script("BroberryShop_Backorders.py", label="BroberryShop_Backorders.py")
            if w:
                w.finished.connect(lambda rc: after_backorders() if rc == 0 else halt("BroberryShop_Backorders.py", rc))
            else:
                halt("BroberryShop_Backorders.py", 1)

        start_backorders()

    # --- Run All Steps (pipeline) ---

    def run_all_steps(self):
        if getattr(self, 'pipeline_active', False):
            self.log("[INFO] Run All Steps is already in progress.")
            return
        self.pipeline_active = True
        ensure_processed_orders_closed_path(self.paths()["processed"], self.log)
        self.log("▶ Run All Steps starting…")

        def halt_pipeline(step_name: str, rc: int):
            self.pipeline_active = False
            self.log(f"[ERROR] {step_name} failed with exit code {rc}.")
            self.log("⏸ Pausing pipeline; no further steps will run.")

        def end_pipeline():
            self.pipeline_active = False
            self.log("◆ Run All Steps complete.")

        def start_extract():
            w = self.run_script("PDFExtract.py")
            if w:
                w.finished.connect(lambda rc: start_broberry() if rc == 0 else halt_pipeline("PDFExtract.py", rc))
            else:
                halt_pipeline("PDFExtract.py", 1)

        def start_broberry():
            w = self.run_script("BroberryShop.py")
            if w:
                w.finished.connect(lambda rc: start_download() if rc == 0 else halt_pipeline("BroberryShop.py", rc))
            else:
                halt_pipeline("BroberryShop.py", 1)

        def start_download():
            w = self.run_script("ShoptoPM.py")
            if w:
                w.finished.connect(lambda rc: start_pm() if rc == 0 else halt_pipeline("ShoptoPM.py", rc))
            else:
                halt_pipeline("ShoptoPM.py", 1)

        def start_pm():
            self.run_pad_flow_sequence(on_complete=lambda rc: start_wrg() if rc == 0 else halt_pipeline("Place in PM", rc))

        def start_wrg():
            self.log("▶ Starting vendor order placement…")
            self.run_orders_with_vendor(on_complete=lambda rc: end_pipeline())

        start_extract()

    # --- Manual PM numbers flow (Run to PM) ---
    def prompt_pm_numbers(self):
        dlg = PMNumberEntryDialog(self)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            nums = dlg.values()
            if not nums:
                self.log("[INFO] No PM numbers entered.")
            return nums
        else:
            self.log("[INFO] PM number entry cancelled.")
            return []

    def write_pm_numbers_excel(self, numbers):
        try:
            from openpyxl import Workbook
        except Exception as e:
            self.log(f"[ERROR] openpyxl not installed: {e}")
            return False
        try:
            p = self.paths()
            os.makedirs(os.path.dirname(p["pmnum"]), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            for i, val in enumerate(numbers, start=1):
                ws.cell(row=i, column=1, value=val)
            wb.save(p["pmnum"])
            self.log(f"💾 Wrote {len(numbers)} PM number(s) to {p['pmnum']}")
            return True
        except Exception as e:
            self.log(f"[ERROR] Failed writing PMNum.xlsx: {e}")
            return False

    def run_to_pm_manual(self):
        if getattr(self, 'pipeline_active', False):
            self.log("[INFO] Another pipeline is in progress.")
            return
        self.pipeline_active = True
        ensure_processed_orders_closed_path(self.paths()["processed"], self.log)
        self.log("▶ Run to PM (manual) starting…")

        def halt_pipeline(step_name: str, rc: int):
            self.pipeline_active = False
            self.log(f"[ERROR] {step_name} failed with exit code {rc}.")
            self.log("⏸ Pausing pipeline; no further steps will run.")

        def end_pipeline():
            self.pipeline_active = False
            self.log("◆ Run to PM (manual) complete.")

        def start_extract():
            w = self.run_script("PDFExtract.py")
            if w:
                w.finished.connect(lambda rc: start_broberry() if rc == 0 else halt_pipeline("PDFExtract.py", rc))
            else:
                halt_pipeline("PDFExtract.py", 1)

        def start_broberry():
            w = self.run_script("BroberryShop.py")
            if w:
                w.finished.connect(lambda rc: start_download() if rc == 0 else halt_pipeline("BroberryShop.py", rc))
            else:
                halt_pipeline("BroberryShop.py", 1)

        def start_download():
            w = self.run_script("ShoptoPM.py")
            if w:
                w.finished.connect(lambda rc: after_download() if rc == 0 else halt_pipeline("ShoptoPM.py", rc))
            else:
                halt_pipeline("ShoptoPM.py", 1)

        def after_download():
            nums = self.prompt_pm_numbers()
            if not nums:
                self.log("[INFO] No PM numbers entered; aborting manual merge.")
                end_pipeline(); return
            if not self.write_pm_numbers_excel(nums):
                end_pipeline(); return
            self.log("▶︎ Running Add_PM_Nums.py…")
            w = self.run_script("Add_PM_Nums.py", label='Add_PM_Nums.py')
            if w:
                w.finished.connect(lambda rc: end_pipeline() if rc == 0 else halt_pipeline("Add_PM_Nums.py", rc))
            else:
                halt_pipeline("Add_PM_Nums.py", 1)

        start_extract()

    # --- Clear PDF folder ---
    def clear_pdf_folder(self):
        purge_pdfs_and_csvs(self.paths()["pdfs"], self.log)

    # --- Graceful shutdown ---
    def closeEvent(self, event):
        try:
            if self.running.pm_to_wrg and self.running.pm_to_wrg.proc:
                self.running.pm_to_wrg.proc.terminate()
        except Exception:
            pass
        return super().closeEvent(event)

# --- Entry point --------------------------------------------------------------
def main():
    app = QtWidgets.QApplication(sys.argv)
    app.setStyleSheet(DARK_STYLESHEET)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()
