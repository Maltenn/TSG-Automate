import os
import re
import sys
import time
import datetime
import pandas as pd

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

# ─── CONFIG ────────────────────────────────────────────────────────────────────
SCRIPT_DIR        = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH        = os.path.join(SCRIPT_DIR, 'Processed_orders.xlsx')

LOGIN_URL            = "https://wranglerb2b.com/login.php/client/NQ=="
ORDER_HISTORY_URL    = "https://wranglerb2b.com/tp_order_history.php/ecat_view"

EMAIL     = os.getenv("WRANGLER_EMAIL") or os.getenv("WRG_EMAIL") or "internal3@broberry.com"
PASSWORD  = os.getenv("WRANGLER_PASSWORD") or os.getenv("WRG_PASSWORD") or "Internal3Broberry!"
# ────────────────────────────────────────────────────────────────────────────────

def login(driver):
    """Log into Wrangler B2B with CAPTCHA pause."""
    driver.get(LOGIN_URL)
    wait = WebDriverWait(driver, 20)

    # Fill credentials
    email_el = wait.until(EC.visibility_of_element_located((By.ID, "login_email")))
    email_el.clear()
    email_el.send_keys(EMAIL)

    pwd_el = driver.find_element(By.ID, "login_password")
    pwd_el.clear()
    pwd_el.send_keys(PASSWORD)

    # Pause for manual CAPTCHA completion before submitting
    print("\n[ACTION REQUIRED]", flush=True)
    print("Please complete any login verification in the browser window now.", flush=True)
    print("Examples: CAPTCHA checkbox, 'I'm not a robot', or any pre-login security step.", flush=True)
    print("When you're done, click 'Verification Complete ✅' button in the app...", flush=True)
    sys.stdout.flush()
    
    # Wait for Enter from app (sent when user clicks "Verification Complete")
    sys.stdin.readline()

    # Submit login after user confirms
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

    # Wait for successful login
    wait.until(EC.presence_of_element_located((By.ID, "p7SOPt_2")))

def find_order_id_for_po(driver, po):
    """
    On the ORDER_HISTORY_URL page, tries in turn to:
      1) scroll down
      2) refresh
      3) click "Next" page
    Returns the Order ID string or None.
    """
    attempts = 0
    while attempts < 4:
        try:
            xpath_li = (
                f"//li[contains(@class,'TD-row') "
                f" and .//span[em[text()='PO#:'] and contains(., '{po}')]]"
            )
            block = driver.find_element(By.XPATH, xpath_li)
            sid = block.find_element(By.XPATH, ".//span[em[text()='Order ID:']]").text
            return sid.replace("Order ID:", "").strip()
        except NoSuchElementException:
            if attempts == 0:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
            elif attempts == 1:
                driver.refresh()
                time.sleep(2)
            elif attempts == 2:
                try:
                    nxt = driver.find_element(By.ID, "next-link")
                    nxt.click()
                    time.sleep(2)
                except NoSuchElementException:
                    break
        attempts += 1
    return None

def main():
    print("")
    print("="*60)
    print("*** GET ORDER IDs FROM WRANGLER B2B ***")
    print("="*60)
    print(f"Script started at: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Excel file: {EXCEL_PATH}")
    print("="*60)
    print("")
    
    # Load the Excel FIRST so we can skip the Wrangler login entirely when
    # this batch contains no Wrangler orders (e.g. Ariat/Propper-only runs).
    print("[INFO] Loading Excel file...")
    df = pd.read_excel(EXCEL_PATH, engine="openpyxl", dtype=str)
    col_g = df.columns[6]   # PO column
    col_k = df.columns[10]  # Vendor column

    wrangler_rows = []
    for idx, po in enumerate(df[col_g]):
        if not po or pd.isna(po):
            continue
        vendor_cell = str(df.at[idx, col_k]) if not pd.isna(df.at[idx, col_k]) else ""
        if "wrangler" in vendor_cell.lower():
            wrangler_rows.append((idx, po))
        else:
            print(f"[{idx+1}/{len(df)}] Skipping PO {po} — vendor '{vendor_cell}' does not include Wrangler")

    if not wrangler_rows:
        print("")
        print("[INFO] No Wrangler orders found in Processed_orders.xlsx (column K).")
        print("[INFO] Nothing to fetch — skipping Wrangler login.")
        return

    driver = webdriver.Chrome()
    try:
        print("[INFO] Logging in to Wrangler B2B...")
        login(driver)
        print("[OK] Login successful!")
        print("")

        # Go to order history once
        print("[INFO] Navigating to Order History...")
        driver.get(ORDER_HISTORY_URL)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "li.TD-row"))
        )
        print("[OK] Order History page loaded")
        print("")

        print("="*60)
        print(f"[INFO] Fetching Order IDs for {len(wrangler_rows)} Wrangler orders")
        print("="*60)
        print("")

        # Loop through the Wrangler POs
        success_count = 0
        fail_count = 0
        found = {}   # df row index -> order id

        for idx, po in wrangler_rows:
            print(f"[{idx+1}/{len(df)}] Looking up Order ID for PO: {po}")
            oid = find_order_id_for_po(driver, po)

            if oid:
                found[idx] = oid
                success_count += 1
                print(f"[OK] PO {po} → Order ID: {oid}")
            else:
                fail_count += 1
                print(f"[WARNING] Order ID not found for PO {po}")
            print("")

        # Save results with targeted openpyxl writes so the rest of the
        # workbook (values, fonts, date formatting) is left untouched.
        print("="*60)
        print("[INFO] Saving results to Excel...")
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        if "Order ID" in headers:
            oid_col = headers.index("Order ID") + 1
        else:
            oid_col = 13  # column M ("Order Number in vendor system")
        for idx, oid in found.items():
            ws.cell(row=idx + 2, column=oid_col).value = oid  # +2: header row + 1-based
        wb.save(EXCEL_PATH)
        print(f"[OK] Saved updated Order IDs to {EXCEL_PATH}")
        print("="*60)
        print("")
        
        # Final summary
        print("="*60)
        print("*** ORDER ID FETCH COMPLETE ***")
        print("="*60)
        print(f"Successfully found: {success_count} Order IDs")
        if fail_count > 0:
            print(f"Failed to find: {fail_count} Order IDs")
        print(f"Excel File Updated: {EXCEL_PATH}")
        print("="*60)
    
    except Exception as e:
        print("")
        print("="*60)
        print("[ERROR] Script encountered an error!")
        print(f"[ERROR] {str(e)}")
        print("="*60)
        raise
    finally:
        print("[INFO] Closing browser...")
        driver.quit()
        print("[INFO] Browser closed.")

if __name__ == "__main__":
    main()
