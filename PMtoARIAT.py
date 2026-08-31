import os
import re
import sys
import glob
import time
import math
import datetime
import traceback
import pandas as pd
from openpyxl import load_workbook

import tsg_runlog

# Make stdout/stderr robust to non-UTF-8 consoles (e.g. Windows cp1252 when this
# script is launched standalone or with its output redirected).  Without this a
# print() containing a non-cp1252 character raised UnicodeEncodeError and killed
# the whole run.  The GUI app already forces PYTHONUTF8=1, but standalone runs
# (double-click, console, redirected logs) did not — making the script appear to
# "fail randomly" depending on how it was launched.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    ElementClickInterceptedException,
    ElementNotInteractableException,
)
from selenium.webdriver.common.action_chains import ActionChains


# ─── CONFIG ────────────────────────────────────────────────────────────────────
SCRIPT_DIR      = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH      = os.path.join(SCRIPT_DIR, "Processed_orders.xlsx")

DOWNLOAD_FOLDER = os.getenv("TSG_DOWNLOAD_DIR", os.path.join(os.path.expanduser("~"), "Downloads"))
PDF_DIR = os.getenv("TSG_PDF_DIR", os.path.join(SCRIPT_DIR, "pdfs"))

ARIAT_URL       = "https://b2b.ariat.com/"

ARIAT_USERNAME  = os.getenv("ARIAT_USERNAME", "internal3")
ARIAT_PASSWORD  = os.getenv("ARIAT_PASSWORD", "5Wft87ptvX68h3h")

WAIT_LONG  = 90
WAIT_MED   = 30
WAIT_SHORT = 10

# Debug mode: set TSG_DEBUG=1 to (a) expose Chrome DevTools on port 9222 so an
# external tool can attach and inspect the live page, (b) dump a screenshot +
# HTML on any failure, and (c) HOLD the browser open on errors instead of
# crashing out, so the failing page state can be examined.
TSG_DEBUG = os.getenv("TSG_DEBUG", "").strip().lower() not in ("", "0", "false", "no")
DEBUG_PORT = 9222
# ────────────────────────────────────────────────────────────────────────────────


def debug_dump(driver, error_name="error"):
    """Save a screenshot + page HTML + URL so a failure can be diagnosed later."""
    try:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        png = os.path.join(SCRIPT_DIR, f"debug_ariat_{error_name}_{ts}.png")
        html = os.path.join(SCRIPT_DIR, f"debug_ariat_{error_name}_{ts}.html")
        driver.save_screenshot(png)
        print(f"[DEBUG] Screenshot saved: {png}")
        with open(html, "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        print(f"[DEBUG] HTML saved: {html}")
        print(f"[DEBUG] URL at failure: {driver.current_url}")
    except Exception as e:
        print(f"[DEBUG] Failed to save debug info: {e}")


def debug_hold(driver, context=""):
    """In TSG_DEBUG mode, keep the browser open and wait for instructions on stdin.

    Returns 'continue' (move on to the next order) or 'abort' (end the run).
    In normal mode returns 'abort' immediately, preserving fail-fast behaviour —
    continuing past a half-finished order risks a dirty cart contaminating the
    next order, so that decision is only offered to a human in debug mode.
    """
    if not TSG_DEBUG:
        return "abort"
    print("")
    print(f"[DEBUG_HOLD] {context}")
    print(f"[DEBUG_HOLD] Browser held open (DevTools on 127.0.0.1:{DEBUG_PORT}).")
    print("[DEBUG_HOLD] Type 'continue' to move to the next order, or 'abort' to end the run.")
    while True:
        try:
            resp = input().strip().lower()
        except EOFError:
            return "abort"
        if resp == "continue":
            print("[DEBUG_HOLD] Continuing with the next order.")
            return "continue"
        if resp == "abort":
            print("[DEBUG_HOLD] Aborting the run.")
            return "abort"
        if resp:
            print(f"[DEBUG_HOLD] Unrecognized input '{resp}' — type 'continue' or 'abort'.")


def _find_displayed(driver, by, sel, timeout=WAIT_MED, context=""):
    """Return the first VISIBLE element matching sel.

    Unlike EC.visibility_of_element_located, this scans ALL matches each poll.
    Dojo keeps the previous order's dialogs/buttons hidden in the DOM, so after
    the first order a first-match lookup often lands on a stale hidden node and
    times out (or worse, reads stale text) even though a fresh visible instance
    exists further down the DOM.
    """
    end = time.time() + timeout
    while time.time() < end:
        for e in driver.find_elements(by, sel):
            try:
                if e.is_displayed():
                    return e
            except Exception:
                continue
        time.sleep(0.25)
    raise TimeoutException(f"No VISIBLE element matching {sel!r} within {timeout}s. {context}")


def _any_displayed(driver, by, sel) -> bool:
    """True if at least one element matching sel is currently visible."""
    try:
        return any(e.is_displayed() for e in driver.find_elements(by, sel))
    except Exception:
        return False


def _wait_none_displayed(driver, by, sel, timeout=WAIT_MED) -> bool:
    """Wait until NO element matching sel is visible. True on success."""
    end = time.time() + timeout
    while time.time() < end:
        if not _any_displayed(driver, by, sel):
            return True
        time.sleep(0.25)
    return False
# ────────────────────────────────────────────────────────────────────────────────


# ─── CART VERIFICATION & SELF-HEALING (added 2026-08-31) ──────────────────────
# Root cause of the doubled-quantity orders: Ariat persists the working draft
# server-side.  A crashed run leaves its items in the draft; the next login
# resumes that SAME draft and the fresh import stacks on top (1 leftover unit +
# 1 imported unit = order placed with 2).  These helpers make every order
# (a) start from a provably empty cart and (b) refuse to submit when the cart
# quantity does not exactly match the upload file.

def expected_units_from_upload(path: str) -> int:
    """Sum the qty column of an Ariat/Carhartt upload workbook (upc, qty)."""
    df = pd.read_excel(path, dtype=str)
    qty_col = None
    for c in df.columns:
        if str(c).strip().lower() in ("qty", "quantity", "units"):
            qty_col = c
            break
    if qty_col is None:
        qty_col = df.columns[1]  # historical layout: A=upc, B=qty
    total = 0
    for v in df[qty_col]:
        s = coerce_str(v)
        if s.isdigit():
            total += int(s)
    return total


def ariat_cart_units(driver) -> int:
    """Read the unit count from the header cart badge ('N Units / $..')."""
    try:
        txt = driver.execute_script("return document.body.innerText") or ""
    except Exception:
        return -1
    m = re.search(r"(\d+)\s*Units\s*/", txt)
    return int(m.group(1)) if m else -1


def open_main_menu_v2(driver, timeout=WAIT_MED):
    """Open the builder Menu via Dojo's own dropdown API.

    The old click-strategy ladder still exists as a fallback, but the widget
    API is deterministic — and immune to the stuck popupactive='true'
    attribute that makes is_main_menu_open() lie after the menu has been used
    once in the same shell.
    """
    opened = driver.execute_script("""
        var t=[...document.querySelectorAll('div.mainMenu')].filter(e=>e.offsetParent)[0];
        if(!t || !window.dijit || !dijit.registry) return false;
        var w = dijit.registry.getEnclosingWidget(t);
        var probe = w, hops = 0;
        while (probe && hops < 4) {
            if (probe.openDropDown) { probe.openDropDown(); return true; }
            probe = dijit.registry.getEnclosingWidget(probe.domNode.parentNode);
            hops++;
        }
        return false;
    """)
    if opened:
        end = time.time() + 5
        while time.time() < end:
            if _any_displayed(driver, By.CSS_SELECTOR, "td.dijitMenuItemLabel"):
                return
            time.sleep(0.2)
    # Fallback: the historical strategy ladder
    open_main_menu(driver, timeout=timeout)


def click_menu_item(driver, label: str, timeout=WAIT_MED):
    """Click a VISIBLE main-menu item by its label text."""
    open_main_menu_v2(driver, timeout=timeout)
    item = _find_displayed(
        driver, By.XPATH,
        f"//td[contains(@class,'dijitMenuItemLabel') and normalize-space()='{label}']",
        timeout=timeout, context=f"(menu item '{label}')")
    safe_click(driver, item)


def ensure_fresh_ariat_order(driver):
    """Guarantee the builder holds an EMPTY cart before an import.

    If the resumed draft has leftover units (crashed previous run), discard it:
    Menu -> New Order -> 'Don't Save' (drops to the marketing page) -> re-enter
    the builder -> confirm 0 units.  Verified live 2026-08-31.
    """
    # Stability poll: right after the shell loads the badge can briefly read 0
    # before the resumed draft's real units render — trust a value only once
    # two consecutive 1s-apart reads agree (max ~6s, usually ~2s).
    units = ariat_cart_units(driver)
    end = time.time() + 6
    while time.time() < end:
        time.sleep(1.0)
        cur = ariat_cart_units(driver)
        if cur == units and cur >= 0:
            break
        units = cur
    if units == 0:
        print("[INFO] Cart check: builder cart is empty — OK to import.")
        return
    print(f"[CART_GUARD] Builder cart has {units} leftover unit(s) from a previous run!")
    print("[CART_GUARD] Discarding the dirty draft via Menu -> New Order -> Don't Save...")
    debug_dump(driver, "leftover_cart_before_discard")

    click_menu_item(driver, "New Order", timeout=WAIT_MED)
    # Confirm dialog: "Your order has not been saved..." -> Don't Save
    try:
        btn = _find_displayed(
            driver, By.XPATH,
            "//button[normalize-space()=\"Don't Save\"] | "
            "//span[contains(@class,'dijitButtonText') and normalize-space()=\"Don't Save\"]/ancestor::*[@role='button'][1]",
            timeout=WAIT_SHORT, context="(New Order confirm dialog)")
        safe_click(driver, btn)
        print("[CART_GUARD] Clicked \"Don't Save\" — dirty draft discarded.")
    except TimeoutException:
        print("[CART_GUARD] No confirm dialog appeared (draft may have been clean-saved).")

    time.sleep(2.0)
    enter_order_builder(driver, timeout=WAIT_LONG)
    wait_ready(driver)
    time.sleep(1.0)

    units = ariat_cart_units(driver)
    if units != 0:
        debug_dump(driver, "cart_not_empty_after_discard")
        raise RuntimeError(
            f"Cart still shows {units} unit(s) after discarding the draft — refusing to import."
        )
    print("[CART_GUARD] Fresh order confirmed: cart is empty.")


def verify_ariat_cart(driver, expected_units: int, context: str, settle_timeout: int = 30):
    """Hard gate: header-badge units must EXACTLY match the upload file.

    Polls up to settle_timeout — the badge updates a few seconds AFTER the
    import dialog closes (server round-trip), so a single immediate read can
    race and report 0 (observed live 2026-08-31)."""
    end = time.time() + settle_timeout
    units = -1
    while time.time() < end:
        units = ariat_cart_units(driver)
        if units == expected_units:
            print(f"[CART_VERIFY] OK ({context}): cart has {units} unit(s), expected {expected_units}.")
            return
        time.sleep(1.0)
    debug_dump(driver, f"cart_mismatch_{context.replace(' ', '_')}")
    raise RuntimeError(
        f"CART MISMATCH ({context}): cart shows {units} unit(s) but the upload file "
        f"contains {expected_units} (waited {settle_timeout}s). A leftover/dirty cart "
        "would double-order — NOT proceeding."
    )
# ────────────────────────────────────────────────────────────────────────────────


US_STATE_ABBR_TO_NAME = {
    "AL":"Alabama","AK":"Alaska","AZ":"Arizona","AR":"Arkansas","CA":"California","CO":"Colorado","CT":"Connecticut",
    "DE":"Delaware","FL":"Florida","GA":"Georgia","HI":"Hawaii","ID":"Idaho","IL":"Illinois","IN":"Indiana",
    "IA":"Iowa","KS":"Kansas","KY":"Kentucky","LA":"Louisiana","ME":"Maine","MD":"Maryland","MA":"Massachusetts",
    "MI":"Michigan","MN":"Minnesota","MS":"Mississippi","MO":"Missouri","MT":"Montana","NE":"Nebraska","NV":"Nevada",
    "NH":"New Hampshire","NJ":"New Jersey","NM":"New Mexico","NY":"New York","NC":"North Carolina","ND":"North Dakota",
    "OH":"Ohio","OK":"Oklahoma","OR":"Oregon","PA":"Pennsylvania","RI":"Rhode Island","SC":"South Carolina",
    "SD":"South Dakota","TN":"Tennessee","TX":"Texas","UT":"Utah","VT":"Vermont","VA":"Virginia","WA":"Washington",
    "WV":"West Virginia","WI":"Wisconsin","WY":"Wyoming","DC":"District of Columbia",
}


def _short_err(e) -> str:
    """First line of an exception message — WebDriver errors append a full
    chromedriver stacktrace to str(e), which drowned the logs on every
    expected-failure click attempt."""
    s = str(e).strip()
    return s.splitlines()[0] if s else type(e).__name__


def coerce_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if math.isnan(val):
            return ""
        if val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()


def cleanup_po_files(client_po: str, csv_path: str = "") -> None:
    """Delete the CSV and PDF in PDF_DIR that belong to a just-placed order.

    CSV: uses csv_path directly if provided, otherwise falls back to
         <PDF_DIR>/<client_po>.csv
    PDF pattern: <PDF_DIR>/PO_Form_Group<client_po>*.pdf
    """
    deleted, skipped = [], []

    # CSV — prefer the exact path we already know
    csv_target = csv_path if csv_path and os.path.exists(csv_path) \
                 else os.path.join(PDF_DIR, f"{client_po}.csv")
    if os.path.exists(csv_target):
        try:
            os.remove(csv_target)
            deleted.append(csv_target)
            print(f"[INFO] Deleted CSV: {csv_target}")
        except Exception as e:
            print(f"[WARN] Could not delete CSV '{csv_target}': {e}")
            skipped.append(csv_target)
    else:
        print(f"[INFO] CSV not found (already removed?): {csv_target}")

    # PDF (vendor name varies, so use a wildcard)
    pdf_matches = glob.glob(os.path.join(PDF_DIR, f"PO_Form_Group{client_po}*.pdf"))
    if not pdf_matches:
        pdf_matches = glob.glob(os.path.join(PDF_DIR, f"PO_Form_Group{client_po}*.PDF"))
    if pdf_matches:
        for pdf_path in pdf_matches:
            try:
                os.remove(pdf_path)
                deleted.append(pdf_path)
                print(f"[INFO] Deleted PDF: {pdf_path}")
            except Exception as e:
                print(f"[WARN] Could not delete PDF '{pdf_path}': {e}")
                skipped.append(pdf_path)
    else:
        print(f"[INFO] PDF not found for client PO {client_po} (already removed?)")

    print(f"[INFO] Cleanup for PO {client_po}: {len(deleted)} deleted, {len(skipped)} skipped.")


def extract_po_key(po_number: str) -> str:
    """
    For values like '162945-297239' return '297239' (last digit chunk).
    If the string is already '297239', returns '297239'.
    """
    s = coerce_str(po_number)
    chunks = re.findall(r"\d+", s)
    return chunks[-1] if chunks else s


def wait_ready(driver, timeout=WAIT_MED):
    WebDriverWait(driver, timeout).until(lambda d: d.execute_script("return document.readyState") == "complete")


def safe_click(driver, el):
    try:
        el.click()
    except (ElementClickInterceptedException, ElementNotInteractableException):
        driver.execute_script("arguments[0].click();", el)


def wait_and_click(driver, by, sel, timeout=WAIT_MED):
    el = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, sel)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    safe_click(driver, el)
    return el


def wait_visible(driver, by, sel, timeout=WAIT_MED):
    return WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((by, sel)))


def wait_present(driver, by, sel, timeout=WAIT_MED):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, sel)))


def click_button_by_text(driver, text, timeout=WAIT_MED):
    xpath = f"//button[normalize-space()='{text}' or .//div[normalize-space()='{text}'] or contains(normalize-space(.), '{text}')]"
    return wait_and_click(driver, By.XPATH, xpath, timeout=timeout)


# ─── DIJIT BUTTON HELPERS (for Save) ───────────────────────────────────────────
def click_dijit_button_by_label(driver, label_text: str, timeout=WAIT_LONG, prefer_id: str | None = None):
    """
    Clicks a Dojo/Dijit button reliably.
    - If prefer_id is provided and exists (e.g. 'dijit_form_Button_40'), click that first.
    - Otherwise find the dijitButtonText span by label and click its clickable ancestor.
    """
    label_text = str(label_text).strip()

    def _visible(el):
        try:
            return el.is_displayed()
        except Exception:
            return False

    def _click_btn_for_label_span(span):
        try:
            btn = span.find_element(By.XPATH, "./ancestor::*[@role='button'][1]")
        except Exception:
            btn = span.find_element(By.XPATH, "./ancestor::*[contains(@class,'dijitButtonNode')][1]")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        safe_click(driver, btn)
        return btn

    # 1) Prefer the stable ID — but ONLY when it is actually visible.  Dojo
    #    regenerates the numeric widget-id suffix and often keeps a hidden
    #    duplicate (e.g. a collapsed form's Save), so a hardcoded id like
    #    'dijit_form_Button_40' frequently points at a hidden/stale button.
    if prefer_id:
        try:
            el = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, prefer_id)))
            if _visible(el):
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                safe_click(driver, el)
                return el
        except Exception:
            pass

    # 2) Find by label, preferring the VISIBLE button instance.
    xpath_label = f"//span[contains(@class,'dijitButtonText') and normalize-space()='{label_text}']"
    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath_label)))
    spans = driver.find_elements(By.XPATH, xpath_label)
    spans = sorted(spans, key=lambda s: 0 if _visible(s) else 1)  # visible first

    last_err = None
    for span in spans:
        try:
            return _click_btn_for_label_span(span)
        except Exception as e:
            last_err = e
            continue
    if last_err:
        raise last_err
    raise TimeoutException(f"Could not click a dijit button labelled '{label_text}'.")


# ─── DOJO MAIN MENU (CRITICAL) ──────────────────────────────────────────────
MAIN_MENU_TRIGGER_ID = "dijit__WidgetsInTemplateMixin_1"
# NOTE: The popup container ID is read dynamically from the trigger's
# aria-owns attribute (typically 'dijit_DropDownMenu_<n>').  Dojo regenerates
# the numeric suffix on re-render, so a hardcoded constant becomes wrong
# over time — that mismatch was the actual cause of the long-standing
# "struggles to select the Import a File menu" bug: the trigger was firing
# correctly, but is_main_menu_open() looked up a non-existent ID and always
# returned False, so open_main_menu()'s retry loop kept JS-clicking the
# trigger and toggling the menu closed again.
# Same lesson as the menu items: do NOT pin to dijit_*_NN IDs.
IMPORT_MENU_LABEL_TD_ID = None  # deprecated; kept for reference only


def _resolve_main_menu_popup_id(driver):
    """Read the live popup container ID from the trigger's aria-owns.

    Returns the ID string, or None if the trigger isn't on the page yet.
    """
    try:
        trig = driver.find_element(By.ID, MAIN_MENU_TRIGGER_ID)
        owns = trig.get_attribute("aria-owns") or trig.get_attribute("aria-controls")
        return owns.strip() if owns else None
    except Exception:
        return None


def is_main_menu_open(driver) -> bool:
    """Truthful detection of the Dojo main menu's open state.

    Three independent signals — ANY ONE is sufficient:
      1. trigger element carries popupactive='true'  (authoritative, set by Dojo)
      2. the popup container referenced by aria-owns is visible
      3. a known menu row (tr.import_csv) is visible in the DOM

    Multiple signals exist because Dojo's ARIA bookkeeping isn't always
    consistent (e.g. popupactive='true' alongside aria-expanded='false').
    """
    # 1) any visible menu-item label — the only signal that cannot lie.
    #    (The old first signal trusted the trigger's popupactive attribute,
    #    but Dojo leaves it stuck on 'true' after the menu has been used once,
    #    so open_main_menu() no-opped on every later call in the same shell.)
    try:
        if any(e.is_displayed() for e in driver.find_elements(By.CSS_SELECTOR, "td.dijitMenuItemLabel")):
            return True
    except Exception:
        pass

    # 2) popup container resolved via aria-owns
    popup_id = _resolve_main_menu_popup_id(driver)
    if popup_id:
        try:
            popup = driver.find_element(By.ID, popup_id)
            style = (popup.get_attribute("style") or "").lower()
            if popup.is_displayed() or "visibility: visible" in style:
                return True
        except Exception:
            pass

    # 3) the menu item we ultimately want is already visible
    try:
        item = driver.find_element(By.CSS_SELECTOR, "tr.import_csv")
        if item.is_displayed():
            return True
    except Exception:
        pass

    return False


def _find_main_menu_trigger(driver, timeout=5):
    """Locate the Menu-button trigger with resilient fallbacks.

    Dijit regenerates the widget ID suffix on re-render.  Try the expected
    ID first, then fall back to stable class/widgetid selectors.
    Returns the outer wrapper element (the div.mainMenu) or None.
    """
    selectors = [
        (By.ID, MAIN_MENU_TRIGGER_ID),
        (By.CSS_SELECTOR, f"*[widgetid='{MAIN_MENU_TRIGGER_ID}']"),
        (By.CSS_SELECTOR, "div.mainMenu.main-menu"),
        (By.CSS_SELECTOR, "div.mainMenu.dijitDownArrowButton"),
        (By.CSS_SELECTOR, "div.mainMenu"),
    ]
    for by, sel in selectors:
        try:
            el = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((by, sel))
            )
            return el
        except TimeoutException:
            continue
    return None


def open_main_menu(driver, timeout=WAIT_LONG):
    """
    Opens the Ariat Dojo main menu (the top-right "Menu" dropdown button).

    This is a dijit click-to-open dropdown (class contains
    'dijitDownArrowButton'), NOT a hover menu — so click strategies come
    first, using methods that fire a real native click event and propagate
    it through Dojo's ondijitclick handler on the dijitButtonNode.

    Strategies (in order), each followed by a short wait for the popup:
      1. Native Selenium .click() on the inner role=button span
      2. Native Selenium .click() on the dijitButtonNode parent (where
         the ondijitclick listener actually lives)
      3. ActionChains real mouse click at the button's center
      4. Keyboard activation: focus the inner button + press Enter/Space
      5. JS click on the inner role=button span
      6. JS click on the offscreen <input> helper (onclick:_onClick)
      7. Synthetic MouseEvent dispatch (mousedown + mouseup + click)
      8. Hover fallbacks (for the rare case the widget behaves as hover)
    """
    end = time.time() + timeout
    last_err = None
    attempt = 0

    while time.time() < end:
        if is_main_menu_open(driver):
            return
        attempt += 1
        print(f"[INFO] open_main_menu: cycle {attempt}")

        trigger_outer = _find_main_menu_trigger(driver, timeout=5)
        if trigger_outer is None:
            last_err = TimeoutException("Could not locate main menu trigger wrapper.")
            print("[WARN] open_main_menu: trigger wrapper not found, retrying...")
            time.sleep(0.5)
            continue

        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", trigger_outer)
        except Exception:
            pass

        # Inner clickable elements
        trigger_inner = None       # inner <span role="button">
        trigger_button_node = None # <span class="dijitButtonNode"> (has ondijitclick)
        trigger_offscreen = None   # <input class="dijitOffScreen">
        try:
            trigger_inner = trigger_outer.find_element(By.CSS_SELECTOR, "span[role='button']")
        except Exception:
            pass
        try:
            trigger_button_node = trigger_outer.find_element(
                By.CSS_SELECTOR, "span.dijitButtonNode"
            )
        except Exception:
            pass
        try:
            trigger_offscreen = trigger_outer.find_element(
                By.CSS_SELECTOR, "input.dijitOffScreen"
            )
        except Exception:
            pass

        def _wait_open(sec=2.0):
            """Poll is_main_menu_open for up to `sec` seconds."""
            deadline = time.time() + sec
            while time.time() < deadline:
                if is_main_menu_open(driver):
                    return True
                time.sleep(0.1)
            return False

        strategies = []
        if trigger_inner is not None:
            strategies.append(("native click inner[role=button]", lambda: trigger_inner.click()))
        if trigger_button_node is not None:
            strategies.append(("native click dijitButtonNode",    lambda: trigger_button_node.click()))
        if trigger_inner is not None:
            strategies.append((
                "ActionChains click inner",
                lambda: ActionChains(driver).move_to_element(trigger_inner).pause(0.1).click().perform()
            ))
            strategies.append((
                "keyboard Enter on inner",
                lambda: trigger_inner.send_keys(Keys.ENTER)
            ))
            strategies.append((
                "JS click inner",
                lambda: driver.execute_script("arguments[0].click();", trigger_inner)
            ))
        if trigger_offscreen is not None:
            strategies.append((
                "JS click offscreen input",
                lambda: driver.execute_script("arguments[0].click();", trigger_offscreen)
            ))
        if trigger_inner is not None:
            strategies.append((
                "dispatch MouseEvent sequence",
                lambda: driver.execute_script(
                    "var el=arguments[0];"
                    "['mousedown','mouseup','click'].forEach(function(t){"
                    "  el.dispatchEvent(new MouseEvent(t,{bubbles:true,cancelable:true,view:window}));"
                    "});",
                    trigger_inner,
                )
            ))
        # Hovers last — included only in case the widget behaves as hover-trigger.
        if trigger_outer is not None:
            strategies.append((
                "hover outer (fallback)",
                lambda: ActionChains(driver).move_to_element(trigger_outer).perform()
            ))
        if trigger_inner is not None:
            strategies.append((
                "hover inner (fallback)",
                lambda: ActionChains(driver).move_to_element(trigger_inner).perform()
            ))

        for name, action in strategies:
            try:
                action()
            except Exception as e:
                last_err = e
                print(f"[WARN] open_main_menu: strategy '{name}' raised: {_short_err(e)}")
                continue

            if _wait_open(sec=2.0):
                print(f"[INFO] open_main_menu: opened via '{name}'")
                return
            else:
                print(f"[WARN] open_main_menu: '{name}' did not open popup within 2s")

        time.sleep(0.5)

    resolved = _resolve_main_menu_popup_id(driver) or "unresolved"
    raise TimeoutException(
        f"Timed out opening main menu popup (trigger={MAIN_MENU_TRIGGER_ID}, "
        f"aria-owns={resolved}). Last error: {last_err}"
    )


def wait_for_import_menu_item(driver, timeout=WAIT_LONG):
    """Wait until the 'Import a File' menu item is present in the DOM (menu is open)."""
    open_main_menu(driver, timeout=timeout)
    # Use the stable import_csv class; fall back to text-content match
    for by, sel in [
        (By.CSS_SELECTOR, "tr.import_csv"),
        (By.XPATH, "//td[normalize-space()='Import a File']/ancestor::tr[1]"),
    ]:
        try:
            return WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((by, sel))
            )
        except TimeoutException:
            continue
    raise TimeoutException("Could not confirm 'Import a File' menu item is present.")


def click_import_a_file(driver, timeout=WAIT_LONG):
    """Activate the 'Import a File' menu item.

    Dijit menu items use event delegation on their parent menu widget —
    a plain JS click on the <tr> is often ignored.  The reliable click
    targets, in order, are:
        - the inner <td class="dijitMenuItemLabel"> (real <td> with text)
        - the <tr role="menuitem"> itself via native/ActionChains click
    We try each with several click methods and detect success by waiting
    for the import dialog's "Paste From Clipboard" control to appear.
    """
    _DIALOG_SIGNAL_XPATH = (
        "//*[contains(@class,'singleValue') and normalize-space()='Paste From Clipboard']"
    )

    def _dialog_open_now() -> bool:
        try:
            return any(e.is_displayed() for e in driver.find_elements(By.XPATH, _DIALOG_SIGNAL_XPATH))
        except Exception:
            return False

    # ── FAST PATH (2026-08-31) ────────────────────────────────────────────────
    # Open the menu via Dojo's own dropdown API and JS-click the visible label:
    # the exact combination that succeeded on every observed run.  The legacy
    # ladder below fired native clicks while the menu was still animating,
    # which double-toggled it closed and burned 10-15s of per-strategy waits
    # (plus stacktrace spam) on every single order.
    for attempt in (1, 2):
        try:
            open_main_menu_v2(driver, timeout=WAIT_MED)
            label = _find_displayed(
                driver, By.XPATH,
                "//tr[contains(@class,'import_csv')]//td[contains(@class,'dijitMenuItemLabel')]"
                " | //td[contains(@class,'dijitMenuItemLabel') and normalize-space()='Import a File']",
                timeout=8, context="(Import a File label)")
            driver.execute_script("arguments[0].click();", label)
        except Exception as e:
            print(f"[WARN] click_import_a_file: fast path attempt {attempt} raised: {_short_err(e)}")
            continue
        deadline = time.time() + 6
        while time.time() < deadline:
            if _dialog_open_now():
                print("[INFO] click_import_a_file: dialog opened (fast path)")
                return
            time.sleep(0.15)
        print(f"[WARN] click_import_a_file: fast path attempt {attempt} did not open the dialog")

    print("[WARN] click_import_a_file: falling back to the legacy click ladder")
    # ── LEGACY LADDER (fallback) ──────────────────────────────────────────────
    open_main_menu(driver, timeout=timeout)
    time.sleep(0.4)  # let menu items finish rendering after popup opens
    _ROW_SELECTORS = [
        (By.CSS_SELECTOR, "tr.import_csv"),
        (By.XPATH,        "//tr[contains(@class,'import_csv')]"),
        (By.XPATH,        "//td[normalize-space()='Import a File']/ancestor::tr[1]"),
    ]

    def _dialog_opened() -> bool:
        try:
            els = driver.find_elements(By.XPATH, _DIALOG_SIGNAL_XPATH)
            return any(e.is_displayed() for e in els)
        except Exception:
            return False

    def _locate_row_and_label():
        row = None
        for by, sel in _ROW_SELECTORS:
            try:
                row = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((by, sel))
                )
                break
            except TimeoutException:
                continue
        if row is None:
            return None, None
        label = None
        try:
            label = row.find_element(By.CSS_SELECTOR, "td.dijitMenuItemLabel")
        except Exception:
            try:
                label = row.find_element(
                    By.XPATH, ".//td[normalize-space()='Import a File']"
                )
            except Exception:
                label = None
        return row, label

    row, label = _locate_row_and_label()
    if row is None:
        raise TimeoutException(
            "Could not locate 'Import a File' menu row using any selector."
        )
    print(f"[INFO] Located 'Import a File' row (label-td: {'yes' if label else 'no'})")

    # (strategy_name, target_kind, action)
    click_actions = [
        ("native click label",       "label",
            lambda el: el.click()),
        ("ActionChains click label", "label",
            lambda el: ActionChains(driver).move_to_element(el).pause(0.08).click().perform()),
        ("JS click label",           "label",
            lambda el: driver.execute_script("arguments[0].click();", el)),
        ("native click row",         "row",
            lambda el: el.click()),
        ("ActionChains click row",   "row",
            lambda el: ActionChains(driver).move_to_element(el).pause(0.08).click().perform()),
        ("JS click row",             "row",
            lambda el: driver.execute_script("arguments[0].click();", el)),
        ("MouseEvent seq on label",  "label",
            lambda el: driver.execute_script(
                "var el=arguments[0];"
                "['mousedown','mouseup','click'].forEach(function(t){"
                "  el.dispatchEvent(new MouseEvent(t,{bubbles:true,cancelable:true,view:window}));"
                "});",
                el,
            )),
    ]

    for name, target_kind, action in click_actions:
        # If the menu closed between attempts (a failed click can dismiss
        # it), re-open before retrying.
        if not is_main_menu_open(driver):
            print(f"[INFO] click_import_a_file: menu closed, re-opening before '{name}'")
            open_main_menu(driver, timeout=timeout)
            time.sleep(0.3)

        # Re-locate fresh each attempt — Dojo may rebuild the DOM.
        row, label = _locate_row_and_label()
        if row is None:
            print(f"[WARN] click_import_a_file: row not found before '{name}', skipping")
            continue

        target = label if (target_kind == "label" and label is not None) else row
        if target is None:
            continue

        try:
            action(target)
        except Exception as e:
            print(f"[WARN] click_import_a_file: '{name}' raised: {_short_err(e)}")
            continue

        # Give the dialog up to 4s to appear after the click
        deadline = time.time() + 4.0
        while time.time() < deadline:
            if _dialog_opened():
                print(f"[INFO] click_import_a_file: dialog opened via '{name}'")
                return row
            time.sleep(0.15)
        print(f"[WARN] click_import_a_file: '{name}' did not open dialog within 4s")

    raise TimeoutException(
        "Could not activate 'Import a File' menu item — all click strategies failed."
    )
# ────────────────────────────────────────────────────────────────────────────────


def wait_for_clipboard_dropdown(driver, timeout=WAIT_LONG):
    xpath = (
        "//div[contains(@class,'css-13483rh-control') and "
        ".//div[contains(@class,'css-1uccc91-singleValue') and normalize-space()='Paste From Clipboard']]"
    )
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath)))


def _react_control_by_display(driver, display_text: str, timeout=WAIT_MED):
    """
    Find the react-select control whose visible label (singleValue or placeholder)
    matches display_text.  Does NOT depend on auto-incremented react-select-N IDs.
    """
    xpath = (
        f"//div[contains(@class,'css-13483rh-control') and "
        f"(.//*[contains(@class,'singleValue') and normalize-space()='{display_text}'] or "
        f"  .//*[contains(@class,'placeholder') and normalize-space()='{display_text}'])]"
    )
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath)))


def _set_react_select_by_display(driver, display_text: str, value_text: str, timeout=WAIT_LONG):
    """
    Set a React Select dropdown by finding it via its current visible label.
    Uses the same send_keys fallback strategy as the original set_react_select_by_input_id
    which is what actually works in this app — click to open, try clicking the option,
    fall back to typing the value + Enter into the hidden input.
    """
    value_text = str(value_text).strip()

    def open_menu():
        ctrl = _react_control_by_display(driver, display_text, timeout=timeout)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ctrl)
        safe_click(driver, ctrl)
        time.sleep(0.15)
        # Find the actual text input inside the control for send_keys fallback
        try:
            inp = ctrl.find_element(By.XPATH, ".//input")
            try:
                safe_click(driver, inp)
            except Exception:
                pass
        except Exception:
            inp = None
        return inp, ctrl

    def pick_option(inp, ctrl):
        # Strategy 1: click a visible [role=option] element
        for opt_xpath in [
            f"//*[@role='option' and normalize-space()='{value_text}']",
            f"//div[contains(@class,'option') and normalize-space()='{value_text}']",
            f"//*[normalize-space()='{value_text}' and (self::div or self::span) and contains(@class,'option')]",
        ]:
            try:
                opt = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, opt_xpath))
                )
                safe_click(driver, opt)
                return True
            except TimeoutException:
                continue

        # Strategy 2: type the value into the input and press Enter
        if inp is not None:
            try:
                inp.send_keys(Keys.CONTROL, "a")
                inp.send_keys(value_text)
                inp.send_keys(Keys.ENTER)
                return True
            except Exception:
                pass

        return False

    for attempt in range(1, 4):
        inp, ctrl = open_menu()
        pick_option(inp, ctrl)

        # Verify the selection took — check singleValue text inside the control
        try:
            WebDriverWait(driver, 8).until(
                lambda d: (
                    bool(ctrl.find_elements(By.XPATH,
                        f".//*[contains(@class,'singleValue') and normalize-space()='{value_text}']"
                    )) or
                    bool(ctrl.find_elements(By.XPATH,
                        f".//*[normalize-space()='{value_text}']"
                    ))
                )
            )
            print(f"[INFO] React-select set: '{display_text}' → '{value_text}'")
            return
        except TimeoutException:
            if attempt == 3:
                current = ctrl.text.strip()
                raise TimeoutException(
                    f"Failed to set '{display_text}' → '{value_text}'. Control text now: '{current}'"
                )
            time.sleep(0.6)


def ensure_custom_file_selected(driver, timeout=WAIT_LONG):
    """Switch the import-mode dropdown from 'Paste From Clipboard' to 'Custom File'."""
    # Short-circuit if already set
    try:
        _react_control_by_display(driver, "Custom File", timeout=3)
        print("[INFO] Import mode already set to 'Custom File'")
        return
    except TimeoutException:
        pass

    # Use the combined display-text finder + send_keys setter
    _set_react_select_by_display(driver, "Paste From Clipboard", "Custom File", timeout=timeout)


def _react_control_for_input(driver, input_id: str):
    inp = WebDriverWait(driver, WAIT_LONG).until(EC.presence_of_element_located((By.ID, input_id)))
    ctrl = inp.find_element(By.XPATH, "./ancestor::div[contains(@class,'css-13483rh-control')][1]")
    return inp, ctrl


def _react_control_text(ctrl) -> str:
    txt = (ctrl.text or "").strip()
    txt = re.sub(r"\s+", " ", txt)
    return txt


def set_react_select_by_input_id(driver, input_id: str, value_text: str, timeout=WAIT_LONG):
    value_text = str(value_text).strip()

    def open_menu():
        inp, ctrl = _react_control_for_input(driver, input_id)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ctrl)
        safe_click(driver, ctrl)
        time.sleep(0.15)
        try:
            safe_click(driver, inp)
        except Exception:
            pass
        return inp, ctrl

    def pick_option(inp, ctrl):
        try:
            opt = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, f"//*[@role='option' and normalize-space()='{value_text}']"))
            )
            safe_click(driver, opt)
            return True
        except TimeoutException:
            pass

        try:
            inp.send_keys(Keys.CONTROL, "a")
            inp.send_keys(value_text)
            inp.send_keys(Keys.ENTER)
            return True
        except Exception:
            return False

    for attempt in range(1, 4):
        inp, ctrl = open_menu()
        _ = pick_option(inp, ctrl)

        try:
            WebDriverWait(driver, 8).until(
                lambda d: (
                    value_text == _react_control_text(ctrl) or
                    f" {value_text} " in f" {_react_control_text(ctrl)} " or
                    bool(ctrl.find_elements(By.XPATH, f".//div[contains(@class,'singleValue') and normalize-space()='{value_text}']")) or
                    bool(ctrl.find_elements(By.XPATH, f".//*[normalize-space()='{value_text}']"))
                )
            )
            return
        except TimeoutException:
            if attempt == 3:
                final_txt = _react_control_text(ctrl)
                raise TimeoutException(
                    f"Failed to set {input_id} to '{value_text}'. Control text now: '{final_txt}'"
                )
            time.sleep(0.6)
# ────────────────────────────────────────────────────────────────────────────────


def find_latest_matching_file(order_no: str) -> str:
    # For Ariat uploads, ONLY look for files containing "ariat" or "carhartt"
    # This prevents accidentally selecting wrangler or other vendor files
    patterns = [
        os.path.join(DOWNLOAD_FOLDER, f"*{order_no}*ariat*carhartt*.*"),
        os.path.join(DOWNLOAD_FOLDER, f"*{order_no}*ariat*.*"),
        os.path.join(DOWNLOAD_FOLDER, f"*{order_no}*carhartt*.*"),
    ]
    candidates = []
    for pat in patterns:
        for f in glob.glob(pat):
            if f.lower().endswith((".xlsx", ".xls", ".xlsm", ".xlsb", ".ods", ".csv", ".txt")):
                candidates.append(f)

    if not candidates:
        available = os.listdir(DOWNLOAD_FOLDER)
        raise FileNotFoundError(
            f"No Ariat upload file found for order {order_no} in {DOWNLOAD_FOLDER}.\n"
            f"Looked for files matching: *{order_no}*ariat* or *{order_no}*carhartt*\n"
            f"Folder contains (first 80): {available[:80]}"
        )

    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def _read_shipto_from_csv(csv_path: str) -> dict:
    df = pd.read_csv(csv_path, dtype=str)
    if df.empty:
        raise ValueError(f"PO CSV {csv_path} has no rows.")
    row = df.iloc[0].to_dict()

    def get(*keys):
        for k in keys:
            if k in row and pd.notna(row[k]):
                return coerce_str(row[k])
        return ""

    company   = get("shipToCompany", "K")
    attention = get("shipToAttention", "L")
    street    = get("shipToStreet", "M")
    city      = get("shipToCity", "N")
    state     = get("shipToState", "O")
    zipc      = get("shipToZip", "P")

    name_line = (company + (attention or "")).strip()  # no separator requested

    return {
        "name_line": name_line,
        "street": street,
        "city": city,
        "state_abbr": state.upper().strip(),
        "zip": zipc,
        "csv_path": csv_path,
    }


def load_shipto_from_po_csv(po_number: str) -> dict:
    """
    Handles naming mismatch:
      - orders sheet might have '162945-297239'
      - CSV is named '297239.csv' and PO column is '297239'
    """
    po_raw = coerce_str(po_number)
    po_key = extract_po_key(po_raw)

    pats = [
        os.path.join(PDF_DIR, f"{po_raw}.csv"),
        os.path.join(PDF_DIR, f"{po_key}.csv"),
        os.path.join(PDF_DIR, f"*{po_key}*.csv"),
        os.path.join(PDF_DIR, f"*{po_raw}*.csv"),
    ]

    matches = []
    for pat in pats:
        matches.extend(glob.glob(pat))

    # If we found matches, prefer exact po_key.csv, then exact po_raw.csv, then most recent
    if matches:
        # De-dupe while preserving order
        seen = set()
        uniq = []
        for m in matches:
            if m not in seen:
                uniq.append(m)
                seen.add(m)

        exact_key = os.path.join(PDF_DIR, f"{po_key}.csv")
        exact_raw = os.path.join(PDF_DIR, f"{po_raw}.csv")

        if exact_key in uniq:
            return _read_shipto_from_csv(exact_key)
        if exact_raw in uniq:
            return _read_shipto_from_csv(exact_raw)

        uniq.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return _read_shipto_from_csv(uniq[0])

    # Fallback: scan CSVs and match their PO column
    all_csvs = glob.glob(os.path.join(PDF_DIR, "*.csv"))
    for csv_path in all_csvs:
        try:
            df = pd.read_csv(csv_path, dtype=str, nrows=1)
            if df.empty:
                continue
            if "PO" in df.columns:
                po_val = extract_po_key(coerce_str(df.iloc[0].get("PO", "")))
                if po_val == po_key:
                    return _read_shipto_from_csv(csv_path)
        except Exception:
            continue

    raise FileNotFoundError(
        f"Could not find PO CSV for '{po_raw}' (PO key '{po_key}') in {PDF_DIR}.\n"
        f"Tried patterns: {pats}"
    )


def wait_for_shipto_data(po_number: str):
    """Return ship-to data for po_number, pausing when the address CSV is missing.

    If no CSV with a usable ship-to address is found, emit an '[ADDRESS_MISSING]'
    marker line (the TSG app watches for it and shows a Try Again / Skip popup)
    and block on stdin for the user's decision:
      - 'retry <po>' → look for the CSV again
      - 'skip <po>'  → return None so the caller skips this order
    Responses tagged with a different PO (stale popup answers) are ignored.
    """
    po_key = extract_po_key(po_number)
    expected_po = str(po_key).strip().lower()
    while True:
        try:
            addr = load_shipto_from_po_csv(po_number)
        except Exception as e:
            print(f"[WARN] Could not load address CSV for PO {po_number}: {e}")
            addr = None
        if addr and addr.get("street"):
            return addr

        print("")
        print(f"[WARN] No usable address CSV for PO {po_number} in {PDF_DIR}")
        print(f"[ADDRESS_MISSING] {po_key}")
        print("[ACTION REQUIRED] Missing ship-to address — choose Try Again or Skip in the TSG app.")
        print(f"(Running standalone? Type 'retry {po_key}' or 'skip {po_key}' and press Enter.)")

        try:
            resp = input().strip().lower()
        except EOFError:
            # No stdin attached — safest is to skip rather than guessing an address.
            print(f"[WARN] Input closed; skipping order for PO {po_number}.")
            return None

        parts = resp.split(None, 1)
        action = parts[0] if parts else ""
        resp_po = parts[1].strip() if len(parts) > 1 else ""

        if resp_po and expected_po and resp_po != expected_po:
            print(f"[INFO] Ignoring stale response for PO '{resp_po}' (waiting on {po_key}).")
            continue
        if action == 'skip':
            print(f"[INFO] User chose to SKIP PO {po_number}.")
            return None
        if action == 'retry':
            print(f"[INFO] Retrying address lookup for PO {po_number}...")
            continue
        # Anything else (e.g., a bare Enter from 'Verification Complete') → check again
        print("[INFO] Unrecognized input; checking for the CSV again...")


def wait_for_submit_enter():
    """Block until the user confirms order submission (bare Enter / app button).

    Ignores stale 'retry <po>' / 'skip <po>' replies left over from the
    address-missing prompt so they can never trigger a premature submit.
    """
    while True:
        resp = input().strip()
        if not resp:
            return
        if resp.lower().startswith(("retry", "skip")):
            print(f"[INFO] Ignoring stale address-prompt reply ('{resp}').")
            continue
        return


def _shell_is_ready(driver) -> bool:
    """True once the Dojo order-builder shell is present (the Menu widget).

    Detected via the stable '.mainMenu' class rather than the auto-incrementing
    dijit widget id — Dojo regenerates the numeric suffix on re-render.
    """
    try:
        return any(e.is_displayed() for e in driver.find_elements(By.CSS_SELECTOR, "div.mainMenu"))
    except Exception:
        return False


def _on_catalog_list(driver) -> bool:
    """True when we've reached the catalog-selection list (post 'Shop Now')."""
    try:
        if "cataloglist" in (driver.current_url or "").lower():
            return True
        return bool(driver.find_elements(By.CSS_SELECTOR, "[data-testid^='card-carousel-image']"))
    except Exception:
        return False


def _click_shop_now(driver, timeout=60) -> bool:
    """Click the post-login 'Shop Now' button reliably.

    The marketing landing page renders 'Shop Now' as a <button> gated by a
    '.not-allowed-wrapper' overlay until catalog data loads, so it is NOT
    immediately interactable after login.  The old one-shot 15s attempt missed
    it on slow loads.  Here we poll for it (up to `timeout`s), JS-click it
    (bypasses the overlay), and confirm we advanced to the catalog list.
    Returns True once we've advanced past the marketing page.
    """
    end = time.time() + timeout
    while time.time() < end:
        if _shell_is_ready(driver) or _on_catalog_list(driver):
            return True

        btn = None
        for e in driver.find_elements(By.XPATH, "//button[normalize-space()='Shop Now']"):
            try:
                if e.is_displayed():
                    btn = e
                    break
            except Exception:
                continue

        if btn is not None:
            try:
                driver.execute_script("arguments[0].click();", btn)
                print("[INFO] Clicked 'Shop Now'")
            except Exception as e:
                print(f"[WARN] 'Shop Now' click raised: {e}")
            # Wait briefly for the click to register / page to advance
            adv_end = time.time() + 8
            while time.time() < adv_end:
                if _shell_is_ready(driver) or _on_catalog_list(driver):
                    return True
                time.sleep(0.4)
        else:
            print("[INFO] 'Shop Now' not present yet — waiting for marketing page to render...")
        time.sleep(0.8)

    return _shell_is_ready(driver) or _on_catalog_list(driver)


def _click_catalog_tile(driver, timeout=30) -> bool:
    """Click the first catalog tile on the catalog-list view to enter the
    order builder.  Tries several selectors and re-checks shell readiness."""
    selectors = [
        "[data-testid='card-carousel-image-0']",
        "[data-testid^='card-carousel-image']",
        ".slick-slide.slick-current",
        ".slick-slide",
    ]
    end = time.time() + timeout
    while time.time() < end:
        if _shell_is_ready(driver):
            return True
        clicked = False
        for sel in selectors:
            for e in driver.find_elements(By.CSS_SELECTOR, sel):
                try:
                    if e.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", e)
                        driver.execute_script("arguments[0].click();", e)
                        print(f"[INFO] Clicked catalog tile ({sel})")
                        clicked = True
                        break
                except Exception:
                    continue
            if clicked:
                break
        # Give the navigation up to 5s to bring up the Dojo shell, else retry
        wait_end = time.time() + 5
        while time.time() < wait_end:
            if _shell_is_ready(driver):
                return True
            time.sleep(0.4)
        time.sleep(0.5)
    return _shell_is_ready(driver)


def enter_order_builder(driver, timeout=WAIT_LONG):
    """Drive the post-login  marketing -> catalog -> Dojo order-builder  hops.

    This sequence is the historical cause of the 'works less than it works'
    flakiness.  The old code fired a single 15s 'Shop Now' click immediately
    after login and, if the page had not finished rendering, silently fell
    through and then waited (in vain) for the Dojo menu that only exists once
    the order builder loads.  We now poll/verify each hop instead.
    """
    if _shell_is_ready(driver):
        print("[INFO] Order-builder shell already present.")
        return

    if not _click_shop_now(driver, timeout=max(timeout, 60)):
        print("[WARN] Could not confirm 'Shop Now'/catalog list — continuing to look for shell anyway.")
    _click_catalog_tile(driver, timeout=30)

    # Wait for the Dojo order-builder shell (Menu widget) — resilient to the
    # auto-incrementing dijit id; match on the stable '.mainMenu' class.
    print("[INFO] Waiting for Dojo order-builder shell (.mainMenu)...")
    end = time.time() + max(timeout, 60)
    while time.time() < end:
        if _shell_is_ready(driver):
            print("[INFO] Order-builder shell is ready.")
            return
        time.sleep(0.5)
    raise TimeoutException(
        f"Order-builder shell (.mainMenu) never appeared. URL={driver.current_url}"
    )


def login_and_land(driver):
    wait = WebDriverWait(driver, WAIT_LONG)
    driver.get(ARIAT_URL)
    wait_ready(driver)

    user = wait.until(EC.visibility_of_element_located((By.NAME, "username")))
    user.clear()
    user.send_keys(ARIAT_USERNAME)

    pwd = wait.until(EC.visibility_of_element_located((By.NAME, "password")))
    pwd.clear()
    pwd.send_keys(ARIAT_PASSWORD)

    click_button_by_text(driver, "Login", timeout=WAIT_LONG)

    # Drive the post-login  marketing -> catalog -> Dojo order-builder  transition
    # robustly.  (This hop was the historical cause of the intermittent failures:
    # a slow marketing page meant 'Shop Now' wasn't clicked within the old 15s
    # window, so the app shell never loaded and the menu wait timed out.)
    enter_order_builder(driver, timeout=WAIT_LONG)

    wait_ready(driver)
    time.sleep(1.5)   # let Dojo finish widget registration before we interact

    # NOTE (2026-08-31): the old code opened the Menu here just to verify the
    # 'Import a File' item exists, then the first order re-opened it — a
    # redundant open (and ladder run) that also left the menu dangling.  The
    # shell-ready check above plus the import step's own menu handling cover it.


def import_file_flow(driver, upload_path: str):
    click_import_a_file(driver, timeout=WAIT_LONG)

    # Wait for the import dialog to be ready (Paste From Clipboard dropdown visible)
    _react_control_by_display(driver, "Paste From Clipboard", timeout=WAIT_LONG)

    # Switch to Custom File mode
    ensure_custom_file_selected(driver, timeout=WAIT_LONG)

    # Upload the file
    file_input = wait_present(driver, By.CSS_SELECTOR, "input[type='file']", timeout=WAIT_LONG)
    file_input.send_keys(upload_path)

    click_button_by_text(driver, "Next", timeout=WAIT_LONG)

    # Map columns by placeholder text — immune to react-select-N ID shifts
    # The UPC/EAN/SKU dropdown maps to column A; Quantity maps to column B
    _set_react_select_by_display(driver, "UPC/EAN/SKU", "A", timeout=WAIT_LONG)
    _set_react_select_by_display(driver, "Quantity",    "B", timeout=WAIT_LONG)

    click_button_by_text(driver, "Next", timeout=WAIT_LONG)


# ─── SALES-PROGRAMS UPSELL MODAL ─────────────────────────────────────────────
# After 'Proceed to Checkout' on the cart page, Ariat sometimes pops a React
# 'Apply Sales Programs' modal (a volume-discount upsell).  It appears for some
# carts and not others depending on the items/quantities — which is exactly why
# the script "worked less than it worked": orders that triggered the modal hung
# at checkout (the modal blocked the drop-ship button), while orders that did
# not sailed through.  We dismiss it by clicking its 'Proceed' button, which
# continues at standard pricing with no program applied (the historical, money-
# neutral behaviour for this unattended automation).  The button classes are
# hashed styled-components, so we match on the modal text + button text.
_SALES_MODAL_XPATH = (
    "//div[contains(@class,'ReactModal__Content')]"
    "[.//*[normalize-space()='Apply Sales Programs']]"
)


def _sales_programs_modal_present(driver) -> bool:
    try:
        return any(e.is_displayed() for e in driver.find_elements(By.XPATH, _SALES_MODAL_XPATH))
    except Exception:
        return False


def _dismiss_sales_programs_modal(driver, timeout=WAIT_SHORT) -> bool:
    """If the 'Apply Sales Programs' upsell modal is showing, click its
    'Proceed' button (continue without applying a program).  Returns True if a
    modal was dismissed."""
    try:
        WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.XPATH, _SALES_MODAL_XPATH))
        )
    except TimeoutException:
        return False

    print("[INFO] 'Apply Sales Programs' modal detected — proceeding without applying a program.")
    for xp in (
        _SALES_MODAL_XPATH + "//button[normalize-space()='Proceed']",
        "//div[contains(@class,'ReactModal__Content')]//button[normalize-space()='Proceed']",
    ):
        try:
            btn = WebDriverWait(driver, WAIT_SHORT).until(EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            safe_click(driver, btn)
            print("[INFO] Clicked 'Proceed' on the Sales Programs modal.")
            try:
                WebDriverWait(driver, WAIT_MED).until(
                    EC.invisibility_of_element_located((By.XPATH, _SALES_MODAL_XPATH))
                )
            except TimeoutException:
                pass
            return True
        except TimeoutException:
            continue
    print("[WARN] Sales Programs modal present but its 'Proceed' button was not clickable.")
    return False


# ─── 'PRODUCT AVAILABILITY ISSUE' (CANCEL DATE) POPUP ─────────────────────────
# On back-orders, some items become available only AFTER the shipment's cancel
# date; Ariat then pops this modal at checkout, which blocks everything under
# it (this froze/killed the 2026-08-31 back-order run for PO 168818-310482).
# Policy (per MY): always choose the FURTHER-OUT cancel date, then Save.
_AVAIL_MODAL_XPATH = (
    "//div[contains(@class,'ReactModal__Content')]"
    "[.//*[contains(normalize-space(.),'Product Availability Issue')]]"
)


def handle_availability_cancel_date_popup(driver, timeout=WAIT_SHORT) -> bool:
    """If the cancel-date modal is showing (or appears within `timeout`),
    select the LATEST date in every shipment's radio group and Save.

    Radio values are ISO dates ('2026-12-14'), grouped per shipment via the
    radio name ('Shipment 1', ...), so a lexicographic max is the latest date.
    Returns True if a modal was handled."""
    end = time.time() + timeout
    modal = None
    while time.time() < end:
        for m in driver.find_elements(By.XPATH, _AVAIL_MODAL_XPATH):
            try:
                if m.is_displayed():
                    modal = m
                    break
            except Exception:
                continue
        if modal is not None:
            break
        time.sleep(0.3)
    if modal is None:
        return False

    print("[INFO] 'Product Availability Issue' popup detected — selecting the further-out cancel date(s).")
    picked = driver.execute_script("""
        var modal = arguments[0];
        var groups = {};
        modal.querySelectorAll('input[type=radio]').forEach(function(r){
            (groups[r.name] = groups[r.name] || []).push(r);
        });
        var out = [];
        Object.keys(groups).forEach(function(name){
            var radios = groups[name].slice().sort(function(a, b){
                return String(a.value || '').localeCompare(String(b.value || ''));
            });
            var latest = radios[radios.length - 1];
            if (latest && !latest.checked) latest.click();
            if (latest) out.push(name + ' -> ' + latest.value);
        });
        return out;
    """, modal)
    for line in picked or []:
        print(f"[INFO] Cancel date choice: {line}")
    time.sleep(0.5)

    # The radio click triggers a React re-render, which STALES the old modal
    # element handle (observed live) — re-locate the visible modal fresh and
    # click Save inside it.
    try:
        save_btn = _find_displayed(
            driver, By.XPATH,
            _AVAIL_MODAL_XPATH + "//button[normalize-space()='Save']",
            timeout=WAIT_SHORT, context="(cancel-date Save button)")
        safe_click(driver, save_btn)
        print("[INFO] Clicked Save on the cancel-date popup.")
    except Exception as e:
        debug_dump(driver, "cancel_date_save_missing")
        raise RuntimeError(f"Cancel-date popup shown but its Save button was not clickable: {_short_err(e)}")

    if _wait_none_displayed(driver, By.XPATH, _AVAIL_MODAL_XPATH, timeout=WAIT_MED):
        print("[INFO] Cancel-date popup closed.")
    else:
        print("[WARN] Cancel-date popup still visible after Save.")
    return True


def proceed_to_checkout_flow(driver):
    # Step 1: 'Proceed to Checkout' on the order page (Dojo proceedBtn) -> cart.
    try:
        wait_and_click(
            driver,
            By.XPATH,
            "//span[contains(@class,'proceedBtn') and .//span[contains(@class,'dijitButtonText') and normalize-space()='Proceed to Checkout']]",
            timeout=WAIT_LONG
        )
    except TimeoutException:
        wait_and_click(driver, By.XPATH, "//*[normalize-space()='Proceed to Checkout']", timeout=WAIT_LONG)

    # Step 2: 'Proceed to Checkout' on the cart page (React button) -> this is
    # what opens the optional Sales Programs modal.  Stop clicking as soon as the
    # modal appears or we've already reached shipping (drop-ship button present).
    # NOTE: visibility check, not presence — after the first order Dojo keeps a
    # stale hidden btnDropShip in the DOM, which made presence checks think we
    # had already reached shipping on the second order.
    for _ in range(3):
        if _sales_programs_modal_present(driver) or _any_displayed(driver, By.CSS_SELECTOR, "span.btnDropShip"):
            break
        try:
            # 8s bound (was 30): the loop itself retries, so a long per-try
            # wait only stretched the rare worst case.
            click_button_by_text(driver, "Proceed to Checkout", timeout=8)
        except TimeoutException:
            pass
        time.sleep(1.0)

    # Steps 3+4: dismiss the optional modal, then confirm we've advanced to the
    # shipping step.  Loop because the exact ordering varies: keep dismissing the
    # modal / re-clicking Proceed until the drop-ship button is present.
    end = time.time() + WAIT_LONG
    while time.time() < end:
        if _any_displayed(driver, By.CSS_SELECTOR, "span.btnDropShip"):
            # Back-orders: the cancel-date modal can arrive with the checkout
            # page and blocks everything below it — resolve it before we
            # declare the checkout ready.
            handle_availability_cancel_date_popup(driver, timeout=3)
            return
        if _sales_programs_modal_present(driver):
            _dismiss_sales_programs_modal(driver, timeout=WAIT_SHORT)
        else:
            try:
                click_button_by_text(driver, "Proceed to Checkout", timeout=3)
            except TimeoutException:
                pass
        time.sleep(1.0)

    raise TimeoutException(
        "span.btnDropShip never appeared after Proceed-to-Checkout / Sales-Programs handling. "
        f"URL={driver.current_url}"
    )



def handle_address_confirmation_popup(driver, timeout=WAIT_LONG):
    """
    Handle the address confirmation popup that appears after saving address.
    - Always selects "Suggested Address" (usually pre-selected)
    - Clicks "Use Selected Address" button
    """
    try:
        # Wait for the modal to appear
        modal = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".ReactModal__Content[aria-label='Confirm Address']"))
        )
        print("[INFO] Address confirmation popup detected")
        
        # The "Suggested Address" radio button is usually already selected by default
        # But let's ensure it's selected by clicking it
        try:
            suggested_radio = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "input[name='selectedAddress'][value='addressFromSmarty']"))
            )
            if not suggested_radio.is_selected():
                suggested_radio.click()
                print("[INFO] Selected 'Suggested Address'")
            else:
                print("[INFO] 'Suggested Address' already selected")
        except Exception as e:
            print(f"[WARN] Could not verify suggested address selection: {e}")
        
        # Click "Use Selected Address" button
        use_button = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Use Selected Address')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", use_button)
        time.sleep(0.5)  # Brief pause to ensure button is ready
        safe_click(driver, use_button)
        print("[INFO] Clicked 'Use Selected Address'")
        
        # Wait for modal to close
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".ReactModal__Content[aria-label='Confirm Address']"))
        )
        print("[INFO] Address confirmation popup closed")
        
    except TimeoutException:
        print("[INFO] No address confirmation popup appeared (this is okay)")
    except Exception as e:
        print(f"[WARN] Error handling address confirmation popup: {e}")


def _handle_address_validation_warning(driver, timeout: int = 6) -> bool:
    """
    After clicking Save on the shipping address form, the site sometimes shows a
    Dijit validation warning:
        'We could not find a match for the address entered below.
         Please double check the fields highlighted in red.
         If the above address is confirmed, please click Save to continue.'

    When this banner is present the Save button must be clicked a second time to
    confirm and proceed.  If the banner does not appear within `timeout` seconds
    we assume the address was accepted on the first click and return False.

    Returns True if the warning was detected and bypassed, False otherwise.
    """
    WARNING_CSS = "div.dijitTextBoxError"
    WARNING_TEXT = "We could not find a match for the address"

    try:
        # Wait briefly for the warning banner to appear
        banner = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, WARNING_CSS))
        )
        if WARNING_TEXT.lower() not in (banner.text or "").lower():
            # Different error — don't swallow it; let the caller surface it
            return False

        print("[WARN] Address validation warning detected — clicking Save again to confirm.")
        click_dijit_button_by_label(driver, "Save", timeout=WAIT_LONG, prefer_id="dijit_form_Button_40")
        print("[INFO] Second Save click sent to confirm unmatched address.")

        # Wait for the warning banner to disappear — confirms the form accepted
        # the second Save and has closed or moved on.
        try:
            WebDriverWait(driver, WAIT_MED).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, WARNING_CSS))
            )
            print("[INFO] Address validation warning dismissed — form closed successfully.")
        except TimeoutException:
            print("[WARN] Warning banner did not disappear after second Save — proceeding anyway.")

        return True

    except TimeoutException:
        # Banner never appeared — address was accepted first time, nothing to do
        return False
    except Exception as e:
        print(f"[WARN] Unexpected error while handling address validation warning: {e}")
        return False


def fill_drop_ship_address(driver, po_number: str, addr: dict = None):
    # Address is normally pre-loaded (and user-confirmed) by wait_for_shipto_data()
    if addr is None:
        addr = load_shipto_from_po_csv(po_number)

    # Defensive: a late-arriving cancel-date modal would block the drop-ship
    # form from opening (exact failure of the 2026-08-31 back-order run).
    handle_availability_cancel_date_popup(driver, timeout=2)

    # Displayed-aware lookup: a hidden stale btnDropShip from the previous
    # order must not be the click target.
    ds_btn = _find_displayed(driver, By.CSS_SELECTOR, "span.btnDropShip",
                             timeout=WAIT_LONG, context="(drop-ship button)")
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ds_btn)
    safe_click(driver, ds_btn)

    def fill_by_input_name(input_name: str, value: str):
        if not value:
            return
        inp = _find_displayed(driver, By.CSS_SELECTOR, f"input[name='{input_name}']",
                              timeout=WAIT_LONG, context=f"(drop-ship field '{input_name}')")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        inp.clear()
        inp.send_keys(value)

    fill_by_input_name("name", addr["name_line"])
    fill_by_input_name("address1", addr["street"])
    fill_by_input_name("city", addr["city"])
    fill_by_input_name("zip", addr["zip"])

    state_ab = addr["state_abbr"]
    state_name = US_STATE_ABBR_TO_NAME.get(state_ab, "")
    if state_name:
        # Displayed-aware: the previous order's drop-ship modal (state dropdown
        # AND its popup menu items) lingers hidden in the DOM, so first-match
        # waits bound to the stale copy and timed out.  Confirmed live on
        # 2026-08-28: order 2's visible dropdown said 'select...' while a
        # hidden 'Louisiana' copy from order 1 soaked up every click attempt.
        st = _find_displayed(driver, By.CSS_SELECTOR, "span.dijitSelect.state",
                             timeout=WAIT_LONG, context="(state dropdown)")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", st)
        safe_click(driver, st)

        item = _find_displayed(
            driver, By.XPATH,
            f"//td[contains(@class,'dijitMenuItemLabel') and normalize-space()='{state_name}']",
            timeout=WAIT_LONG, context=f"(state menu item '{state_name}')")
        safe_click(driver, item)
    else:
        print(f"[WARN] Unknown/blank state abbreviation '{state_ab}' for PO {po_number}. Please select state manually.")

    # ✅ Save (Dijit) - click the actual button node, not just the inner text span
    click_dijit_button_by_label(driver, "Save", timeout=WAIT_LONG, prefer_id="dijit_form_Button_40")

    # ✅ Combined post-Save poll (2026-08-31): wait for whichever comes FIRST —
    #    (a) the address-validation warning (needs a second Save to confirm),
    #    (b) the React 'Confirm Address' modal (pick suggested address),
    #    (c) the form simply closing (address accepted silently).
    #    The old sequence always burned a fixed 6s waiting for a warning that
    #    rarely appears before it even started waiting for the modal.
    WARNING_CSS = "div.dijitTextBoxError"
    CONFIRM_CSS = ".ReactModal__Content[aria-label='Confirm Address']"
    handled_warning = False
    end = time.time() + WAIT_LONG
    while time.time() < end:
        # (a) validation warning → second Save confirms the unmatched address
        if not handled_warning:
            try:
                banners = [e for e in driver.find_elements(By.CSS_SELECTOR, WARNING_CSS) if e.is_displayed()]
            except Exception:
                banners = []
            if any("could not find a match" in (b.text or "").lower() for b in banners):
                print("[WARN] Address validation warning detected — clicking Save again to confirm.")
                click_dijit_button_by_label(driver, "Save", timeout=WAIT_MED)
                handled_warning = True
                time.sleep(0.5)
                continue
        # (b) confirm-address modal
        if _any_displayed(driver, By.CSS_SELECTOR, CONFIRM_CSS):
            handle_address_confirmation_popup(driver, timeout=WAIT_MED)
            break
        # (c) form closed with no modal pending → done
        if not _any_displayed(driver, By.CSS_SELECTOR, "input[name='address1']"):
            print("[INFO] Address form closed (no confirmation popup needed).")
            break
        time.sleep(0.3)
    else:
        print("[WARN] Post-Save state still unresolved after wait — continuing.")

    return addr

def fill_po_number_field(driver, po_number: str):
    po = coerce_str(po_number)
    # Displayed-aware: the fixed dijit id regenerates per render, and after the
    # first order a stale hidden poNumber input can linger — first-match
    # visibility waits then time out even though a fresh visible input exists.
    inp = _find_displayed(driver, By.XPATH, "//input[contains(@id,'poNumber') and @type='text']",
                          timeout=WAIT_LONG, context="(PO number field)")
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
    inp.clear()
    inp.send_keys(po)



def click_place_order_button(driver, timeout=WAIT_LONG):
    """
    Click the 'Place Order' button to initiate order submission.
    """
    print("[INFO] Clicking 'Place Order' button...")
    try:
        # Try by widgetid first
        place_order_btn = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "*[widgetid='finalSubmitButton']"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", place_order_btn)
        time.sleep(0.5)
        safe_click(driver, place_order_btn)
        print("[INFO] 'Place Order' button clicked")
    except TimeoutException:
        # Fallback: try clicking by button text
        print("[INFO] Trying fallback method to find 'Place Order' button...")
        click_dijit_button_by_label(driver, "Place Order", timeout=timeout)
        print("[INFO] 'Place Order' button clicked (fallback method)")


def handle_order_confirmation_popup(driver, timeout=WAIT_LONG):
    """
    Handle the order confirmation popup that appears after clicking 'Place Order'.
    Clicks the 'Submit' button in the confirmation dialog.
    """
    print("[INFO] Waiting for order confirmation popup...")
    try:
        # Wait for a VISIBLE confirmation dialog.  Dojo keeps the previous
        # order's dialog hidden in the DOM, so a presence check on the second
        # order matched the stale order-1 dialog instantly, "clicked" its
        # hidden Submit, then saw it "close" — while the real dialog for this
        # order was never touched.
        dialog = _find_displayed(driver, By.CSS_SELECTOR, ".dijitDialog.modal-confirm",
                                 timeout=timeout, context="(order confirmation dialog)")
        print("[INFO] Order confirmation popup detected (visible instance)")

        # Wait a moment for the dialog to fully render
        time.sleep(1)

        # Click the Submit button INSIDE this visible dialog — never by the
        # auto-incrementing dijit_form_Button_NN id, which points at a
        # different widget every render.
        try:
            submit_btn = dialog.find_element(
                By.XPATH,
                ".//span[contains(@class,'dijitButtonText') and normalize-space()='Submit']"
                "/ancestor::*[@role='button'][1]"
            )
        except Exception:
            submit_btn = dialog.find_element(
                By.XPATH,
                ".//span[contains(@class,'dijitButtonText') and normalize-space()='Submit']/.."
            )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", submit_btn)
        safe_click(driver, submit_btn)
        print("[INFO] Clicked 'Submit' button in confirmation popup")

        # Wait until NO confirm dialog is visible any more (hidden stale copies
        # in the DOM don't count).
        if _wait_none_displayed(driver, By.CSS_SELECTOR, ".dijitDialog.modal-confirm", timeout=timeout):
            print("[INFO] Order confirmation popup closed")
        else:
            print("[WARN] Order confirmation popup still visible after Submit click")

    except TimeoutException:
        print("[WARN] Order confirmation popup did not appear within timeout")
    except Exception as e:
        print(f"[WARN] Error handling order confirmation popup: {e}")


def extract_order_id_from_success_popup(driver, timeout=WAIT_LONG):
    """
    Wait for the order submission success popup and extract the order ID.
    Returns the order ID string (e.g., "10744371") or None if not found.
    """
    print("[INFO] Waiting for order submission success popup...")
    try:
        # Wait for a VISIBLE success dialog and read the order id from INSIDE
        # that instance.  Presence-based lookups matched the hidden stale
        # dialog from the previous order and could return order 1's ID for
        # order 2 (or "succeed" before this order's dialog even opened).
        success_dialog = _find_displayed(driver, By.CSS_SELECTOR, ".dijitDialog.submitOkModal",
                                         timeout=timeout, context="(order success dialog)")
        print("[INFO] Order submission success popup detected (visible instance)")

        # Find the description paragraph containing the order ID — scoped to
        # the visible dialog.
        description = success_dialog.find_element(
            By.CSS_SELECTOR, ".submitOkModalContents p[data-dojo-attach-point='description']"
        )

        # Extract the text (poll briefly — the text can render a beat later)
        success_text = ""
        end = time.time() + 10
        while time.time() < end:
            success_text = (description.text or "").strip()
            if success_text:
                break
            time.sleep(0.3)
        print(f"[INFO] Success message: {success_text}")

        # Extract order ID using regex (e.g., "Order 10744371 submitted successfully...")
        match = re.search(r"Order\s+(\d+)\s+submitted", success_text, re.IGNORECASE)
        if match:
            order_id = match.group(1)
            print(f"[SUCCESS] Extracted Order ID: {order_id}")

            # Click "Okay" INSIDE the visible dialog to close it
            try:
                okay_btn = success_dialog.find_element(
                    By.XPATH,
                    ".//span[contains(@class,'dijitButtonText') and normalize-space()='Okay']/.."
                )
                safe_click(driver, okay_btn)
                print("[INFO] Clicked 'Okay' to close success popup")
            except Exception:
                print("[WARN] Could not find 'Okay' button, popup may close automatically")

            # Let the dialog actually close before the caller moves on to the
            # next order — starting the next import while it is still up was
            # part of the second-order breakage.
            if _wait_none_displayed(driver, By.CSS_SELECTOR, ".dijitDialog.submitOkModal", timeout=WAIT_MED):
                print("[INFO] Success popup closed")
            else:
                print("[WARN] Success popup still visible — continuing anyway")

            return order_id
        else:
            print(f"[ERROR] Could not extract order ID from success message: {success_text}")
            debug_dump(driver, "success_text_unparsed")
            return None

    except TimeoutException:
        print("[ERROR] Order submission success popup did not appear within timeout")
        debug_dump(driver, "no_success_popup")
        return None
    except Exception as e:
        print(f"[ERROR] Error extracting order ID from success popup: {e}")
        debug_dump(driver, "success_popup_error")
        return None


def update_order_id_in_excel(excel_path: str, row_index: int, order_id: str):
    """
    Update the Order ID in column M (index 12) of the Excel file.
    If there's already a value in column M, append the new order ID with a space separator.
    
    Args:
        excel_path: Path to the Excel file
        row_index: The pandas DataFrame index (row number)
        order_id: The order ID to add
    """
    last_err = None
    for attempt in range(1, 4):
        try:
            print(f"[INFO] Updating Excel file with Order ID: {order_id} (attempt {attempt})")

            # Targeted openpyxl write: only touch the one column-M cell so the
            # rest of the workbook (values, fonts, date formatting) is preserved.
            # A full pandas read/to_excel round-trip would reset all formatting.
            wb = load_workbook(excel_path)
            ws = wb.active
            cell = ws.cell(row=row_index + 2, column=13)  # +2: header row, 1-based

            existing_value = coerce_str(cell.value)

            # Append or set the order ID
            if existing_value:
                new_value = f"{existing_value} {order_id}"
                print(f"[INFO] Appending to existing value: '{existing_value}' → '{new_value}'")
            else:
                new_value = order_id
                print(f"[INFO] Setting new Order ID: '{new_value}'")

            cell.value = new_value
            wb.save(excel_path)
            print(f"[SUCCESS] Excel file updated: {excel_path}")

            return True

        except Exception as e:
            # Most common cause: the workbook is open in Excel (PermissionError
            # on save).  Retry a couple of times, then record the ID in a
            # fallback file so a placed order's ID is never lost.
            last_err = e
            print(f"[ERROR] Failed to update Excel file (attempt {attempt}): {e}")
            time.sleep(2)

    try:
        fallback = os.path.join(SCRIPT_DIR, "order_ids_fallback.txt")
        with open(fallback, "a", encoding="utf-8") as f:
            f.write(f"{datetime.datetime.now():%Y-%m-%d %H:%M:%S}  row={row_index + 2}  order_id={order_id}\n")
        print(f"[WARN] Excel locked/unwritable — Order ID saved to {fallback}")
        print("[WARN] Close Processed_orders.xlsx in Excel and copy the ID into column M.")
    except Exception as e2:
        print(f"[ERROR] Could not write fallback order-id file either: {e2} (original error: {last_err})")
    return False


def main():
    options = webdriver.ChromeOptions()
    if TSG_DEBUG:
        options.add_argument(f"--remote-debugging-port={DEBUG_PORT}")
        print(f"[DEBUG] TSG_DEBUG on — Chrome DevTools will listen on 127.0.0.1:{DEBUG_PORT}")
    driver = webdriver.Chrome(options=options)

    try:
        try:
            login_and_land(driver)
        except Exception as e:
            print(f"[ERROR] Login / landing failed: {e}")
            traceback.print_exc()
            debug_dump(driver, "login_failed")
            debug_hold(driver, f"Login/landing failed: {e}")
            raise

        df = pd.read_excel(EXCEL_PATH, engine="openpyxl", dtype=str)

        # Column G (index 6) = PO number
        # Column J (index 9) = upload identifier
        # Column K (index 10) = Vendor
        col_g = df.columns[6]
        col_j = df.columns[9]
        col_k = df.columns[10]
        col_d = df.columns[3]   # Client PO number (used for file cleanup)

        skipped = []
        failed = []

        for idx, row in df.iterrows():
            po_number   = coerce_str(row[col_g])
            order_field = coerce_str(row[col_j])
            vendor      = coerce_str(row[col_k])
            client_po   = coerce_str(row[col_d])

            if not po_number:
                print(f"[SKIP] Row {idx}: blank PO in column G.")
                continue

            # CRITICAL: Only process Ariat orders (skip Wrangler, Propper, etc.)
            if "ariat" not in vendor.lower():
                print(f"[SKIP] Row {idx}: Not an Ariat order (Vendor: {vendor})")
                continue

            # CHECKPOINT: never place an order twice.  The ledger is
            # vendor-aware (safe for split rows); the column-M check is kept
            # as a backstop but only for single-vendor rows — on a split row
            # column M may hold the OTHER vendor's ID.
            prior = tsg_runlog.already_placed(SCRIPT_DIR, po_number, "ariat")
            if prior:
                print(f"[SKIP] Row {idx}: PO {po_number} already placed with Ariat on "
                      f"{prior.get('when','?')} (Order ID: {prior.get('order_id') or 'n/a'}).")
                continue
            existing_id = coerce_str(row[df.columns[12]])
            if existing_id and "/" not in vendor:
                print(f"[SKIP] Row {idx}: already placed (Order ID '{existing_id}' in column M).")
                continue

            m = re.search(r"\d+", order_field)
            if not m:
                raise ValueError(f"Row {idx}: cannot parse upload identifier from '{order_field}' (column J).")
            order_no = m.group()

            print(f"\n=== ARIAT ORDER START: PO={po_number}  UploadID={order_no} ===")

            # Require a usable ship-to address BEFORE touching the site so a
            # missing CSV can be fixed (Try Again) or the order skipped cleanly.
            addr = wait_for_shipto_data(po_number)
            if addr is None:
                skipped.append((po_number, client_po))
                print(f"[SKIP] Order PO={po_number} (Client PO {client_po}) skipped — no address CSV.")
                continue

            try:
                # After an order completes, Ariat may land outside the Dojo
                # order-builder (marketing/catalog page) or the shell may be
                # mid re-render.  Ensure the builder is up before importing —
                # starting the second order's import without this was a main
                # source of the "crashes on the second order" failures.
                if not _shell_is_ready(driver):
                    print("[INFO] Order-builder shell not present — re-entering order builder...")
                    enter_order_builder(driver, timeout=WAIT_LONG)
                    wait_ready(driver)
                    time.sleep(1.5)

                # CART GUARD: never import on top of a leftover cart — this is
                # what produced double-quantity orders after crashes/restarts.
                ensure_fresh_ariat_order(driver)

                upload_path = find_latest_matching_file(order_no)
                print(f"[INFO] Using upload file: {upload_path}")
                expected_units = expected_units_from_upload(upload_path)
                print(f"[INFO] Upload file contains {expected_units} unit(s).")

                import_file_flow(driver, upload_path)
                verify_ariat_cart(driver, expected_units, "after import")
                proceed_to_checkout_flow(driver)

                addr = fill_drop_ship_address(driver, po_number, addr=addr)
                print(f"[INFO] Address loaded from: {addr['csv_path']}")

                fill_po_number_field(driver, po_number)

                # Verify AGAIN on the checkout page before asking for review —
                # the badge stays visible there, so a mismatch is caught even
                # if something changed between import and checkout.
                verify_ariat_cart(driver, expected_units, "pre-review")

                # Wait for user to review and press Enter
                print("\n[ACTION REQUIRED]")
                print("Review the cart / shipping / totals")
                print("When ready, press Enter to automatically submit the order...")
                wait_for_submit_enter()

                # Automatically submit the order
                try:
                    verify_ariat_cart(driver, expected_units, "final pre-submit")
                    # A cancel-date modal can also (re)appear right before
                    # submission — resolve it so Place Order isn't blocked.
                    handle_availability_cancel_date_popup(driver, timeout=2)
                    click_place_order_button(driver, timeout=WAIT_LONG)
                    handle_order_confirmation_popup(driver, timeout=WAIT_LONG)
                    order_id = extract_order_id_from_success_popup(driver, timeout=WAIT_LONG)

                    if order_id:
                        tsg_runlog.record_placed(SCRIPT_DIR, po_number, "ariat", order_id)
                        # Update Excel with the order ID
                        update_order_id_in_excel(EXCEL_PATH, idx, order_id)
                        print(f"[SUCCESS] Order submitted successfully! Order ID: {order_id}")
                    else:
                        # Submission was attempted (Place Order + confirm both
                        # clicked) — checkpoint it so a re-run can't double-place.
                        # Marked UNCONFIRMED: verify on the vendor site, and if it
                        # truly did not go through, delete this PO's entry from
                        # placed_orders.json before re-running.
                        tsg_runlog.record_placed(SCRIPT_DIR, po_number, "ariat", "UNCONFIRMED")
                        print("[WARNING] Order may have been submitted, but Order ID could not be extracted.")
                        print("[WARNING] Checkpointed as UNCONFIRMED — check the vendor site; if it was NOT")
                        print("[WARNING] placed, remove this PO from placed_orders.json and re-run.")

                except Exception as e:
                    print(f"[ERROR] Error during order submission: {e}")
                    print("You may need to complete the order manually.")
                    input("Press Enter to continue to next order...")

            except Exception as e:
                # Per-order guard: capture the failing page, then (in debug
                # mode) hold the browser open for live inspection instead of
                # killing the whole run.
                print(f"[ORDER_ERROR] PO={po_number}: {e}")
                traceback.print_exc()
                debug_dump(driver, f"order_{extract_po_key(po_number)}")
                action = debug_hold(driver, f"Order PO={po_number} failed mid-flow: {e}")
                if action != "continue":
                    raise
                failed.append((po_number, client_po))
                continue

            print(f"=== ARIAT ORDER DONE: PO={po_number} ===")

        if failed:
            print("")
            print(f"[INFO] {len(failed)} Ariat order(s) FAILED mid-flow (see debug dumps):")
            for po, cpo in failed:
                print(f"  - PO {po} (Client PO {cpo})")
        if skipped:
            print("")
            print(f"[INFO] {len(skipped)} Ariat order(s) skipped (missing address CSV):")
            for po, cpo in skipped:
                print(f"  - PO {po} (Client PO {cpo})")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
