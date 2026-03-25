import os
import glob
import time
import re
import pandas as pd
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from openpyxl import load_workbook

# --------------------------------------------------------------------
# Environment-driven settings (provided by the controller app)
# --------------------------------------------------------------------
INITIALS        = os.getenv("ORDER_USER_INITIALS", "MY")
ADMIN_EMAIL     = os.getenv("BROBERRY_ADMIN_EMAIL", "internal3@broberry.com")
ADMIN_PASSWORD  = os.getenv("BROBERRY_ADMIN_PASSWORD", "MYoung454$")

# Workspace and folders
APP_DIR         = os.path.dirname(os.path.abspath(__file__))
WORKSPACE_DIR   = os.getenv("TSG_WORKSPACE_DIR", APP_DIR)

CSV_FOLDER      = os.path.join(WORKSPACE_DIR, "pdfs")  # <— where CSVs live (per your request)
DOWNLOAD_FOLDER = os.getenv("TSG_DOWNLOAD_DIR", os.path.join(os.path.expanduser("~"), "Downloads"))

TEMPLATE_XLSX   = os.path.join(WORKSPACE_DIR, "Example.xlsx")
OUTPUT_XLSX     = os.path.join(WORKSPACE_DIR, "Processed_orders.xlsx")

ADMIN_LOGIN_URL      = "https://admin.broberry.com/login"
ORDERS_URL           = "https://admin.broberry.com/orders"

# Written by BroberryShop.py when an order is skipped (back order / unresolvable)
SKIPPED_ORDERS_PATH  = os.path.join(WORKSPACE_DIR, "skipped_orders.xlsx")

# --------------------------------------------------------------------
# Vendor detection from SKU (so we can click the right XLS export buttons)
# --------------------------------------------------------------------
WRANGLER_SKU_PREFIXES = ("3W0", "10FR")
PROPPER_SKU_PREFIXES  = ("F52944X", "F52594X")
ARIAT_EXACT_SKUS      = {"10030232"}

def detect_vendor_from_sku(sku: str):
    if not sku:
        return None
    s = str(sku).strip().upper()
    if any(s.startswith(p) for p in WRANGLER_SKU_PREFIXES):
        return "wrangler"
    if any(s.startswith(p) for p in PROPPER_SKU_PREFIXES):
        return "propper"
    if s in ARIAT_EXACT_SKUS:
        return "ariat"
    return None

def detect_vendors_from_df(df: pd.DataFrame):
    # Try to find a SKU/product column in a tolerant way
    col_map = {c.lower().strip(): c for c in df.columns}
    for key in ("productid", "product_id", "sku", "product"):
        if key in col_map:
            sku_col = col_map[key]
            break
    else:
        return []

    vendors = set()
    for raw in df[sku_col].dropna().astype(str).tolist():
        v = detect_vendor_from_sku(raw)
        if v:
            vendors.add(v)
    return sorted(vendors)

# --------------------------------------------------------------------
# Skipped-orders loader (produced by BroberryShop.py)
# --------------------------------------------------------------------
def load_skipped_pos(path: str) -> set:
    """
    Returns a set of PO strings that BroberryShop.py logged as skipped.
    If the file doesn't exist (no orders were skipped) returns an empty set.
    """
    if not os.path.exists(path):
        return set()
    try:
        df = pd.read_excel(path, dtype=str)
        if "PO" not in df.columns:
            print(f"⚠️  skipped_orders.xlsx found but has no 'PO' column — ignoring.")
            return set()
        skipped = set(df["PO"].dropna().str.strip().tolist())
        print(f"\n=== Skipped Orders ===")
        print(f"Found {len(skipped)} skipped PO(s) in skipped_orders.xlsx: {', '.join(sorted(skipped))}")
        return skipped
    except Exception as e:
        print(f"⚠️  Could not read skipped_orders.xlsx: {e} — proceeding without skip list.")
        return set()


# --------------------------------------------------------------------

def get_records(csv_folder: str):
    """
    Reads first row from each CSV in csv_folder and returns a list of dicts.
    FIXED: Now handles both old and new column name formats (case-insensitive)
    """
    records = []
    csv_files = glob.glob(os.path.join(csv_folder, "*.csv"))
    
    print(f"\n=== CSV Discovery ===")
    print(f"Looking in folder: {csv_folder}")
    print(f"Found {len(csv_files)} CSV file(s)")
    
    for path in csv_files:
        print(f"\n--- Processing: {os.path.basename(path)} ---")
        try:
            df = pd.read_csv(path, dtype=str)
        except Exception as e:
            print(f"  ⚠️  Error reading with default encoding: {e}")
            try:
                df = pd.read_csv(path, dtype=str, encoding="latin-1")
                print(f"  ✓ Successfully read with latin-1 encoding")
            except Exception as e2:
                print(f"  ✖ Failed with latin-1 too: {e2}")
                continue
        
        if df.empty:
            print(f"  ⚠️  CSV is empty, skipping")
            continue
        
        # Create case-insensitive column lookup
        col_map = {col.lower(): col for col in df.columns}
        print(f"  Columns found: {list(df.columns)}")
        
        row = df.iloc[0]
        
        # Look for email column (case-insensitive)
        email = ""
        for col_key in ["email", "account", "acct"]:
            if col_key in col_map:
                actual_col = col_map[col_key]
                email = (row.get(actual_col) or "").strip()
                print(f"  ✓ Found email in column '{actual_col}': {email}")
                break
        
        if not email:
            print(f"  ⚠️  No email column found, using first column")
            email = str(row.iloc[0]).strip()
        
        # Look for PO column (case-insensitive) - THIS IS THE FIX
        po = ""
        for col_key in ["po", "p.o.", "purchase order", "purchaseorder"]:
            if col_key in col_map:
                actual_col = col_map[col_key]
                po = (row.get(actual_col) or "").strip()
                print(f"  ✓ Found PO in column '{actual_col}': {po}")
                break
        
        if not po:
            print(f"  ✖ No PO column found in columns: {list(df.columns)}")
        
        # Look for Order-Cost column (case-insensitive)
        order_cost = ""
        for col_key in ["order-cost", "order cost", "ordercost", "cost", "total"]:
            if col_key in col_map:
                actual_col = col_map[col_key]
                order_cost = (row.get(actual_col) or "").strip()
                print(f"  ✓ Found cost in column '{actual_col}': {order_cost}")
                break
        
        if not order_cost:
            print(f"  ⚠️  No Order-Cost column found")
        
        record = {
            "email": email,
            "PO": po,
            "Order-Cost": order_cost,
            "vendors": detect_vendors_from_df(df),
        }
        
        records.append(record)
        print(f"  ✓ Added record - Email: {email}, PO: {po}, Cost: {order_cost}")
    
    print(f"\n=== Summary ===")
    print(f"Total records extracted: {len(records)}")
    for i, rec in enumerate(records, 1):
        print(f"  {i}. Email: {rec['email']}, PO: {rec['PO']}, Cost: {rec['Order-Cost']}, Vendors: {','.join(rec.get('vendors', []))}")
    
    return records

# --------------------------------------------------------------------
# Selenium helpers
# --------------------------------------------------------------------
def setup_driver(download_folder: str):
    os.makedirs(download_folder, exist_ok=True)

    opts = Options()
    # Keep browser visible (comment next line if you want headless)
    # opts.add_argument("--headless=new")

    # Safer default prefs for downloading
    prefs = {
        "download.default_directory": download_folder,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "profile.default_content_settings.popups": 0,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "safebrowsing.enabled": True,
        "safebrowsing.disable_download_protection": True,
    }
    opts.add_experimental_option("prefs", prefs)
    opts.add_argument("--safebrowsing-disable-download-protection")

    # Let Selenium Manager resolve the driver (Selenium 4.10+)
    driver = webdriver.Chrome(options=opts)
    driver.maximize_window()
    return driver

def login(driver, login_url, user, pwd):
    driver.get(login_url)
    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.NAME, "email")))
    driver.find_element(By.NAME, "email").clear()
    driver.find_element(By.NAME, "email").send_keys(user)
    driver.find_element(By.NAME, "password").clear()
    driver.find_element(By.NAME, "password").send_keys(pwd)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    WebDriverWait(driver, 20).until(EC.invisibility_of_element_located((By.NAME, "email")))

def process_orders(driver, records, orders_url, skipped_pos: set = None):
    driver.get(orders_url)
    skipped_pos = skipped_pos or set()

    for rec in records:
        po = rec["PO"]

        # Skip any PO that BroberryShop.py flagged as unresolvable
        if po in skipped_pos:
            print(f"⏭️  Skipping PO {po} — logged as skipped by BroberryShop.")
            rec["skipped"] = True
            continue

        # 1) Search for the PO
        search = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "search"))
        )
        search.clear()
        search.send_keys(po, Keys.RETURN)
        time.sleep(2)

        # 2) Grab the order number text from the FIRST cell's <p>
        try:
            table_row = WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
            )
        except TimeoutException:
            raise TimeoutException("Orders table did not render; selector may have changed.")

        try:
            # Strictly pull from the first <td> -> <p> (e.g., <td>...<p>13266</p></td>)
            order_p = table_row.find_element(By.XPATH, ".//td[1]//p")
            text = order_p.text.strip()
            # Fallback: ensure it's numeric; if not, scan for the first purely numeric text in TD[1]
            if not text.isdigit():
                candidates = [
                    e.text.strip()
                    for e in table_row.find_elements(By.XPATH, ".//td[1]//*[self::p|self::span|self::div]")
                    if e.text and e.text.strip()
                ]
                text = next((t for t in candidates if t.isdigit()), "")
            if not text:
                raise TimeoutException("First cell did not contain a numeric order number.")
            order_num = text
            rec["order_num"] = order_num
        except Exception:
            raise TimeoutException("Could not extract order number from the first table cell.")


        # 3) Click each download link for this order (based on vendors present in the CSV)
        titles = ["Download XML"]
        vendors = set((rec.get("vendors") or []))
        if 'propper' in vendors and 'ariat' in vendors:
            print(f"⚠️  PO {po}: vendors include both Propper and Ariat. Skipping vendor XLS downloads to avoid wrong export.")
            vendors.discard('propper')
            vendors.discard('ariat')
        if "wrangler" in vendors:
            titles.append("Download Wrangler")
        if "propper" in vendors:
            titles.append("Download Propper")
        if "ariat" in vendors:
            titles.append("Download Ariat/Carhartt")

        for title in titles:
            for attempt in range(3):
                try:
                    row = driver.find_element(By.XPATH, f"//tr[.//p[normalize-space()='{order_num}'] or .//a[normalize-space()='{order_num}']]")
                    link = row.find_element(By.XPATH, f".//a[@title='{title}']")
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
                    driver.execute_script("arguments[0].click();", link)
                    time.sleep(1)
                    break
                except StaleElementReferenceException:
                    time.sleep(1)
                except Exception:
                    if attempt == 2:
                        raise RuntimeError(f"Could not click '{title}' for order {order_num}")
                    time.sleep(1)

        print(f"Downloaded order {rec['order_num']}, waiting 5s before next...")
        time.sleep(5)

# --------------------------------------------------------------------
# Excel writer (unchanged layout; paths now under workspace)
# --------------------------------------------------------------------
def write_to_excel(template_path, output_path, records):
    wb = load_workbook(template_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    today = datetime.now().strftime("%m/%d/%Y")

    for rec in records:
        # Don't write rows for orders that were skipped by BroberryShop
        if rec.get("skipped"):
            print(f"⏭️  Omitting skipped PO {rec['PO']} from Excel output.")
            continue
        row_map = {
            "Name": today,
            "Customer": "The Sourcing Group",
            "Ack": rec["email"],
            "Client PO #": rec["PO"],
            "Cust Acct": "",
            "Who began order": INITIALS,
            "BMI Order #/ Full retailers PO": "",
            "Who finalized order": INITIALS,
            "Date PO finalized": today,
            "Notes/F/up date & who": f"Order #: {rec.get('order_num','')}",
            "Vendor": " / ".join((rec.get("vendors") or ["Wrangler"])).title(),
            "Transaction ID": "Terms",
            "Order Number in vendor system": "",
            "GP%": "19%",
            "Item Amount": rec["Order-Cost"],
            "Freight (CC or N30)": "Cust Acct"
        }
        ws.append([row_map.get(h, "") for h in headers])

    wb.save(output_path)

# --------------------------------------------------------------------
# Main
# --------------------------------------------------------------------
def main():
    # Ensure folders exist
    os.makedirs(CSV_FOLDER, exist_ok=True)
    os.makedirs(os.path.dirname(OUTPUT_XLSX), exist_ok=True)

    records = get_records(CSV_FOLDER)
    if not records:
        print(f"No CSV files found in: {CSV_FOLDER}")
        return

    # Load any POs that BroberryShop.py marked as skipped so we don't crash on them
    skipped_pos = load_skipped_pos(SKIPPED_ORDERS_PATH)

    driver = setup_driver(DOWNLOAD_FOLDER)
    try:
        login(driver, ADMIN_LOGIN_URL, ADMIN_EMAIL, ADMIN_PASSWORD)
        process_orders(driver, records, ORDERS_URL, skipped_pos)
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    write_to_excel(TEMPLATE_XLSX, OUTPUT_XLSX, records)
    written = sum(1 for r in records if not r.get("skipped"))
    skipped = sum(1 for r in records if r.get("skipped"))
    print(f"Done! Wrote {written} row(s) to {OUTPUT_XLSX}" +
          (f" ({skipped} skipped PO(s) omitted)." if skipped else "."))

if __name__ == "__main__":
    main()
