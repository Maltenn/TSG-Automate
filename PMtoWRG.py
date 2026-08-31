import os
import re
import glob
import csv
import datetime
import pandas as pd
import time
import math

from openpyxl import load_workbook

import tsg_runlog

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

# ─── CONFIG ────────────────────────────────────────────────────────────────────
SCRIPT_DIR           = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH           = os.path.join(SCRIPT_DIR, 'Processed_orders.xlsx')
DOWNLOAD_FOLDER      = os.getenv("TSG_DOWNLOAD_DIR", os.path.join(os.path.expanduser("~"), "Downloads"))
LOGIN_URL            = "https://wranglerb2b.com/login.php/client/NQ=="
BATCH_ORDER_URL      = "https://wranglerb2b.com/batch_order.php/ecat_view"
CHECKOUT_URL         = "https://wranglerb2b.com/tp_checkout.php/ecat_checkout"
ORDER_HISTORY_URL    = "https://wranglerb2b.com/tp_order_history.php/ecat_view"

# Folder that contains per-PO CSVs produced by your PDF extraction pipeline.
# Example: C:\TSG_Automate\pdfs\297361.csv
PDFS_DIR             = os.path.join(SCRIPT_DIR, "pdfs")

# If the extracted ship-to value matches this (after normalization), we proceed
# with the normal checkout flow (select radio and done).
DEFAULT_SHIPTO_VALUE = "THE SOURCING GROUP, INC. | 4560 36TH STREET | ORLANDO, FL 32811 | FedEx Ground: 955617339,"

EMAIL    = os.getenv("WRANGLER_EMAIL")  or os.getenv("WRG_EMAIL")  or "internal3@broberry.com"
PASSWORD = os.getenv("WRANGLER_PASSWORD") or os.getenv("WRG_PASSWORD") or "Internal3Broberry!"

# Default Selenium waits (seconds)
WAIT_SHORT = 10
WAIT_LONG = 25
WAIT_XLONG = 60

# Debug mode: set TSG_DEBUG=1 to (a) expose Chrome DevTools on port 9224 so an
# external tool can attach and inspect the live page, (b) HOLD the browser open
# on errors instead of crashing out.  Failure screenshots/HTML are always saved
# (see debug_dump below).
TSG_DEBUG = os.getenv("TSG_DEBUG", "").strip().lower() not in ("", "0", "false", "no")
DEBUG_PORT = 9224

# State abbreviation to full name mapping
STATE_ABBREV_MAP = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "DC": "District of Columbia", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii",
    "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine",
    "MD": "Maryland", "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota",
    "MS": "Mississippi", "MO": "Missouri", "MT": "Montana", "NE": "Nebraska",
    "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico",
    "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island",
    "SC": "South Carolina", "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas",
    "UT": "Utah", "VT": "Vermont", "VA": "Virginia", "WA": "Washington",
    "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming"
}

# ────────────────────────────────────────────────────────────────────────────────

def log(msg: str) -> None:
    """Simple logger used throughout the script."""
    print(msg, flush=True)

def debug_dump(driver, error_name="error"):
    """Save screenshot and HTML for debugging."""
    try:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        screenshot_path = os.path.join(SCRIPT_DIR, f"debug_{error_name}_{timestamp}.png")
        html_path = os.path.join(SCRIPT_DIR, f"debug_{error_name}_{timestamp}.html")
        
        driver.save_screenshot(screenshot_path)
        log(f"[DEBUG] Screenshot saved: {screenshot_path}")
        
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(driver.page_source)
        log(f"[DEBUG] HTML saved: {html_path}")
    except Exception as e:
        log(f"[DEBUG] Failed to save debug info: {e}")

# ─── CART (ORDER PAD) VERIFICATION & SELF-HEALING (added 2026-08-31) ──────────
# Reproduced live: Wrangler's order pad is PER-DRAFT, but create_new_draft()
# with a name that already exists RESUMES the existing draft — items included.
# A crash-and-restart therefore resumed the dirty draft (same PO name), the
# re-upload stacked on top, and the auto-submit placed DOUBLE quantities.
# Guards: after draft creation the pad must be empty (else it is cleared), and
# after upload / before submit the pad must EXACTLY match the upload file.

def wrangler_pad_units(driver) -> int:
    """Unit count from the header order-pad badge ('N Items / N Units').

    The badge (.nav-rSideInfo) is present on every page, checkout included."""
    try:
        txt = driver.execute_script(
            "var b=document.querySelector('.nav-rSideInfo'); return b ? b.textContent : ''") or ""
    except Exception:
        return -1
    m = re.search(r"(\d+)\s*Units", txt.replace("\n", " "))
    return int(m.group(1)) if m else -1


def expected_units_from_upload_file(path: str) -> int:
    """Sum the units column of a Wrangler batch upload file (.xlsx or .xml)."""
    try:
        if path.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(path, dtype=str)
            col = None
            for c in df.columns:
                if str(c).strip().lower() in ("units", "qty", "quantity"):
                    col = c
                    break
            if col is None:
                return -1
            total = 0
            for v in df[col]:
                s = coerce_str(v)
                if s.isdigit():
                    total += int(s)
            return total
        # Best-effort XML: sum numeric qty/units tags or attributes
        with open(path, "r", encoding="utf-8", errors="replace") as fh:
            xml = fh.read()
        nums = re.findall(r"<(?:qty|units|quantity)>\s*(\d+)\s*</", xml, re.I)
        if not nums:
            nums = re.findall(r"(?:qty|units|quantity)\s*=\s*[\"'](\d+)[\"']", xml, re.I)
        return sum(int(n) for n in nums) if nums else -1
    except Exception as e:
        log(f"[WARN] Could not read expected units from '{path}': {e}")
        return -1


def wrangler_clear_pad(driver) -> None:
    """Empty the active draft's order pad via the site's own clear function.

    (The Clear All link runs confirm_clear_ecatalogs() which shows an in-page
    Yes/No modal — and the DOM contains a DECOY 'Yes' with no handler, so we
    call the real clear_ecatalogs() directly.)"""
    log("[CART_GUARD] Clearing the Wrangler order pad (clear_ecatalogs)...")
    driver.execute_script("if (typeof clear_ecatalogs === 'function') { clear_ecatalogs(); }")
    time.sleep(3.0)
    driver.refresh()
    wait_ready(driver, timeout=25)
    time.sleep(1.5)
    units = wrangler_pad_units(driver)
    if units != 0:
        debug_dump(driver, "pad_clear_failed")
        raise RuntimeError(f"Order pad still shows {units} unit(s) after Clear All — refusing to continue.")
    log("[CART_GUARD] Order pad is now empty.")


def ensure_fresh_wrangler_pad(driver) -> None:
    """Guarantee the just-activated draft's pad is EMPTY before uploading.

    If create_new_draft resumed an existing same-name draft (crashed run),
    its leftover items are cleared here — the upload then rebuilds the order
    from scratch with correct quantities."""
    time.sleep(1.5)
    units = wrangler_pad_units(driver)
    if units < 0:
        # Badge unreadable on this page — go somewhere it exists and retry.
        driver.get(BATCH_ORDER_URL)
        wait_ready(driver, timeout=25)
        time.sleep(1.0)
        units = wrangler_pad_units(driver)
    if units == 0:
        log("[INFO] Pad check: order pad is empty — OK to upload.")
        return
    log(f"[CART_GUARD] Order pad has {units} leftover unit(s) — the draft was RESUMED "
        "from a previous (crashed) run, not newly created!")
    debug_dump(driver, "leftover_pad_before_clear")
    wrangler_clear_pad(driver)


def verify_wrangler_pad(driver, expected_units: int, context: str, settle_timeout: int = 30) -> None:
    """Hard gate: pad units must EXACTLY match the upload file (polls while
    the site finishes adding batch items)."""
    if expected_units is None or expected_units < 0:
        log(f"[CART_VERIFY] SKIPPED ({context}): expected units unknown for this upload file type.")
        return
    end = time.time() + settle_timeout
    units = -1
    while time.time() < end:
        units = wrangler_pad_units(driver)
        if units == expected_units:
            log(f"[CART_VERIFY] OK ({context}): pad has {units} unit(s), expected {expected_units}.")
            return
        time.sleep(1.0)
    debug_dump(driver, f"pad_mismatch_{context.replace(' ', '_')}")
    raise RuntimeError(
        f"ORDER PAD MISMATCH ({context}): pad shows {units} unit(s) but the upload file "
        f"contains {expected_units} (waited {settle_timeout}s). NOT submitting — this is "
        "exactly how double-quantity orders happened."
    )


def debug_hold(driver, context=""):
    """In TSG_DEBUG mode, keep the browser open and wait for instructions on stdin.

    Returns 'continue' (move on to the next order) or 'abort' (end the run).
    In normal mode returns 'abort' immediately (fail-fast, as before) —
    continuing past a half-finished order risks putting the next order's items
    into a stale draft, so that decision is only offered to a human in debug mode.
    """
    if not TSG_DEBUG:
        return "abort"
    log("")
    log(f"[DEBUG_HOLD] {context}")
    log(f"[DEBUG_HOLD] Browser held open (DevTools on 127.0.0.1:{DEBUG_PORT}).")
    log("[DEBUG_HOLD] Type 'continue' to move to the next order, or 'abort' to end the run.")
    while True:
        try:
            resp = input().strip().lower()
        except EOFError:
            return "abort"
        if resp == "continue":
            log("[DEBUG_HOLD] Continuing with the next order.")
            return "continue"
        if resp == "abort":
            log("[DEBUG_HOLD] Aborting the run.")
            return "abort"
        if resp:
            log(f"[DEBUG_HOLD] Unrecognized input '{resp}' — type 'continue' or 'abort'.")


def cleanup_old_debug_files():
    """Remove old debug screenshots and HTML files from script directory."""
    try:
        import glob
        debug_files = glob.glob(os.path.join(SCRIPT_DIR, "debug_*.png")) + \
                     glob.glob(os.path.join(SCRIPT_DIR, "debug_*.html"))
        
        if debug_files:
            log(f"[INFO] Cleaning up {len(debug_files)} old debug files...")
            for file in debug_files:
                try:
                    os.remove(file)
                except Exception as e:
                    log(f"[WARN] Could not delete {file}: {e}")
            log(f"[INFO] Debug file cleanup complete")
        else:
            log("[INFO] No old debug files to clean up")
    except Exception as e:
        log(f"[WARN] Error during debug file cleanup: {e}")


def coerce_str(val) -> str:
    """Convert Excel/CSV values (floats, NaN, ints, None, etc.) to a clean string."""
    if val is None:
        return ""
    if isinstance(val, float):
        if math.isnan(val):
            return ""
        # If it's an integer-like float (e.g., 13092.0), return without .0
        if val.is_integer():
            return str(int(val))
        return str(val)
    if isinstance(val, (int,)):
        return str(val)
    return str(val).strip()


def _normalize_shipto(s: str) -> str:
    """Normalize ship-to strings for robust comparisons."""
    if s is None:
        return ""
    # Standardize whitespace + pipe formatting.
    s = str(s).replace("\r", " ").replace("\n", " ")
    s = re.sub(r"\s*\|\s*", " | ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _canonical_shipto(s: str) -> str:
    """A looser canonicalization used for fuzzy ship-to matching.

    Wrangler's saved ship-to labels often abbreviate (ST vs STREET), omit punctuation,
    and may include ZIP+4. The PDF-extracted 'shipTo' string can include commas/INC/etc.
    This helper reduces those differences so we can reliably detect the default
    "THE SOURCING GROUP" destination.
    """
    if s is None:
        return ""
    s = str(s).upper()
    # Replace common punctuation with spaces
    s = re.sub(r"[\.,]", " ", s)
    # Normalize street words
    s = s.replace("STREET", "ST")
    s = s.replace("AVENUE", "AVE")
    s = s.replace("ROAD", "RD")
    s = s.replace("DRIVE", "DR")
    # Remove common legal suffixes
    for tok in (" INC ", " INCORPORATED ", " LLC "):
        s = s.replace(tok, " ")
    # Collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_default_sourcing_group_shipto(shipto_text: str) -> bool:
    """Return True if shipto_text refers to THE SOURCING GROUP at 4560 36TH ST, Orlando, FL 32811."""
    norm = _normalize_shipto(shipto_text)
    # Fast-path exact-ish match
    if norm == _normalize_shipto(DEFAULT_SHIPTO_VALUE):
        return True

    canon = _canonical_shipto(shipto_text)
    # Fuzzy token match: tolerate abbreviations and ZIP+4.
    required = [
        "THE SOURCING GROUP",
        "4560",
        "36TH",
        "ORLANDO",
        "FL",
        "32811",
    ]
    return all(tok in canon for tok in required)


def find_po_csv_path(po_number: str) -> str:
    """Find a PO CSV in PDFS_DIR.

    Prefers an exact '<po>.csv'. Otherwise, falls back to a wildcard
    search for '*<po>*...*.csv' and chooses the most recently modified.
    Returns an empty string if nothing is found.
    """
    exact = os.path.join(PDFS_DIR, f"{po_number}.csv")
    if os.path.exists(exact):
        return exact

    try:
        pattern = os.path.join(PDFS_DIR, f"*{po_number}*.csv")
        candidates = glob.glob(pattern)
        if not candidates:
            # Sometimes extensions can be uppercase, depending on how it was saved
            pattern2 = os.path.join(PDFS_DIR, f"*{po_number}*.CSV")
            candidates = glob.glob(pattern2)
        if not candidates:
            return ""
        candidates.sort(key=lambda fp: os.path.getmtime(fp), reverse=True)
        return candidates[0]
    except Exception:
        return ""


def load_shipto_data_from_csv(po_number: str) -> dict:
    """Load ship-to data from the PO's CSV.
    
    Returns a dict with keys:
    - 'shipTo': Full ship-to text from column C (index 2)
    - 'company': shipToCompany from column K (index 10)
    - 'attention': shipToAttention from column L (index 11)
    - 'street': shipToStreet from column M (index 12)
    - 'city': shipToCity from column N (index 13)
    - 'state': shipToState from column O (index 14)
    - 'zip': shipToZip from column P (index 15)
    """
    csv_path = find_po_csv_path(po_number)
    if not csv_path:
        log(f"[WARN] PO CSV not found for {po_number}: {os.path.join(PDFS_DIR, f'{po_number}.csv')}")
        return {}

    # Try multiple encodings in order of likelihood
    encodings = ['utf-8-sig', 'cp1252', 'latin-1', 'utf-8', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            with open(csv_path, newline='', encoding=encoding) as f:
                reader = csv.reader(f)
                rows = list(reader)
                
                # Skip header row (row 0), data starts at row 1
                if len(rows) < 2:
                    log(f"[WARN] CSV file for {po_number} has insufficient data rows")
                    return {}
                
                data_row = rows[1]  # First data row
                
                # Build result dict with coerced values
                result = {
                    'shipTo': coerce_str(data_row[2]) if len(data_row) > 2 else "",
                    'company': coerce_str(data_row[10]) if len(data_row) > 10 else "",
                    'attention': coerce_str(data_row[11]) if len(data_row) > 11 else "",
                    'street': coerce_str(data_row[12]) if len(data_row) > 12 else "",
                    'city': coerce_str(data_row[13]) if len(data_row) > 13 else "",
                    'state': coerce_str(data_row[14]) if len(data_row) > 14 else "",
                    'zip': coerce_str(data_row[15]) if len(data_row) > 15 else "",
                }
                
                log(f"[INFO] Successfully read CSV with {encoding} encoding")
                return result
                
        except UnicodeDecodeError:
            # Try next encoding
            continue
        except Exception as e:
            log(f"[WARN] Failed reading PO CSV '{csv_path}' with {encoding}: {e}")
            return {}
    
    # If all encodings failed
    log(f"[ERROR] Could not read CSV '{csv_path}' with any supported encoding")
    return {}


def wait_for_shipto_data(client_po: str):
    """Return ship-to data for client_po, pausing when the address CSV is missing.

    If no CSV with a ship-to address is found, emit an '[ADDRESS_MISSING]'
    marker line (the TSG app watches for it and shows a Try Again / Skip
    popup) and block on stdin for the user's decision:
      - 'retry <po>' → look for the CSV again
      - 'skip <po>'  → return None so the caller skips this order
    Responses tagged with a different PO (stale popup answers) are ignored.
    Never falls back to the default Sourcing Group address.
    """
    expected_po = str(client_po).strip().lower()
    while True:
        shipto_data = load_shipto_data_from_csv(client_po)
        if shipto_data and shipto_data.get('shipTo'):
            return shipto_data

        log("")
        log(f"[WARN] No address CSV found for Client PO {client_po} in {PDFS_DIR}")
        log(f"[ADDRESS_MISSING] {client_po}")
        log("[ACTION REQUIRED] Missing ship-to address — choose Try Again or Skip in the TSG app.")
        log(f"(Running standalone? Type 'retry {client_po}' or 'skip {client_po}' and press Enter.)")

        try:
            resp = input().strip().lower()
        except EOFError:
            # No stdin attached — safest is to skip rather than silently
            # shipping to the default Sourcing Group address.
            log(f"[WARN] Input closed; skipping order for Client PO {client_po}.")
            return None

        parts = resp.split(None, 1)
        action = parts[0] if parts else ""
        resp_po = parts[1].strip() if len(parts) > 1 else ""

        if resp_po and expected_po and resp_po != expected_po:
            log(f"[INFO] Ignoring stale response for PO '{resp_po}' (waiting on {client_po}).")
            continue
        if action == 'skip':
            log(f"[INFO] User chose to SKIP Client PO {client_po}.")
            return None
        if action == 'retry':
            log(f"[INFO] Retrying address lookup for Client PO {client_po}...")
            continue
        # Anything else (e.g., a bare Enter from 'Verification Complete') → check again
        log("[INFO] Unrecognized input; checking for the CSV again...")


def coerce_date(val) -> str:
    """Return a mm/dd/yyyy string for dates coming from Excel/datetime/strings."""
    if val is None:
        return ""
    # Already a string → trust it
    if isinstance(val, str):
        return val.strip()
    # datetime/date → format it
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%m/%d/%Y")
    # Excel may pass floats (serials) or something odd; fall back to str
    return coerce_str(val)

def get_next_business_day(from_date=None):
    d = from_date or datetime.date.today()
    one_day = datetime.timedelta(days=1)
    d += one_day
    while d.weekday() > 4:  # Sat=5, Sun=6
        d += one_day
    return d

def login(driver):
    driver.get(LOGIN_URL)
    wait = WebDriverWait(driver, 20)

    # Fill credentials
    email_el = wait.until(EC.visibility_of_element_located((By.ID, "login_email")))
    email_el.clear()
    email_el.send_keys(EMAIL)

    pwd_el = driver.find_element(By.ID, "login_password")
    pwd_el.clear()
    pwd_el.send_keys(PASSWORD)

    # Pause for manual verification before submitting
    print("\n[ACTION REQUIRED]")
    print("Please complete any login verification in the browser window now.")
    print("Examples: CAPTCHA checkbox, 'I'm not a robot', or any pre-login security step.")
    input("When you're done and ready for the script to click Sign In, press Enter here... ")

    # Submit login after user confirms
    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

    # Continue with your original post-login wait
    wait.until(EC.presence_of_element_located((By.ID, "p7SOPt_2")))


def open_order_menu(driver):
    wait = WebDriverWait(driver, 10)
    
    # Wait for any overlays/preloaders to disappear
    wait_for_overlay_gone(driver, timeout=15)
    time.sleep(0.5)
    
    # Wait for the order menu button to be clickable
    order_menu_btn = wait.until(EC.element_to_be_clickable((By.ID, "p7SOPt_2")))
    
    # Use safe_click to handle any remaining interception issues
    safe_click(driver, order_menu_btn)
    
    # Wait for the New Draft option to appear
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.pop-newdraft")))


def to_text(x) -> str:
    """Coerce Excel/CSV values (floats, NaN, ints, None, etc.) to a clean string."""
    if x is None:
        return ""
    if isinstance(x, float):
        if math.isnan(x):
            return ""
        if x.is_integer():
            return str(int(x))
        return str(x)
    return str(x).strip()


def create_new_draft(driver, draft_name, ship_date):
    wait = WebDriverWait(driver, 10)
    
    # Wait for any overlays to disappear first
    wait_for_overlay_gone(driver, timeout=15)
    time.sleep(0.5)

    # Open the "New Draft" popup - wait for it to be clickable and use safe_click
    new_draft_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.pop-newdraft")))
    safe_click(driver, new_draft_btn)

    # Safely send the draft name and tab to the date field
    safe_name = to_text(draft_name)
    wait.until(EC.visibility_of_element_located(
        (By.ID, "pfm-newdraft"))
    ).send_keys(safe_name + "\t")

    # Pick the ship date from the calendar
    day = ship_date.day
    xpath = f"//td[@data-handler='selectDay']/a[text()='{day}']"
    wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()

    # Click "Save New Draft"
    driver.find_element(
        By.XPATH, "//button[@onclick='save_new_draft()']"
    ).click()

    # Make closing the popup optional / resilient
    try:
        # Old behaviour: explicit close button
        close_btn = WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[@onclick='preloadCloseWindow()']")
            )
        )
        close_btn.click()
    except TimeoutException:
        # If the site changed and there is no close button anymore,
        # just wait briefly for the popup/overlay to disappear and continue.
        try:
            WebDriverWait(driver, 8).until(
                EC.invisibility_of_element_located(
                    (By.ID, "fancybox-wrap")
                )
            )
        except TimeoutException:
            # As a last resort, just continue
            pass
    
    # Final wait for any overlays to clear before moving on
    wait_for_overlay_gone(driver, timeout=10)



def upload_batch_order(driver, order_no):
    wait = WebDriverWait(driver, 60)
    driver.get(BATCH_ORDER_URL)
    
    # Wait for page to load and any overlays to clear
    wait_ready(driver, timeout=25)
    wait_for_overlay_gone(driver, timeout=15)

    file_input = wait.until(EC.presence_of_element_located((By.ID, "load_items_file")))
    driver.execute_script("arguments[0].style.display = 'block';", file_input)

    pattern1 = os.path.join(DOWNLOAD_FOLDER, f"*{order_no}*wrangler*.*")
    candidates = [f for f in glob.glob(pattern1) if f.lower().endswith((".xml", ".xlsx"))]
    if not candidates:
        pattern2 = os.path.join(DOWNLOAD_FOLDER, f"*{order_no}*.*")
        candidates = [f for f in glob.glob(pattern2) if f.lower().endswith((".xml", ".xlsx"))]
    if not candidates:
        available = os.listdir(DOWNLOAD_FOLDER)
        raise FileNotFoundError(f"No file for order {order_no} in {DOWNLOAD_FOLDER}. Contains: {available}")

    # Newest file first — an old export for a re-used order number must not win
    candidates.sort(key=lambda fp: os.path.getmtime(fp), reverse=True)
    xml_path = candidates[0]
    print(f"[INFO] Uploading file: {xml_path}")
    file_input.send_keys(xml_path)
    # Shortened fixed waits (were 9s/16s): the Add button already has a
    # clickable-wait, and verify_wrangler_pad() polls the badge for up to 30s
    # after this returns, absorbing slow server-side adds.
    time.sleep(3)
    add_btn = wait.until(EC.element_to_be_clickable((
        By.XPATH, "//button[contains(@onclick,'add_ecat_items_to_cart_alert')]"
    )))
    add_btn.click()
    time.sleep(6)
    return xml_path


def wait_ready(driver, timeout=25):
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )

def wait_for_overlay_gone(driver, timeout=20):
    """Wait for common overlays/spinners to go away."""
    wait = WebDriverWait(driver, timeout)
    try:
        wait.until(EC.invisibility_of_element_located(
            (By.CSS_SELECTOR, ".fancybox-overlay, .modal-backdrop, .blockUI, .loading-overlay, .loading, [id^='fs-preloader']")
        ))
    except TimeoutException:
        pass


def safe_click(driver, el):
    """Click element, falling back to JavaScript if intercepted or not interactable."""
    from selenium.common.exceptions import ElementClickInterceptedException, ElementNotInteractableException
    try:
        el.click()
    except (ElementClickInterceptedException, ElementNotInteractableException) as e:
        log(f"[INFO] Regular click failed ({type(e).__name__}), using JavaScript click")
        driver.execute_script("arguments[0].click();", el)

def wait_modal_open(driver, timeout=10):
    """
    Fancybox usually injects .fancybox-overlay + .fancybox-wrap/.fancybox-inner.
    Wait until the radio list is present, VISIBLE, and clickable.
    """
    wait = WebDriverWait(driver, timeout)
    
    log("[DEBUG] Waiting for modal to open...")
    
    # Check if fancybox overlay appeared
    try:
        overlay = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, ".fancybox-overlay, .fancybox-wrap")
        ))
        log("[DEBUG] Fancybox overlay detected")
    except TimeoutException:
        log("[ERROR] Fancybox overlay never appeared!")
        debug_dump(driver, "modal_overlay_timeout")
        log("[DEBUG] Current URL: " + driver.current_url)
        raise
    
    # Wait for fancybox-inner (the actual modal content container) to be visible
    try:
        inner = wait.until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, ".fancybox-inner")
        ))
        log("[DEBUG] Fancybox inner content is visible")
    except TimeoutException:
        log("[WARN] Fancybox inner content not detected, continuing anyway...")
    
    # Wait for radio buttons to be present
    try:
        radio = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, "input[name='add_addresses1']")
        ))
        log("[DEBUG] Radio buttons found in DOM")
    except TimeoutException:
        log("[ERROR] Radio buttons never appeared in DOM!")
        debug_dump(driver, "radio_buttons_timeout")
        raise
    
    # CRITICAL: Wait for radio buttons to become VISIBLE (not just present)
    # The modal animates in, so buttons exist but are hidden initially
    log("[DEBUG] Waiting for radio buttons to become visible...")
    
    # Try scrolling within the modal in case there's an inner scroll container
    try:
        scroll_container = driver.find_element(By.CSS_SELECTOR, ".stylescrollA, .fancybox-inner")
        driver.execute_script("arguments[0].scrollTop = 0;", scroll_container)
        log("[DEBUG] Scrolled modal content to top")
    except:
        pass
    
    max_attempts = 15  # Try for up to 3 seconds (15 * 0.2s)
    for attempt in range(max_attempts):
        try:
            radio = driver.find_element(By.CSS_SELECTOR, "input[name='add_addresses1']")
            
            # Check if it's actually visible with real dimensions
            is_displayed = radio.is_displayed()
            size = radio.size
            location = radio.location
            
            log(f"[DEBUG] Attempt {attempt+1}: displayed={is_displayed}, size={size}, location={location}")
            
            if is_displayed and size['height'] > 0 and size['width'] > 0:
                log("[DEBUG] Radio button is now visible with real dimensions!")
                break
                
            time.sleep(0.2)
            
        except Exception as e:
            log(f"[DEBUG] Attempt {attempt+1} check failed: {e}")
            time.sleep(0.2)
    else:
        # If we exhausted all attempts
        log("[ERROR] Radio buttons never became visible after 3 seconds!")
        debug_dump(driver, "radio_not_visible")
        raise TimeoutException("Radio buttons present but never became visible")
    
    # Now wait for it to be clickable
    log("[DEBUG] Waiting for radio button to be clickable...")
    try:
        clickable_radio = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "input[name='add_addresses1']")
        ))
        log("[DEBUG] Radio button is now clickable")
        return clickable_radio
    except TimeoutException:
        log("[ERROR] Radio button never became clickable!")
        debug_dump(driver, "radio_not_clickable")
        raise

def wait_modal_close(driver, timeout=10):
    """Wait for modal and overlay to disappear."""
    wait = WebDriverWait(driver, timeout)
    wait.until(EC.invisibility_of_element_located(
        (By.CSS_SELECTOR, ".fancybox-overlay")
    ))


def _open_and_choose_ship_to_legacy(
    driver,
    preferred_radio_id: str = None,
    preferred_value_contains: str = None,
    preferred_label_contains: str = None,
    preferred_account_number: str = None,
    max_retries: int = 5,
):
    """
    LEGACY — replaced by the rewritten open_and_choose_ship_to below (kept for
    reference).  Its open-check and radio-visibility check lived in different
    retry scopes, so a modal that opened briefly and closed again passed the
    open-check and then hung on the radio wait.
    """
    wait = WebDriverWait(driver, 25)
    
    for attempt in range(max_retries):
        try:
            log(f"[DEBUG] Ship-To modal attempt {attempt + 1}/{max_retries}")
            
            # Check for any alerts or error messages first
            if attempt == 0:  # Only check on first attempt
                log("[DEBUG] Checking for alerts or error messages...")
                try:
                    alerts = driver.find_elements(By.CSS_SELECTOR, ".alert, .error, .warning, [role='alert']")
                    if alerts:
                        visible_alerts = [a for a in alerts if a.is_displayed()]
                        if visible_alerts:
                            log(f"[WARN] Found {len(visible_alerts)} visible alerts on page!")
                            for i, alert in enumerate(visible_alerts[:3]):
                                log(f"[WARN] Alert {i+1}: {alert.text[:100]}")
                except Exception as e:
                    log(f"[DEBUG] Error checking alerts: {e}")
            
            # Re-find the button on each attempt (avoid stale element)
            log("[DEBUG] Looking for Ship-To button...")
            try:
                shiptos_btn = wait.until(EC.element_to_be_clickable((
                    By.CSS_SELECTOR, "button.pop-myShipTos-1, button[class*='pop-myShipTos']"
                )))
                log(f"[DEBUG] Found Ship-To button with selector: button.pop-myShipTos-1")
            except TimeoutException:
                log("[DEBUG] First selector failed, trying XPath...")
                shiptos_btn = wait.until(EC.element_to_be_clickable((
                    By.XPATH, "//button[contains(., \"Available Ship-To\") or contains(., \"Ship-To\")]"
                )))
                log(f"[DEBUG] Found Ship-To button with XPath")
            
            log(f"[DEBUG] Ship-To button text: '{shiptos_btn.text}'")
            log(f"[DEBUG] Ship-To button is_displayed: {shiptos_btn.is_displayed()}, is_enabled: {shiptos_btn.is_enabled()}")
            
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", shiptos_btn)
            log("[DEBUG] Scrolled to Ship-To button")
            
            safe_click(driver, shiptos_btn)
            log("[DEBUG] Clicked Ship-To button")
            
            time.sleep(1.5)  # Longer wait for modal to appear
            log("[DEBUG] Waited 1.5s for modal animation to start")
            
            # Verify modal actually opened
            log("[DEBUG] Checking if modal opened after button click...")
            try:
                WebDriverWait(driver, 3).until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, ".fancybox-overlay, .fancybox-wrap")
                ))
                log("[DEBUG] Modal overlay detected after click")
                
                # If we got here, modal opened successfully - break out of retry loop
                break
                
            except TimeoutException:
                log(f"[WARN] Modal did not open on attempt {attempt + 1}")
                debug_dump(driver, f"modal_not_opened_attempt_{attempt + 1}")
                
                if attempt < max_retries - 1:
                    # Try refreshing the page for next attempt
                    log("[INFO] Refreshing page and retrying...")
                    driver.refresh()
                    time.sleep(4)  # Longer wait after refresh to let page fully load
                    
                    # Wait for page to be ready
                    WebDriverWait(driver, 15).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                    time.sleep(1)
                else:
                    # Last attempt failed
                    log("[ERROR] All attempts to open modal failed!")
                    raise TimeoutException("Ship-To modal never opened after multiple attempts")
        
        except TimeoutException:
            if attempt == max_retries - 1:
                raise
            log(f"[WARN] Attempt {attempt + 1} failed with timeout, will retry...")
            continue
    
    # Wait for modal to fully open
    wait_modal_open(driver, timeout=12)
    
    # Find and select the radio button
    all_radios = driver.find_elements(By.CSS_SELECTOR, "input[name='add_addresses1']")
    chosen_radio = None

    for radio in all_radios:
        radio_id = radio.get_attribute("id")
        radio_val = radio.get_attribute("value") or ""
        
        # Get the label text
        try:
            label = driver.find_element(By.XPATH, f"//label[input[@id='{radio_id}']]")
            label_text = label.text
        except NoSuchElementException:
            label_text = ""

        # Match criteria
        if preferred_radio_id and radio_id == preferred_radio_id:
            chosen_radio = radio
            log(f"[INFO] Matched ship-to by radio ID: {radio_id}")
            break
        if preferred_value_contains and preferred_value_contains in radio_val:
            chosen_radio = radio
            log(f"[INFO] Matched ship-to by value substring: '{preferred_value_contains}'")
            break
        if preferred_label_contains and preferred_label_contains in label_text:
            chosen_radio = radio
            log(f"[INFO] Matched ship-to by label substring: '{preferred_label_contains}'")
            break
        if preferred_account_number and f"account_number={preferred_account_number}" in radio_val:
            chosen_radio = radio
            log(f"[INFO] Matched ship-to by account number: {preferred_account_number}")
            break

    if chosen_radio is None:
        # Never gamble on an arbitrary ship-to — a wrong pick here means the
        # order ships/bills against the wrong account, silently.
        debug_dump(driver, "shipto_no_match")
        labels = []
        for radio in all_radios[:10]:
            try:
                rid = radio.get_attribute("id")
                lab = driver.find_element(By.XPATH, f"//label[input[@id='{rid}']]").text
                labels.append(f"{rid}: {lab[:80]}")
            except Exception:
                continue
        raise RuntimeError(
            "No ship-to radio matched THE SOURCING GROUP criteria. "
            f"Available options: {labels}"
        )

    # Select the radio button
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", chosen_radio)
    
    if not chosen_radio.is_selected():
        # Wait for the specific radio to be clickable
        try:
            wait.until(EC.element_to_be_clickable(chosen_radio))
            log("[DEBUG] Radio button is clickable, attempting click...")
        except TimeoutException:
            log("[WARN] Radio button wait timed out, will try clicking anyway...")
            time.sleep(0.5)
        
        # Try to click it
        try:
            safe_click(driver, chosen_radio)
            log("[DEBUG] Radio button clicked successfully")
        except Exception as e:
            log(f"[WARN] Regular click failed: {e}, using JavaScript to select radio...")
            # Fallback: Use JavaScript to directly set the radio as checked
            driver.execute_script("""
                arguments[0].checked = true;
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('click', { bubbles: true }));
            """, chosen_radio)
            log("[DEBUG] Radio button selected via JavaScript")


def _dismiss_checkout_preloader(driver, timeout=25) -> None:
    """Dismiss the fs-preloader 'Continue' overlay on the checkout page.

    The overlay runs an inventory check (3-15s) and blocks every click until
    its Continue button is pressed.  Safe to call when the overlay is absent.
    """
    try:
        wait = WebDriverWait(driver, timeout)
        wait.until(EC.presence_of_element_located((By.ID, "fs-preloader-1")))
        continue_btn = wait.until(EC.visibility_of_element_located((
            By.XPATH, "//div[@id='fs-preload-continue']//button[@onclick='preloadCloseWindow()']"
        )))
        try:
            continue_btn.click()
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", continue_btn)
            except Exception:
                try:
                    driver.execute_script("preloadCloseWindow();")
                except Exception as e:
                    log(f"[WARN] Could not dismiss checkout preloader: {e}")
        time.sleep(1)
        wait_for_overlay_gone(driver, timeout=10)
        log("[INFO] Checkout preloader dismissed.")
    except TimeoutException:
        pass  # overlay never appeared — nothing to dismiss


def open_and_choose_ship_to(
    driver,
    preferred_radio_id: str = None,
    preferred_value_contains: str = None,
    preferred_label_contains: str = None,
    preferred_account_number: str = None,
    max_retries: int = 5,
):
    """
    Open the Ship-To modal and select the specified ship-to address.

    Rewritten 2026-08-28 after catching the failure live:  the native click on
    the Ship-To button is reliably intercepted on this page, and the old code
    then fired a JS click as fallback — but the intercepted click had ALREADY
    triggered the fancybox, so the fallback click toggled it closed again.
    The old open-check saw the overlay during its brief open window, moved on,
    and then waited forever for radios inside a closed modal ("Wrangler just
    stops and moves on").

    Fixes here:
      - a SINGLE JS click per attempt (never native-then-JS double fire)
      - modal-open and radios-VISIBLE are verified as one unit per attempt
      - page refresh between attempts re-dismisses the inventory preloader
      - only visible radios are considered for selection
      - no silent fallback to an arbitrary radio
    """
    wait = WebDriverWait(driver, 25)

    def _visible_radios():
        out = []
        try:
            for r in driver.find_elements(By.CSS_SELECTOR, "input[name='add_addresses1']"):
                try:
                    if r.is_displayed() and (r.size or {}).get("height", 0) > 0:
                        out.append(r)
                except StaleElementReferenceException:
                    continue
        except Exception:
            pass
        return out

    def _click_shipto_button():
        try:
            btn = wait.until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "button.pop-myShipTos-1, button[class*='pop-myShipTos']"
            )))
        except TimeoutException:
            btn = wait.until(EC.presence_of_element_located((
                By.XPATH, "//button[contains(., \"Available Ship-To\") or contains(., \"Ship-To\")]"
            )))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        driver.execute_script("arguments[0].click();", btn)   # single JS click only
        log("[DEBUG] Ship-To button JS-clicked")

    opened = False
    for attempt in range(1, max_retries + 1):
        log(f"[DEBUG] Ship-To modal attempt {attempt}/{max_retries}")
        if _visible_radios():
            opened = True
            break
        try:
            _click_shipto_button()
        except Exception as e:
            log(f"[WARN] Could not click Ship-To button: {e}")
        end = time.time() + 6
        while time.time() < end:
            if _visible_radios():
                opened = True
                break
            time.sleep(0.25)
        if opened:
            break
        log(f"[WARN] Ship-To modal not open with visible radios after attempt {attempt}.")
        debug_dump(driver, f"modal_not_open_attempt_{attempt}")
        if attempt < max_retries:
            log("[INFO] Refreshing checkout page and retrying...")
            driver.refresh()
            time.sleep(4)
            try:
                WebDriverWait(driver, 15).until(
                    lambda d: d.execute_script("return document.readyState") == "complete"
                )
            except TimeoutException:
                pass
            _dismiss_checkout_preloader(driver, timeout=20)
            wait_for_overlay_gone(driver, timeout=10)
            time.sleep(1)

    if not opened:
        raise TimeoutException(
            "Ship-To modal never opened with visible radios after multiple attempts."
        )
    log("[DEBUG] Ship-To modal open — radios visible.")

    # ── Select the requested radio (visible instances only) ──────────────────
    all_radios = _visible_radios()
    chosen_radio = None

    for radio in all_radios:
        radio_id = radio.get_attribute("id")
        radio_val = radio.get_attribute("value") or ""
        try:
            label = driver.find_element(By.XPATH, f"//label[input[@id='{radio_id}']]")
            label_text = label.text
        except NoSuchElementException:
            label_text = ""

        if preferred_radio_id and radio_id == preferred_radio_id:
            chosen_radio = radio
            log(f"[INFO] Matched ship-to by radio ID: {radio_id}")
            break
        if preferred_value_contains and preferred_value_contains in radio_val:
            chosen_radio = radio
            log(f"[INFO] Matched ship-to by value substring: '{preferred_value_contains}'")
            break
        if preferred_label_contains and preferred_label_contains in label_text:
            chosen_radio = radio
            log(f"[INFO] Matched ship-to by label substring: '{preferred_label_contains}'")
            break
        if preferred_account_number and f"account_number={preferred_account_number}" in radio_val:
            chosen_radio = radio
            log(f"[INFO] Matched ship-to by account number: {preferred_account_number}")
            break

    if chosen_radio is None:
        # Never gamble on an arbitrary ship-to — a wrong pick means the order
        # ships/bills against the wrong account, silently.
        debug_dump(driver, "shipto_no_match")
        labels = []
        for radio in all_radios[:10]:
            try:
                rid = radio.get_attribute("id")
                lab = driver.find_element(By.XPATH, f"//label[input[@id='{rid}']]").text
                labels.append(f"{rid}: {lab[:80]}")
            except Exception:
                continue
        raise RuntimeError(
            "No ship-to radio matched THE SOURCING GROUP criteria. "
            f"Available options: {labels}"
        )

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", chosen_radio)
    if not chosen_radio.is_selected():
        try:
            safe_click(driver, chosen_radio)
            log("[DEBUG] Radio button clicked successfully")
        except Exception as e:
            log(f"[WARN] Regular click failed: {e}, using JavaScript to select radio...")
            driver.execute_script("""
                arguments[0].checked = true;
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('click', { bubbles: true }));
            """, chosen_radio)
            log("[DEBUG] Radio button selected via JavaScript")


def fill_drop_ship_form(driver, shipto_data: dict):
    """
    Fill in the drop ship form with data from the CSV.
    
    Args:
        driver: Selenium WebDriver instance
        shipto_data: Dict containing address data with keys:
                     company, attention, street, city, state, zip
    """
    wait = WebDriverWait(driver, 15)

    log("[INFO] Filling drop ship form...")

    # Required-field failures are collected and raised at the end — the old
    # behaviour logged nine [WARN] lines when the form never even opened, then
    # declared success and let a wrong/blank-address order sail on.
    failed_required = []

    # 1. Set Country to United States
    try:
        country_select = wait.until(EC.element_to_be_clickable(
            (By.ID, "fm-shipTo-country")
        ))
        Select(country_select).select_by_value("USA")
        log("[INFO] Set country to United States")
        time.sleep(1)  # Wait for state dropdown to populate
    except Exception as e:
        log(f"[WARN] Failed to set country: {e}")
        failed_required.append("country")
    
    # 2. Fill Contact Name (Column K - shipToCompany)
    try:
        contact_name = shipto_data.get('company', '')
        if contact_name:
            contact_input = wait.until(EC.presence_of_element_located(
                (By.ID, "fm-addrbook-contactName")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", contact_input)
            contact_input.clear()
            contact_input.send_keys(contact_name)
            log(f"[INFO] Set contact name: {contact_name}")
    except Exception as e:
        log(f"[WARN] Failed to set contact name: {e}")
        failed_required.append("contact name")
    
    # 3. Fill Address 1 (Column M - shipToStreet)
    try:
        street = shipto_data.get('street', '')
        if street:
            addr1_input = wait.until(EC.presence_of_element_located(
                (By.ID, "fm-shipTo-addr-1")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", addr1_input)
            addr1_input.clear()
            addr1_input.send_keys(street)
            log(f"[INFO] Set address 1: {street}")
    except Exception as e:
        log(f"[WARN] Failed to set address 1: {e}")
        failed_required.append("street")
    
    # 4. Fill Address 2 (Column L - shipToAttention) - max 50 chars, can be blank
    try:
        attention = shipto_data.get('attention', '')
        if attention:
            # Truncate to 50 characters if needed
            attention = attention[:50]
            addr2_input = wait.until(EC.presence_of_element_located(
                (By.ID, "fm-shipTo-addr-2")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", addr2_input)
            addr2_input.clear()
            addr2_input.send_keys(attention)
            log(f"[INFO] Set address 2: {attention}")
    except Exception as e:
        log(f"[WARN] Failed to set address 2: {e}")
    
    # 5. Fill City (Column N - shipToCity)
    try:
        city = shipto_data.get('city', '')
        if city:
            city_input = wait.until(EC.presence_of_element_located(
                (By.ID, "fm-shipTo-city")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", city_input)
            city_input.clear()
            city_input.send_keys(city)
            log(f"[INFO] Set city: {city}")
    except Exception as e:
        log(f"[WARN] Failed to set city: {e}")
        failed_required.append("city")
    
    # 6. Select State (Column O - shipToState)
    try:
        state_abbrev = shipto_data.get('state', '').upper()
        if state_abbrev:
            state_select = wait.until(EC.element_to_be_clickable(
                (By.ID, "fm-shipTo-state")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", state_select)
            # Select by abbreviation value
            Select(state_select).select_by_value(state_abbrev)
            log(f"[INFO] Set state: {state_abbrev}")
    except Exception as e:
        log(f"[WARN] Failed to set state: {e}")
        failed_required.append("state")
    
    # 7. Fill Zip Code (Column P - shipToZip)
    try:
        zipcode = shipto_data.get('zip', '')
        if zipcode:
            # Ensure zip is max 7 characters (some forms have maxlength=7)
            zipcode = str(zipcode)[:7]
            zip_input = wait.until(EC.presence_of_element_located(
                (By.ID, "fm-shipTo-zipcode")
            ))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", zip_input)
            zip_input.clear()
            zip_input.send_keys(zipcode)
            log(f"[INFO] Set zip code: {zipcode}")
    except Exception as e:
        log(f"[WARN] Failed to set zip code: {e}")
        failed_required.append("zip")
    
    # 8. Fill Email field with required addresses
    try:
        email_text = "sales@broberry.com"
        email_input = wait.until(EC.presence_of_element_located(
            (By.ID, "email")
        ))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", email_input)
        email_input.clear()
        email_input.send_keys(email_text)
        log(f"[INFO] Set email: {email_text}")
    except Exception as e:
        log(f"[WARN] Failed to set email: {e}")
    
    # 9. Fill Special Instructions with FedEx Ground number
    try:
        instructions_text = "FedEx Ground 955617339"
        instructions_input = wait.until(EC.presence_of_element_located(
            (By.ID, "fm-shipTo-instructions")
        ))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", instructions_input)
        instructions_input.clear()
        instructions_input.send_keys(instructions_text)
        log(f"[INFO] Set special instructions: {instructions_text}")
    except Exception as e:
        log(f"[WARN] Failed to set special instructions: {e}")

    if failed_required:
        debug_dump(driver, "dropship_form_failed")
        raise RuntimeError(
            f"Drop ship form: required field(s) could not be filled: {', '.join(failed_required)}. "
            "The form probably never opened — NOT submitting this order."
        )

    log("[INFO] Drop ship form filled successfully")


def handle_address_verification_popup(driver, timeout=10):
    """
    Handle the 'Verify Your Address' popup that may appear after order submission.
    
    This popup shows ORIGINAL and SUGGESTED addresses from USPS verification.
    We select the SUGGESTED radio button and click Continue.
    
    Returns True if popup was handled, False if popup didn't appear.
    """
    try:
        # Wait for the address verification popup to appear
        popup = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located(
                (By.ID, "pop-chk-address-verify-1")
            )
        )
        log("[INFO] Address verification popup detected!")
        
        # Check which section is visible (verify or invalid)
        try:
            verify_section = driver.find_element(By.ID, "address-chk-verify-2")
            verify_visible = driver.execute_script(
                "return window.getComputedStyle(arguments[0]).display !== 'none';", 
                verify_section
            )
            
            if verify_visible:
                log("[INFO] USPS suggested address corrections detected.")
                
                # Select the SUGGESTED radio button
                try:
                    suggested_radio = driver.find_element(By.ID, "fm-choutNumbShipTo-suggest-s1")
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", suggested_radio)
                    
                    if not suggested_radio.is_selected():
                        safe_click(driver, suggested_radio)
                        log("[INFO] Selected SUGGESTED address")
                        time.sleep(0.3)
                except NoSuchElementException:
                    log("[WARN] Could not find SUGGESTED radio button, trying ORIGINAL")
                    # Fallback to ORIGINAL if SUGGESTED not found
                    try:
                        original_radio = driver.find_element(By.ID, "fm-choutNumbShipTo-orig-s1")
                        if not original_radio.is_selected():
                            safe_click(driver, original_radio)
                            log("[INFO] Selected ORIGINAL address")
                            time.sleep(0.3)
                    except NoSuchElementException:
                        log("[WARN] Could not find ORIGINAL radio button either")
                
                # Click Continue button
                try:
                    continue_btn = driver.find_element(By.ID, "continue_chk_address")
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", continue_btn)
                    safe_click(driver, continue_btn)
                    log("[INFO] Clicked Continue on address verification popup")
                    time.sleep(1)
                    return True
                except NoSuchElementException:
                    log("[WARN] Could not find Continue button")
                    
        except NoSuchElementException:
            pass
        
        # Check if it's the "Invalid Address" popup instead
        try:
            invalid_section = driver.find_element(By.ID, "address-chk-invalid-2")
            invalid_visible = driver.execute_script(
                "return window.getComputedStyle(arguments[0]).display !== 'none';", 
                invalid_section
            )
            
            if invalid_visible:
                log("[WARN] Invalid address popup detected!")
                log("[INFO] Clicking 'Use as Entered' button...")
                
                try:
                    use_as_entered_btn = driver.find_element(By.ID, "use_as_entered")
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", use_as_entered_btn)
                    safe_click(driver, use_as_entered_btn)
                    log("[INFO] Clicked 'Use as Entered' button")
                    time.sleep(1)
                    return True
                except NoSuchElementException:
                    log("[WARN] Could not find 'Use as Entered' button")
        except NoSuchElementException:
            pass
            
    except TimeoutException:
        # Popup didn't appear - this is normal for many orders
        return False
    except Exception as e:
        log(f"[WARN] Error handling address verification popup: {e}")
        return False
    
    return False


def submit_checkout(driver, timeout=25):
    """Submit the order checkout form."""
    wait = WebDriverWait(driver, timeout)
    
    # 1) Find and click the submit button
    # Try multiple selectors since the button might have different attributes
    submit_btn = None
    
    try:
        # First try: look for the button by ID (most reliable)
        submit_btn = wait.until(EC.element_to_be_clickable((By.ID, "submit_order")))
        log("[INFO] Found submit button by ID")
    except TimeoutException:
        try:
            # Second try: look for validate_checkout_form onclick
            submit_btn = wait.until(EC.element_to_be_clickable((
                By.XPATH, "//button[contains(@onclick,'validate_checkout_form')]"
            )))
            log("[INFO] Found submit button by validate_checkout_form onclick")
        except TimeoutException:
            try:
                # Third try: look for ecat_submit_order onclick (older version)
                submit_btn = wait.until(EC.element_to_be_clickable((
                    By.XPATH, "//button[contains(@onclick,'ecat_submit_order')]"
                )))
                log("[INFO] Found submit button by ecat_submit_order onclick")
            except TimeoutException:
                # Fourth try: look for Submit Order button by text
                submit_btn = wait.until(EC.element_to_be_clickable((
                    By.XPATH, "//button[contains(text(),'Submit Order')]"
                )))
                log("[INFO] Found submit button by text content")
    
    if submit_btn is None:
        raise RuntimeError("Could not find Submit Order button")
    
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", submit_btn)
    safe_click(driver, submit_btn)
    log("[INFO] Clicked Submit Order button")

    # 2) Handle the "Please review your order" confirmation alert
    try:
        WebDriverWait(driver, 5).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        alert.accept()
        log("[INFO] Accepted confirmation alert")
    except TimeoutException:
        pass

    # 3) Handle "DON'T MISS OUT!" free-shipping modal if it shows up
    try:
        popup = WebDriverWait(driver, 4).until(
            EC.visibility_of_element_located(
                (By.ID, "not-all-qualified-pop-alert-1")
            )
        )
        log("[INFO] Free shipping popup detected")
        try:
            proceed_btn = popup.find_element(
                By.XPATH, ".//button[contains(@onclick, 'ecat_submit_order')]"
            )
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", proceed_btn
            )
            safe_click(driver, proceed_btn)
            log("[INFO] Clicked Proceed on free shipping popup")
        except NoSuchElementException:
            pass
    except TimeoutException:
        pass

    # 4) Give the site a beat to process
    time.sleep(0.6)
    wait_for_overlay_gone(driver, timeout=timeout)

    # 5) Handle address verification popup if it appears
    handle_address_verification_popup(driver, timeout=8)

    # 6) Check for error banner
    try:
        err = driver.find_element(By.ID, "submit_order_error_text")
        visible = driver.execute_script(
            "return window.getComputedStyle(arguments[0]).display !== 'none';", err
        )
        if visible:
            raise RuntimeError(
                "Order submission appears to be disabled (error banner shown)."
            )
    except NoSuchElementException:
        pass


def capture_wrangler_order_id(driver, po_number: str, timeout: int = 30) -> str:
    """Grab the vendor Order ID right after placement (added 2026-08-31).

    After submit, Wrangler lands on an order page showing 'Order ID: <id>'
    together with the PO.  Requiring BOTH on the page prevents grabbing some
    other order's id.  Falls back to the Order History list (same lookup
    GetOrderId.py uses).  Returns '' when nothing could be found — the
    GetOrderId.py fallback pass can still fetch it later."""
    po = str(po_number).strip()
    end = time.time() + timeout
    while time.time() < end:
        try:
            body = driver.execute_script("return document.body.innerText") or ""
        except Exception:
            body = ""
        if po and po in body:
            m = re.search(r"Order ID:\s*([A-Za-z0-9]+)", body)
            if m:
                return m.group(1)
        time.sleep(1.0)

    log("[INFO] Order ID not on the post-submit page — checking Order History...")
    try:
        driver.get(ORDER_HISTORY_URL)
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "li.TD-row")))
        block = driver.find_element(
            By.XPATH,
            f"//li[contains(@class,'TD-row') and .//span[em[text()='PO#:'] and contains(., '{po}')]]")
        sid = block.find_element(By.XPATH, ".//span[em[text()='Order ID:']]").text
        return sid.replace("Order ID:", "").strip()
    except Exception as e:
        log(f"[WARN] Could not find Order ID in Order History either: {type(e).__name__}")
        return ""


def update_order_id_in_excel(excel_path: str, row_index: int, order_id: str) -> bool:
    """Append the vendor Order ID to column M (split orders may hold several
    vendors' IDs, so never overwrite).  Retries around Excel file locks and
    falls back to order_ids_fallback.txt so a placed order's ID is never lost."""
    last_err = None
    for attempt in range(1, 4):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            cell = ws.cell(row=row_index + 2, column=13)  # +2: header row, 1-based
            existing = coerce_str(cell.value)
            if order_id in existing.split():
                log(f"[INFO] Order ID {order_id} already recorded in column M.")
                return True
            cell.value = f"{existing} {order_id}".strip()
            wb.save(excel_path)
            log(f"[OK] Column M updated with Wrangler Order ID: {order_id}")
            return True
        except Exception as e:
            last_err = e
            log(f"[ERROR] Failed to update Excel (attempt {attempt}): {e}")
            time.sleep(2)
    try:
        fallback = os.path.join(SCRIPT_DIR, "order_ids_fallback.txt")
        with open(fallback, "a", encoding="utf-8") as f:
            f.write(f"{datetime.datetime.now():%Y-%m-%d %H:%M:%S}  row={row_index + 2}  order_id={order_id}\n")
        log(f"[WARN] Excel locked/unwritable — Order ID saved to {fallback}")
    except Exception as e2:
        log(f"[ERROR] Could not write fallback order-id file either: {e2} (original: {last_err})")
    return False


def verify_order_submitted(driver, timeout=60) -> None:
    """Confirm the checkout actually went through before declaring success.

    The old flow clicked Submit and reported '[OK] Order placed' with no
    verification at all — a silently failed submit looked identical to a
    placed order ("Wrangler just stops and moves on").

    Success signals (either):
      - the browser navigated away from the checkout page (receipt page), or
      - the Submit Order button is gone/hidden with no loading overlay,
        on two consecutive polls (page content replaced in place).
    An error banner, or neither signal within `timeout`, raises.
    """
    log("[INFO] Verifying the order actually submitted...")
    end = time.time() + timeout
    quiet_polls = 0
    while time.time() < end:
        # Error banner → definite failure
        try:
            err = driver.find_element(By.ID, "submit_order_error_text")
            if driver.execute_script(
                "return window.getComputedStyle(arguments[0]).display !== 'none';", err
            ):
                debug_dump(driver, "submit_error_banner")
                raise RuntimeError("Wrangler shows a submit-order error banner — order NOT placed.")
        except NoSuchElementException:
            pass

        url = (driver.current_url or "")
        if "tp_checkout" not in url:
            log(f"[OK] Left the checkout page (now at {url}) — order submitted.")
            return

        # Still on a tp_checkout URL (receipt can share it): check whether the
        # submit button is gone and nothing is still loading.
        try:
            btns = driver.find_elements(By.ID, "submit_order") or driver.find_elements(
                By.XPATH,
                "//button[contains(@onclick,'validate_checkout_form') or contains(@onclick,'ecat_submit_order')]",
            )
            submit_visible = any(b.is_displayed() for b in btns)
        except Exception:
            submit_visible = False
        try:
            overlays = driver.find_elements(
                By.CSS_SELECTOR,
                ".fancybox-overlay, .blockUI, .loading-overlay, [id^='fs-preloader']",
            )
            overlay_visible = any(o.is_displayed() for o in overlays)
        except Exception:
            overlay_visible = False

        if not submit_visible and not overlay_visible:
            quiet_polls += 1
            if quiet_polls >= 2:
                log("[OK] Submit button gone and page settled — order submitted.")
                return
        else:
            quiet_polls = 0
        time.sleep(1.0)

    debug_dump(driver, "submit_unverified")
    raise RuntimeError(
        f"Could not confirm the order submitted within {timeout}s — still on the "
        "checkout page with the Submit button visible. NOT counting this order as placed."
    )


def checkout_and_ship(driver, po_number: str, client_po: str, shipto_data: dict = None,
                      expected_units: int = None):
    """
    Navigate to checkout and handle ship-to selection based on address:
    - If default Sourcing Group address: select radio and proceed normally
    - If non-default address: select Sourcing Group radio, click Select,
      click Drop Ship, and fill the drop ship form
    """
    # Load the extracted ship-to data for this client PO (unless pre-loaded by caller)
    if shipto_data is None:
        shipto_data = load_shipto_data_from_csv(client_po)

    if not shipto_data or not shipto_data.get('shipTo'):
        # Safety net: never silently submit to the default Sourcing Group
        # address when the PO's address data is missing. main() prompts the
        # user (Try Again / Skip) via wait_for_shipto_data() before calling us.
        raise RuntimeError(
            f"No ship-to address available for Client PO {client_po}; "
            "refusing to submit with the default Sourcing Group address."
        )

    is_default_shipto = is_default_sourcing_group_shipto(shipto_data['shipTo'])
    
    # 1) Navigate to checkout
    driver.get(CHECKOUT_URL)
    wait_ready(driver, timeout=25)
    
    # Handle the fs-preloader overlay that checks inventory (takes 3-15 seconds)
    try:
        wait = WebDriverWait(driver, 30)
        
        # First, wait for the preloader div to appear
        log("[INFO] Waiting for inventory check preloader...")
        preloader = wait.until(EC.presence_of_element_located((By.ID, "fs-preloader-1")))
        
        # Wait for the "Continue" section to become visible (display: block)
        # This happens after the inventory check completes
        log("[INFO] Waiting for Continue button to appear...")
        continue_section = wait.until(EC.visibility_of_element_located((By.ID, "fs-preload-continue")))
        
        # Now wait for the Continue button itself to be present and visible
        continue_btn = wait.until(EC.visibility_of_element_located((
            By.XPATH, "//div[@id='fs-preload-continue']//button[@onclick='preloadCloseWindow()']"
        )))
        
        # Try multiple methods to click the button
        log("[INFO] Clicking Continue button on checkout overlay...")
        try:
            # Method 1: Try regular click first
            continue_btn.click()
            log("[INFO] Continue button clicked (regular click)")
        except Exception as e1:
            log(f"[INFO] Regular click failed: {e1}, trying JavaScript click...")
            try:
                # Method 2: Try JavaScript click
                driver.execute_script("arguments[0].click();", continue_btn)
                log("[INFO] Continue button clicked (JavaScript click)")
            except Exception as e2:
                log(f"[INFO] JavaScript click failed: {e2}, trying direct function call...")
                try:
                    # Method 3: Directly execute the onclick function
                    driver.execute_script("preloadCloseWindow();")
                    log("[INFO] Continue button clicked (direct function call)")
                except Exception as e3:
                    log(f"[WARN] All click methods failed: {e3}")
        
        # Wait for the overlay to disappear
        time.sleep(1)
        wait_for_overlay_gone(driver, timeout=10)
        log("[INFO] Checkout overlay dismissed successfully")
        
    except TimeoutException:
        log("[INFO] No Continue overlay found (may have already been dismissed)")
    
    wait = WebDriverWait(driver, 25)
    
    if is_default_shipto:
        # OLD WAY: Just select The Sourcing Group radio button and click Select
        log("[INFO] Default ship-to address detected; using old method (select radio only).")
        open_and_choose_ship_to(
            driver,
            preferred_radio_id="add_addresses-4",
            preferred_value_contains="store=THE SOURCING GROUP",
            preferred_label_contains="THE SOURCING GROUP",
            preferred_account_number="1000263820",
        )
        
        # Click Select button
        try:
            select_btn = wait.until(EC.element_to_be_clickable((
                By.XPATH, "//button[@onclick='return selected_my_shiptos()']"
            )))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", select_btn)
            safe_click(driver, select_btn)
            log("[INFO] Clicked Select button")
        except TimeoutException:
            log("[WARN] Could not find Select button")
        
        # Wait for modal to close
        try:
            wait_modal_close(driver, timeout=12)
        except TimeoutException:
            pass
        
    else:
        # NEW WAY: Select The Sourcing Group, click Select, then Drop Ship, then fill form
        log("[INFO] Non-default ship-to address detected; using new method (Drop Ship form).")
        
        # Step 1: Select The Sourcing Group radio button
        open_and_choose_ship_to(
            driver,
            preferred_radio_id="add_addresses-4",
            preferred_value_contains="store=THE SOURCING GROUP",
            preferred_label_contains="THE SOURCING GROUP",
            preferred_account_number="1000263820",
        )
        
        # Step 2: Click Select button
        try:
            select_btn = wait.until(EC.element_to_be_clickable((
                By.XPATH, "//button[@onclick='return selected_my_shiptos()']"
            )))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", select_btn)
            safe_click(driver, select_btn)
            log("[INFO] Clicked Select button")
        except TimeoutException:
            log("[WARN] Could not find Select button")
        
        # Wait for modal to close
        try:
            wait_modal_close(driver, timeout=12)
        except TimeoutException:
            pass
        
        wait_for_overlay_gone(driver, timeout=20)
        time.sleep(1)
        
        # Step 3: Click Drop Ship button — this MUST succeed for a non-default
        # address; warning-and-continuing here is how orders went out against
        # the wrong ship-to.
        try:
            drop_ship_btn = wait.until(EC.element_to_be_clickable((
                By.XPATH, "//button[@onclick='BTNaddNewShipToAddress()']"
            )))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", drop_ship_btn)
            safe_click(driver, drop_ship_btn)
            log("[INFO] Clicked Drop Ship button")
            time.sleep(1)
        except TimeoutException:
            debug_dump(driver, "dropship_btn_missing")
            raise RuntimeError(
                "Drop Ship button never became clickable — cannot enter the "
                "drop-ship address for this order."
            )
        
        # Step 4: Fill in the drop ship form
        fill_drop_ship_form(driver, shipto_data)
    
    wait_for_overlay_gone(driver, timeout=20)
    time.sleep(0.3)
    
    # Fill PO number (same for both paths)
    try:
        po_input = WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input#fm-shipTo-po-Order1"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", po_input)
        po_input.clear()
        po_input.send_keys(po_number)
    except TimeoutException:
        inputs = wait.until(EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "input[name='po_order_number[]']")
        ))
        for inp in inputs:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
            try:
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable(inp))
                inp.clear()
                inp.send_keys(po_number)
            except TimeoutException:
                driver.execute_script("arguments[0].value = arguments[1];", inp, po_number)
    
    # FINAL QUANTITY GATE: the checkout badge must still match the upload
    # file exactly.  This is the last line of defense against doubled orders.
    if expected_units is not None:
        verify_wrangler_pad(driver, expected_units, "final pre-submit", settle_timeout=15)

    # Auto-submit the order
    log(f"[INFO] Submitting order for PO '{po_number}' (Client PO {client_po})...")
    submit_checkout(driver, timeout=25)
    verify_order_submitted(driver, timeout=60)
    log(f"[OK] Order submitted successfully!")
    time.sleep(0.5)



def main():
    """Main automation script that places orders with Wrangler."""
    log("")
    log("="*60)
    log("*** WRANGLER B2B ORDER AUTOMATION SCRIPT ***")
    log("="*60)
    log(f"Script started at: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log(f"Excel file: {EXCEL_PATH}")
    log(f"PDF directory: {PDFS_DIR}")
    log("="*60)
    
    # Clean up old debug files from previous runs
    cleanup_old_debug_files()
    log("")
    
    options = webdriver.ChromeOptions()
    if TSG_DEBUG:
        options.add_argument(f"--remote-debugging-port={DEBUG_PORT}")
        log(f"[DEBUG] TSG_DEBUG on — Chrome DevTools will listen on 127.0.0.1:{DEBUG_PORT}")
    driver = webdriver.Chrome(options=options)
    try:
        log("[INFO] Logging in to Wrangler B2B...")
        try:
            login(driver)
        except Exception as e:
            log(f"[ERROR] Wrangler login failed: {e}")
            debug_dump(driver, "login_failed")
            debug_hold(driver, f"Login failed: {e}")
            raise
        log("[OK] Login successful!")
        
        log("[INFO] Loading Excel file...")
        df = pd.read_excel(EXCEL_PATH, engine="openpyxl", dtype=str)
        col_g = df.columns[6]   # draft-name / PO column
        col_d = df.columns[3]   # Client PO # (used to locate PDFs_DIR\\<Client PO #>.csv)
        col_j = df.columns[9]   # the raw order field
        col_k = df.columns[10]  # brand column – only process rows containing "Wrangler"

        # Filter to only rows where column K contains "Wrangler" (case-insensitive).
        # NOTE: original sheet indices are kept (no reset_index) — they are what
        # update_order_id_in_excel needs to hit the right row in column M.
        df = df[df[col_k].fillna("").str.contains("Wrangler", case=False, na=False)]

        # track which row/index and PO we processed / skipped
        processed = []
        skipped = []

        total_orders = len(df)
        log("")
        log("="*60)
        log(f"[INFO] Starting Order Placement - {total_orders} Wrangler orders to process")
        log("="*60)

        # Place all orders
        for order_num, (idx, row) in enumerate(df.iterrows(), 1):
            draft_name  = coerce_str(row[col_g])
            client_po   = coerce_str(row[col_d])
            order_field = row[col_j]
            m = re.search(r"\d+", order_field)
            if not m:
                raise ValueError(f"Cannot parse order number from '{order_field}'")
            order_no  = m.group()
            ship_date = get_next_business_day()
            
            log("")
            log(f"[{order_num}/{total_orders}] Processing Order:")
            log(f"  - PO Number: {draft_name}")
            log(f"  - Client PO: {client_po}")
            log(f"  - Order File: {order_no}")

            # CHECKPOINT: skip orders this vendor already placed (survives
            # crashes/restarts; vendor-aware so split orders are safe).
            prior = tsg_runlog.already_placed(SCRIPT_DIR, draft_name, "wrangler")
            if prior:
                log(f"[SKIP] {draft_name} already placed with Wrangler on {prior.get('when','?')} "
                    f"(Order ID: {prior.get('order_id') or 'n/a'}).")
                processed.append((idx, draft_name, client_po))
                continue

            # Require ship-to data BEFORE touching the vendor site so a
            # missing address CSV can be fixed (Try Again) or the order
            # skipped cleanly from the TSG app.
            shipto_data = wait_for_shipto_data(client_po)
            if shipto_data is None:
                skipped.append((idx, draft_name, client_po))
                log(f"[SKIP] Order {draft_name} (Client PO {client_po}) skipped — no address CSV.")
                continue

            try:
                open_order_menu(driver)
                create_new_draft(driver, draft_name, ship_date)

                # CART GUARD: if the draft name already existed, the site just
                # RESUMED that draft — leftover items and all.  Clear before
                # uploading so quantities can never stack across restarts.
                ensure_fresh_wrangler_pad(driver)

                upload_path = upload_batch_order(driver, order_no)
                expected_units = expected_units_from_upload_file(upload_path)
                log(f"[INFO] Upload file expects {expected_units} unit(s).")
                verify_wrangler_pad(driver, expected_units, "after upload")

                checkout_and_ship(driver, draft_name, client_po, shipto_data,
                                  expected_units=expected_units)
            except Exception as e:
                log(f"[ORDER_ERROR] {draft_name} (Client PO {client_po}): {e}")
                import traceback as _tb
                _tb.print_exc()
                debug_dump(driver, f"order_{client_po}")
                action = debug_hold(driver, f"Order {draft_name} failed mid-flow: {e}")
                if action != "continue":
                    raise
                skipped.append((idx, draft_name, client_po))
                continue

            # Record the placement checkpoint IMMEDIATELY (before ID capture:
            # even if the capture hiccups, a re-run must never re-place this).
            tsg_runlog.record_placed(SCRIPT_DIR, draft_name, "wrangler")

            # Capture the vendor Order ID right now instead of relying on the
            # end-of-pipeline GetOrderId pass (which never runs if a later
            # vendor script crashes).
            order_id = capture_wrangler_order_id(driver, draft_name)
            if order_id:
                log(f"[OK] Wrangler Order ID captured at placement: {order_id}")
                update_order_id_in_excel(EXCEL_PATH, idx, order_id)
                tsg_runlog.record_placed(SCRIPT_DIR, draft_name, "wrangler", order_id)
            else:
                log("[WARN] Order ID not captured — the Get Order IDs step can still fetch it later.")

            processed.append((idx, draft_name, client_po))
            log(f"[OK] Order {draft_name} placed successfully!")

            # Small pause between orders to let system settle
            time.sleep(2)

        # Final summary
        log("")
        log("="*60)
        log("*** ORDER PLACEMENT COMPLETE ***")
        log("="*60)
        log(f"Total Orders Placed: {len(processed)}")
        if skipped:
            log(f"Orders Skipped (missing address CSV): {len(skipped)}")
            for _, dn, cpo in skipped:
                log(f"  - {dn} (Client PO {cpo})")
        log(f"Excel File: {EXCEL_PATH}")
        log("")
        log("[INFO] Use 'Get Order IDs' button to fetch Order IDs after placement")
        log("="*60)

    except Exception as e:
        log("")
        log("="*60)
        log("[ERROR] Script encountered an error!")
        log(f"[ERROR] {str(e)}")
        log("="*60)
        raise
    finally:
        log("[INFO] Closing browser...")
        driver.quit()
        log("[INFO] Browser closed.")

if __name__ == "__main__":
    main()
