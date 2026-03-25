#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Merge PMNum.xlsx into Processed_orders.xlsx column G as "<PMNum>-<D value>".
Pairs PMNum A1 -> Processed_orders row 2, A2 -> row 3, etc.

Usage: run this file (double-click or `python add_pm_nums.py`).
Place PMNum.xlsx and Processed_orders.xlsx in the same folder as this script.
"""

from __future__ import annotations
import sys
from pathlib import Path
from typing import List, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl\nInstall with:  pip install openpyxl")
    sys.exit(1)

# === CONFIG ===
BASE_DIR = Path(__file__).resolve().parent
PMNUM_FILE: Path = BASE_DIR / "PMNum.xlsx"
PROCESSED_FILE: Path = BASE_DIR / "Processed_orders.xlsx"


def read_pmnums(path: Path) -> List[str]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        pmnums: List[str] = []
        row = 1
        while True:
            val = ws.cell(row=row, column=1).value  # Column A
            if val is None or str(val).strip() == "":
                break
            pmnums.append(str(val).strip())
            row += 1
        if not pmnums:
            raise ValueError("No PM numbers found in PMNum.xlsx column A.")
        return pmnums
    finally:
        wb.close()


def count_processed_rows_with_d(ws) -> int:
    """Count non-empty D values from row 2 downward until the first blank encountered."""
    count = 0
    row = 2
    while True:
        val = ws.cell(row=row, column=4).value  # Column D
        if val is None or str(val).strip() == "":
            break
        count += 1
        row += 1
    return count


def apply_pairs(pmnums: List[str], processed_path: Path) -> Tuple[int, Path]:
    wb = load_workbook(processed_path, data_only=False)
    try:
        ws = wb.active

        # Validate row count in D against pmnums length
        d_count = count_processed_rows_with_d(ws)
        if d_count != len(pmnums):
            raise ValueError(
                "Row count mismatch:\n"
                f" - PMNum.xlsx non-empty in A: {len(pmnums)}\n"
                f" - Processed_orders.xlsx non-empty in D (from row 2): {d_count}\n"
                "These must match 1-to-1."
            )

        # Apply pairs: PMNum A1 -> G2, A2 -> G3, etc.
        for i, pm in enumerate(pmnums, start=2):
            d_val = ws.cell(row=i, column=4).value  # Column D
            if d_val is None or str(d_val).strip() == "":
                raise ValueError(f"Missing value in Processed_orders.xlsx D{i}.")
            g_text = f"{pm}-{str(d_val).strip()}"
            ws.cell(row=i, column=7).value = g_text  # Column G

        wb.save(processed_path)
        return len(pmnums), processed_path
    finally:
        wb.close()


def main() -> int:
    print("=== Merge PMNum -> Processed_orders (G column) ===")
    print(f"PMNum file:        {PMNUM_FILE.name}")
    print(f"Processed file:    {PROCESSED_FILE.name}")

    if not PMNUM_FILE.exists():
        print(f"ERROR: PMNum file not found at {PMNUM_FILE}")
        return 2
    if not PROCESSED_FILE.exists():
        print(f"ERROR: Processed_orders file not found at {PROCESSED_FILE}")
        return 2

    pmnums = read_pmnums(PMNUM_FILE)
    print(f"Found {len(pmnums)} PM numbers in PMNum.xlsx (A1..A{len(pmnums)}).")

    updated, out_path = apply_pairs(pmnums, PROCESSED_FILE)
    print(f"Wrote {updated} values to column G in {out_path.name} (rows G2..G{updated+1}).")
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
