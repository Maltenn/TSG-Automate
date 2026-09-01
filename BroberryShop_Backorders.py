import os
import re
import time
import argparse
from datetime import datetime, date
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
)
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys

# ─── STATE ABBREVIATION → DROPDOWN VISIBLE NAME ───────────────────────────────
STATE_ABBR_TO_NAME = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming", "DC": "District Of Columbia",
}

# ─── CUSTOM EXCEPTIONS ────────────────────────────────────────────────────────
class UnorderableSizeError(Exception):
    """Raised when the desired size exists on the page but is not orderable (no qty input)."""


# ─── CONFIG ───────────────────────────────────────────────────────────────────
CREDENTIALS = {
    "jmccarthy@thesourcinggroup.com": "TSG2025$",
    "ashotwell@thesourcinggroup.com": "Welcome2TSG!",
    "mdelgado@thesourcinggroup.com": "TSG2024$",
}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_DIR = os.path.join(SCRIPT_DIR, "pdfs")
CSV_DIRS = [PDF_DIR, SCRIPT_DIR]

LOGIN_URL = "https://shop.broberry.com/login"
SUMMARY_URL = "https://shop.broberry.com/shop/order/summary"
ADDRESS_URL = "https://shop.broberry.com/shop/order/address"

SKIPPED_ORDERS_PATH = os.path.join(SCRIPT_DIR, "skipped_orders.xlsx")

PRODUCT_MAP = {
    "3W045CH": {
        "url": "https://shop.broberry.com/shop/product/1462529",
        "sizes": [30, 31, 32, 33, 34, 35, 36, 38, 40, 42, 44, 46],
        "mode": "grid",
    },
    "3W045DK": {
        "url": "https://shop.broberry.com/shop/product/1462564",
        "sizes": [30, 31, 32, 33, 34, 35, 36, 38, 40, 42, 44, 46],
        "mode": "grid",
    },
    "3W060BR": {
        "url": "https://shop.broberry.com/shop/product/1462875",
        "sizes": [30, 31, 32, 33, 34, 35, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54, 56, 58, 60, 62],
        "mode": "grid",
    },
    "10FR13MWZ": {
        "url": "https://shop.broberry.com/shop/product/1444629",
        "sizes": [30, 31, 32, 33, 34, 35, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54],
        "mode": "auto",
    },
    "10FR13MMS": {
        "url": "https://shop.broberry.com/shop/product/1444608",
        "sizes": [30, 31, 32, 33, 34, 35, 36, 38, 40, 42],
        "mode": "auto",
    },
    "10FR47MLW": {
        "url": "https://shop.broberry.com/shop/product/1445120",
        "sizes": [30, 31, 32, 33, 34, 35, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54],
        "mode": "auto",
    },
    "F52944X250": {
        "url": "https://shop.broberry.com/shop/product/1095437",
        "sizes": [30, 32, 34, 36, 38, 40, 42, 44, 46, 48, 50, 52, 54, 56],
        "mode": "auto",
    },
    "F52594X250": {
        "url": "https://shop.broberry.com/shop/product/1094476",
        "sizes": [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24],
        "mode": "length_grid",
    },
    "10030232": {
        "url": "https://shop.broberry.com/shop/product/1083190",
        "sizes": [28, 29, 30, 31, 32, 33, 34, 35, 36, 38, 40, 42, 44],
        "mode": "auto",
    },
}

def sync_product_links_from_main_script():
    """Adopt current product links (and size lists) from BroberryShop.py.

    BroberryShop.py is the source of truth for shop product URLs — they change
    whenever products are republished in the shop, and this file's local copy
    has gone stale before (orders then silently fail as 'unavailable' against
    dead product pages). At startup we import its PRODUCT_MAP and adopt 'url'
    and 'sizes' for SKUs we share. 'mode' is NEVER adopted: modes select
    size-entry BEHAVIOR in this file's own locator code, so they are kept in
    sync by hand, not at runtime. Degrades to the local values if the import
    fails.
    """
    try:
        from BroberryShop import PRODUCT_MAP as MAIN_MAP
    except Exception as e:
        print(f"⚠️  Could not read product links from BroberryShop.py ({e}) — "
              "using this file's local links.")
        return

    updated = []
    for sku, info in PRODUCT_MAP.items():
        main_info = MAIN_MAP.get(sku)
        if not main_info:
            print(f"⚠️  {sku} is not in BroberryShop.py's PRODUCT_MAP — keeping local entry.")
            continue
        if main_info.get("url") and main_info["url"] != info["url"]:
            updated.append(f"{sku}: {info['url']} → {main_info['url']}")
            info["url"] = main_info["url"]
        if main_info.get("sizes") and main_info["sizes"] != info.get("sizes"):
            info["sizes"] = list(main_info["sizes"])
            updated.append(f"{sku}: sizes updated")

    if updated:
        print("🔗 Product links refreshed from BroberryShop.py:")
        for line in updated:
            print(f"   {line}")
        print("   (Consider syncing the PRODUCT_MAP literal in this file too.)")
    else:
        print("✓ Product links match BroberryShop.py")


# ─── AUTOMATIC SUBSTITUTION PAIRS ─────────────────────────────────────────────
# Mirrors PAIRABLE in BroberryShop.py — kept local to match the file's existing
# pattern of carrying its own copy of PRODUCT_MAP.  Bidirectional map: each SKU
# points at the one it may sub for.
PAIRABLE = {
    # CH <-> DK can sub for each other
    "3W045CH": "3W045DK",
    "3W045DK": "3W045CH",

    # 10FR13MWZ <-> 10FR13MMS can sub for each other (back order / unavailable)
    "10FR13MWZ": "10FR13MMS",
    "10FR13MMS": "10FR13MWZ",
}

# ─── LENGTH NORMALISATION (F52594X250: Short/Regular/Long/Unhemmed row labels) ──
# Ported from BroberryShop.py so length_grid products behave identically here.
LENGTH_ALIASES = {
    # Short
    "s": "Short", "sh": "Short", "sht": "Short", "short": "Short",
    # Regular
    "r": "Regular", "reg": "Regular", "regular": "Regular", "regl": "Regular",
    # Long
    "l": "Long", "lng": "Long", "long": "Long",
    # Unhemmed
    "u": "Unhemmed", "unh": "Unhemmed", "unhemmed": "Unhemmed", "unhem": "Unhemmed",
}


def normalize_length(val):
    """Normalise a length abbreviation → 'Short'/'Regular'/'Long'/'Unhemmed'.
    Returns None if the value is not a recognised length alias."""
    if val is None:
        return None
    return LENGTH_ALIASES.get(str(val).strip().lower())


def resolve_length_grid_dims(size1_raw, size2_raw):
    """Determine which CSV field is the numeric size and which is the length label
    for 'length_grid' products (F52594X250).

    Strategy:
      1. If one field is a plain integer and the other normalises to a length → clear win.
      2. If only one is an integer → treat as size; attempt to normalise the other as length.
      3. If only one normalises as a length → treat as length; other assumed to be size.
      4. Ambiguous (both text, both lengths) → prefer size1 as length, warn.

    Returns (numeric_size: int | None, length_label: str | None).
    """
    def _to_int(v):
        try:
            return int(str(v).strip())
        except (ValueError, TypeError):
            return None

    s1 = str(size1_raw).strip() if size1_raw is not None else ""
    s2 = str(size2_raw).strip() if size2_raw is not None else ""

    n1, n2 = _to_int(s1), _to_int(s2)
    l1, l2 = normalize_length(s1), normalize_length(s2)

    # Unambiguous cases
    if n1 is not None and l2 is not None:
        return n1, l2        # size1=number, size2=length  ← typical
    if n2 is not None and l1 is not None:
        return n2, l1        # size2=number, size1=length  ← reversed

    # One side is numeric, length side unrecognised
    if n1 is not None:
        return n1, l2        # l2 may be None; caller will log warning
    if n2 is not None:
        return n2, l1

    # Neither is numeric; both text
    if l1 is not None and l2 is None:
        return None, l1
    if l2 is not None and l1 is None:
        return None, l2

    # Both look like length aliases (e.g. "S", "L") — default: size1=length
    if l1 is not None:
        print(f"⚠️  Ambiguous dims for length_grid product "
              f"(size1={s1!r}, size2={s2!r}). Treating size1 as length.")
        return None, l1

    return None, None


# Module-level flag set by main() from argparse.  When True, process_backorder_csv
# will compare the restock dates of an item and its registered sub and place
# whichever has the sooner date.  See pick_best_sku() below.
PREFER_SOONER_BO = False

# ─── DRIVER SETUP ─────────────────────────────────────────────────────────────
def init_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--incognito")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--disable-save-password-bubble")
    opts.add_argument("--disable-features=AutofillKeyBoardAccessoryView,PasswordManagerOnboarding,OptimizationHints")
    opts.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
    })
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=opts)
    driver.maximize_window()
    driver.implicitly_wait(5)
    return driver


# ─── LOGIN ────────────────────────────────────────────────────────────────────
def login(driver, email, password):
    print(f"⇢ Logging in as {email} …")
    driver.get(LOGIN_URL)
    wait = WebDriverWait(driver, 20)

    email_in = wait.until(EC.element_to_be_clickable((By.NAME, "email")))
    email_in.clear(); email_in.send_keys(email)

    pwd_in = wait.until(EC.element_to_be_clickable((By.NAME, "password")))
    pwd_in.clear(); pwd_in.send_keys(password)

    sign_in_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[@type='submit' and normalize-space(text())='Sign in']")))
    sign_in_btn.click()

    try:
        WebDriverWait(driver, 20).until(
            EC.any_of(
                EC.url_contains("/account"),
                EC.presence_of_element_located((By.XPATH, "//*[contains(.,'Order History') or contains(.,'My Account')]"))
            )
        )
        print(f"→ Logged in as {email}")
        return True
    except TimeoutException:
        print(f"✖ Login did not complete for {email}. Check credentials or MFA prompts.")
        return False


# ─── SKU / QTY HELPERS ────────────────────────────────────────────────────────
def _locate_qty_input_and_context(driver, sku, waist, inseam):
    sizes = PRODUCT_MAP[sku].get("sizes", [])
    mode  = PRODUCT_MAP[sku].get("mode", "auto")

    try:
        waist_i = int(waist) if waist is not None and str(waist).strip() != "" else None
    except Exception:
        waist_i = None
    try:
        inseam_i = int(inseam) if inseam is not None and str(inseam).strip() != "" else None
    except Exception:
        inseam_i = None
    # Raw string form needed for length_grid products (Short/Regular/Long/Unhemmed)
    inseam_str = str(inseam).strip() if inseam is not None else ""

    # Shared helper: derive the 1-based column index for a given waist/size value
    # by scanning the table header.  Defined at outer scope so both try_grid and
    # try_length_grid can call it.
    def _col_index_for_waist(table, w):
        w = str(w).strip()
        xpaths = [
            f".//thead//tr//*[self::th or self::td][normalize-space()='{w}' and not(.//input)]",
            f".//tr[1]//*[self::th or self::td][normalize-space()='{w}' and not(.//input)]",
            f".//*[self::th or self::td][normalize-space()='{w}' and not(.//input)]",
        ]
        for xp in xpaths:
            els = table.find_elements(By.XPATH, xp)
            if not els:
                continue
            el = els[0]
            return len(el.find_elements(By.XPATH, "preceding-sibling::*[self::th or self::td]")) + 1
        return None

    def try_grid():
        if inseam_i is None:
            return None
        if sizes and waist_i not in sizes:
            return None

        header_td = driver.find_element(
            By.XPATH,
            f"//td[contains(@class,'sticky') and normalize-space()='{inseam_i}']"
        )

        row   = header_td.find_element(By.XPATH, "ancestor::tr[1]")
        table = header_td.find_element(By.XPATH, "ancestor::table[1]")
        col_idx = _col_index_for_waist(table, waist_i)

        if col_idx is None and sizes:
            col_idx = sizes.index(waist_i) + 2

        if col_idx is None:
            return None

        row_cells = row.find_elements(By.XPATH, "./*[self::td or self::th]")
        if col_idx < 1 or col_idx > len(row_cells):
            return None
        cell = row_cells[col_idx - 1]
        inputs = cell.find_elements(By.CSS_SELECTOR, "input[type='number']")
        if inputs:
            return inputs[0], cell
        err = UnorderableSizeError(
            f"{sku} size {waist_i}{('x'+str(inseam_i)) if inseam_i is not None else ''} is not orderable (no qty input)"
        )
        err.cell = cell   # the greyed cell — restock fix reads its hidden ids
        raise err

    def try_row():
        row = driver.find_element(By.XPATH, f"//tr[.//*[self::td or self::th][normalize-space()='{waist_i}']]")
        qty_inputs = row.find_elements(By.CSS_SELECTOR, "input[type='number']")
        if not qty_inputs:
            # Only trust "row exists but no qty input" when it's a real product
            # row — header rows of grid tables also contain the bare size text
            # but never carry a $ price pill.
            if not row.find_elements(By.XPATH, ".//span[contains(normalize-space(), '$')]"):
                return None
            err = UnorderableSizeError(f"{sku} size {waist_i} is not orderable (no qty input)")
            err.cell = row
            raise err
        return qty_inputs[0], row

    def try_length_grid():
        """Find the qty input for length_grid products (e.g. F52594X250) whose row
        labels are text (Short / Regular / Long / Unhemmed) rather than integers."""
        # inseam holds the normalised length label; waist holds the numeric size.
        length_label = normalize_length(inseam_str) or inseam_str
        if not length_label or waist_i is None:
            return None

        try:
            header_td = driver.find_element(
                By.XPATH,
                f"//td[contains(@class,'sticky') and normalize-space()='{length_label}']"
            )
        except NoSuchElementException:
            return None

        row   = header_td.find_element(By.XPATH, "ancestor::tr[1]")
        table = header_td.find_element(By.XPATH, "ancestor::table[1]")
        col_idx = _col_index_for_waist(table, waist_i)

        if col_idx is None and sizes:
            try:
                col_idx = sizes.index(waist_i) + 2  # +1 sticky label + 1-based
            except ValueError:
                return None

        if col_idx is None:
            return None

        row_cells = row.find_elements(By.XPATH, "./*[self::td or self::th]")
        if col_idx < 1 or col_idx > len(row_cells):
            return None

        cell   = row_cells[col_idx - 1]
        inputs = cell.find_elements(By.CSS_SELECTOR, "input[type='number']")
        if inputs:
            return inputs[0], cell
        err = UnorderableSizeError(
            f"{sku} size {waist_i} {length_label} is not orderable (no qty input)"
        )
        err.cell = cell
        raise err

    # UnorderableSizeError must propagate out of EVERY mode here (the main
    # BroberryShop.py swallows it for grid/auto): a greyed-out cell is the
    # trigger for the admin restock-date fix on a back-order run.
    if mode == "grid":
        try:
            return try_grid()
        except UnorderableSizeError:
            raise
        except Exception:
            return None
    if mode == "row":
        try:
            return try_row()
        except UnorderableSizeError:
            raise
        except Exception:
            return None
    if mode == "length_grid":
        try:
            return try_length_grid()
        except UnorderableSizeError:
            raise
        except Exception:
            return None

    # auto: try grid first, then row-based
    try:
        res = try_grid()
        if res:
            return res
    except UnorderableSizeError:
        raise
    except Exception:
        pass

    try:
        return try_row()
    except UnorderableSizeError:
        raise
    except Exception:
        return None


def try_add_line(driver, sku, waist, inseam, qty):
    driver.get(PRODUCT_MAP[sku]["url"])
    time.sleep(0.5)

    try:
        located = _locate_qty_input_and_context(driver, sku, waist, inseam)
    except UnorderableSizeError as e:
        # Sold out with NO restock date: the shop renders the size cell as a
        # bare greyed price pill with no qty input.  The caller repairs this
        # through the admin panel (fix_missing_restock_date) and retries.
        return ('greyed_out', str(e))
    if not located:
        return ('unavailable', 'size not found on page')

    qty_input, context = located

    if qty_input.get_attribute("disabled") or qty_input.get_attribute("readonly"):
        return ('unavailable', 'qty input disabled')

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", qty_input)
    # JS focus instead of .click(): each cell now has a hover-triggered price tooltip
    # (absolute-positioned w-80 div, group-hover:flex) that intercepts mouse clicks.
    driver.execute_script("arguments[0].focus();", qty_input)
    qty_input.send_keys(Keys.CONTROL, "a")
    qty_input.send_keys(Keys.DELETE)
    qty_input.send_keys(str(qty))
    qty_input.send_keys(Keys.TAB)

    try:
        form    = context.find_element(By.XPATH, "ancestor::form[1]")
        add_btn = form.find_element(
            By.XPATH,
            ".//button[@type='submit' and (contains(@class,'bg-green-600') or contains(., 'Add'))]"
        )
    except NoSuchElementException:
        add_btn = driver.find_element(
            By.XPATH,
            "//button[@type='submit' and (contains(@class,'bg-green-600') or contains(., 'Add'))]"
        )

    try:
        driver.execute_script("arguments[0].click();", add_btn)
    except Exception:
        add_btn.click()

    time.sleep(1)
    return ('added', None)


# ─── GREYED-OUT SIZE FIX (admin restock-date placeholder) ─────────────────────
# A size that is sold out AND has no Restock At date renders as a bare greyed
# price pill with no qty input, so it cannot be ordered even as a back-order.
# This script only runs after the customer has approved the back order, so
# instead of skipping the line we repair the product in the admin panel: open
# its edit page, set the line's Restock At field to a far-future placeholder,
# save, and retry the add — the shop then renders the size as an orderable
# back-order line.  Mechanics mirror DeleteLineItems.py (edit page → find the
# product_items row → act on it → one Update save).
#
# The placeholder is 12/31/9999, NOT the office convention 99/99/9999:
# restock_at is a DATE column and the admin backend runs the submitted text
# through a date parser — an unparseable 99/99/9999 is silently stored as
# 01/01/1970 (verified live 2026-09-01).  12/31/9999 is the largest date the
# column can hold and reads just as clearly as "no real date yet".
RESTOCK_PLACEHOLDER_DATE = "12/31/9999"
PRODUCT_EDIT_URL_TMPL    = "https://admin.broberry.com/products/edit/{product_id}"

_restock_admin_driver = None   # one admin session, reused for every fix this run


def _get_restock_admin_driver():
    """Boot (once) and return the admin-panel driver used for restock fixes."""
    global _restock_admin_driver
    if _restock_admin_driver is not None:
        try:
            _ = _restock_admin_driver.current_url   # liveness check
            return _restock_admin_driver
        except Exception:
            try:
                _restock_admin_driver.quit()
            except Exception:
                pass
            _restock_admin_driver = None
    driver = _setup_admin_driver()
    _admin_login(driver)
    _restock_admin_driver = driver
    return driver


def shutdown_restock_admin_driver():
    """Close the restock-fix admin session (safe to call when never booted)."""
    global _restock_admin_driver
    if _restock_admin_driver is not None:
        try:
            _restock_admin_driver.quit()
        except Exception:
            pass
        _restock_admin_driver = None


def _harvest_cell_identity(shop_driver, sku, waist, inseam):
    """On the CURRENT shop product page, pull the admin identifiers needed to
    fix a greyed-out size: the product_item id (exact row match) and the
    product id (admin edit-page URL).

    Greyed cells carry NO hidden inputs at all, so the item id is usually
    unavailable — but every orderable cell of the page shares the same
    [product_id] hidden input, which is all we need to open the edit page.

    Returns (item_id or None, product_id or None).
    """
    item_id = product_id = None
    cell = None
    try:
        _locate_qty_input_and_context(shop_driver, sku, waist, inseam)
    except UnorderableSizeError as e:
        cell = getattr(e, "cell", None)
    except Exception:
        pass

    if cell is not None:
        try:
            for h in cell.find_elements(By.CSS_SELECTOR, "input[type='hidden']"):
                name = h.get_attribute("name") or ""
                if name.endswith("[product_id]") and not product_id:
                    product_id = (h.get_attribute("value") or "").strip() or None
                elif name.endswith("[id]") and not item_id:
                    item_id = (h.get_attribute("value") or "").strip() or None
        except Exception:
            pass

    if not product_id:
        try:
            for h in shop_driver.find_elements(
                    By.CSS_SELECTOR, "input[type='hidden'][name$='[product_id]']"):
                val = (h.get_attribute("value") or "").strip()
                if val:
                    product_id = val
                    break
        except Exception:
            pass

    return item_id, product_id


def _find_admin_item_row(admin_driver, item_id, waist, inseam):
    """Locate the product_items <tr> on the admin edit page.

    Prefer the exact product_item id (hidden [id] input value); fall back to
    matching the row's [size]/[width] fields against the ordered waist/inseam
    (admin size == shop waist, admin width == shop inseam; width is empty for
    single-dimension products).  Returns the row element, or None — never
    guesses when the match is ambiguous.
    """
    if item_id:
        for el in admin_driver.find_elements(
                By.CSS_SELECTOR, f"input[name$='[id]'][value='{item_id}']"):
            try:
                return el.find_element(By.XPATH, "./ancestor::tr[1]")
            except Exception:
                continue
        print(f"   ⚠️  Restock fix: no admin row with item id {item_id} — trying size match.")

    waist_s  = "" if waist  is None else str(waist).strip()
    inseam_s = "" if inseam is None else str(inseam).strip()
    matches = []
    for size_in in admin_driver.find_elements(By.CSS_SELECTOR, "input[name$='[size]']"):
        try:
            if (size_in.get_attribute("value") or "").strip() != waist_s:
                continue
            row = size_in.find_element(By.XPATH, "./ancestor::tr[1]")
            width_val = ""
            width_els = row.find_elements(By.CSS_SELECTOR, "input[name$='[width]']")
            if width_els:
                width_val = (width_els[0].get_attribute("value") or "").strip()
            if width_val.lower() == inseam_s.lower():
                matches.append(row)
        except StaleElementReferenceException:
            continue
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        print(f"   ⚠️  Restock fix: {len(matches)} admin rows match "
              f"size {waist_s}/{inseam_s} — refusing to guess.")
    return None


def _read_admin_item_fields(admin_driver, item_id, waist, inseam,
                            fields=("restock_at", "stock")):
    """Re-find the row and return {field: current value} for the given
    product_items input suffixes, or None when the row can't be found."""
    row = _find_admin_item_row(admin_driver, item_id, waist, inseam)
    if row is None:
        return None
    out = {}
    for f in fields:
        els = row.find_elements(By.CSS_SELECTOR, f"input[name$='[{f}]']")
        out[f] = (els[0].get_attribute("value") or "").strip() if els else None
    return out


def fix_missing_restock_date(shop_driver, sku, waist, inseam):
    """Set RESTOCK_PLACEHOLDER_DATE on the admin product line for (sku, waist,
    inseam) so the greyed-out size becomes orderable as a back-order.

    Expects the shop driver to still be on the product page (try_add_line has
    just navigated there).  Returns True only when the saved value was
    re-read from a fresh admin page load and matches the placeholder.
    """
    size_str = f"{waist}{('x' + str(inseam)) if inseam is not None else ''}"
    item_id, product_id = _harvest_cell_identity(shop_driver, sku, waist, inseam)
    if not product_id:
        print(f"   ✖ Restock fix: no [product_id] hidden input on the {sku} page — "
              "cannot open the admin edit page.")
        return False
    print(f"   🛠  Restock fix: {sku} {size_str} → admin product {product_id}"
          + (f", item {item_id}" if item_id else f", matching row by size {size_str}"))

    try:
        admin = _get_restock_admin_driver()
    except Exception as e:
        print(f"   ✖ Restock fix: admin login failed: {e}")
        return False

    edit_url = PRODUCT_EDIT_URL_TMPL.format(product_id=product_id)
    try:
        admin.get(edit_url)
        WebDriverWait(admin, 15).until(EC.presence_of_element_located(
            (By.XPATH, "//input[contains(@name, '[restock_at]')]")))
        time.sleep(0.3)
    except TimeoutException:
        print(f"   ✖ Restock fix: no line-item table on admin edit page for "
              f"product {product_id}.")
        return False

    row = _find_admin_item_row(admin, item_id, waist, inseam)
    if row is None:
        print(f"   ✖ Restock fix: could not find the {sku} {size_str} line on the "
              "admin edit page.")
        return False

    # The shop only renders a qty input when stock >= 0: 0 + restock date is a
    # proper red back-order cell, but a NEGATIVE count (oversell artifact)
    # keeps the cell locked no matter what the restock date says — verified
    # live 2026-09-01.  Normalise negative stock to 0 alongside the date.
    stock_note = ""
    try:
        stock_in = row.find_element(By.CSS_SELECTOR, "input[name$='[stock]']")
        raw_stock = (stock_in.get_attribute("value") or "").strip()
        try:
            stock_val = int(float(raw_stock)) if raw_stock else None
        except ValueError:
            stock_val = None
        if stock_val is not None and stock_val < 0:
            admin.execute_script(
                "arguments[0].scrollIntoView({block:'center', inline:'center'});", stock_in)
            stock_in.clear()
            stock_in.send_keys("0")
            stock_note = f", stock {stock_val} → 0"
    except Exception as e:
        print(f"   ⚠️  Restock fix: could not read/adjust the stock field: {e}")

    try:
        restock_in = row.find_element(By.CSS_SELECTOR, "input[name$='[restock_at]']")
        admin.execute_script(
            "arguments[0].scrollIntoView({block:'center', inline:'center'});", restock_in)
        restock_in.clear()
        restock_in.send_keys(RESTOCK_PLACEHOLDER_DATE)
    except Exception as e:
        print(f"   ✖ Restock fix: could not fill the Restock At field: {e}")
        return False

    try:
        update_btn = WebDriverWait(admin, 12).until(EC.element_to_be_clickable(
            (By.XPATH, "//button[@type='submit' and normalize-space()='Update']")))
        admin.execute_script("arguments[0].scrollIntoView({block:'center'});", update_btn)
        time.sleep(0.1)
        admin.execute_script("arguments[0].click();", update_btn)
    except Exception as e:
        print(f"   ✖ Restock fix: could not click Update: {e}")
        return False

    # The form posts and the page navigates/re-renders; wait for that, then
    # reload the edit page fresh and verify the value actually persisted.
    try:
        WebDriverWait(admin, 10).until(EC.staleness_of(restock_in))
    except TimeoutException:
        time.sleep(2)   # in-place save — give the POST a moment anyway
    try:
        admin.get(edit_url)
        WebDriverWait(admin, 15).until(EC.presence_of_element_located(
            (By.XPATH, "//input[contains(@name, '[restock_at]')]")))
        time.sleep(0.3)
        vals = _read_admin_item_fields(admin, item_id, waist, inseam)
    except Exception as e:
        print(f"   ⚠️  Restock fix: could not verify the save ({e}) — treating as failed.")
        return False

    saved = (vals or {}).get("restock_at") or ""
    saved_stock = (vals or {}).get("stock") or ""
    try:
        stock_ok = int(float(saved_stock)) >= 0
    except (ValueError, TypeError):
        stock_ok = False

    # The backend may re-render the stored date in another format; the year
    # 9999 is proof enough that the placeholder (and not 01/01/1970 from a
    # failed parse, or the old value) is what persisted.
    if "9999" in saved and stock_ok:
        print(f"   ✓ Restock fix saved: {sku} {size_str} restock date = {saved}"
              f"{stock_note} (stock now {saved_stock})")
        return True
    print(f"   ⚠️  Restock fix: saved restock={saved!r} stock={saved_stock!r} — "
          "treating as failed.")
    return False


# ─── RESTOCK DATE LOOKUP (for --prefer-sooner mode) ───────────────────────────
_RESTOCK_DATE_RE = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b")


def read_restock_info(driver, sku, waist, inseam):
    """Look up orderability AND the backorder Restock Date for one
    (sku, waist, inseam) on the product page.

    Returns a dict::

        {"orderable": bool, "restock_date": datetime.date | None}

    - ``orderable`` is True only when the size cell exists on the page AND
      contains an enabled qty <input>.  It is False when the size isn't on
      the grid, the cell has no input, or the input is disabled/readonly.
    - ``restock_date`` is the parsed date from the tooltip's 'Restock Date'
      row; None when no such row is visible (the item is in stock at this
      size) OR when the lookup failed.

    These two flags must be kept distinct: an earlier version returned
    only the date, so "in-stock" and "unorderable at this size" both
    collapsed to None — which caused pick_best_sku to swap to an
    unorderable sub.  Callers MUST check ``orderable`` before acting on
    the date.

    Never raises on failure — a failed lookup must not abort the order.
    """
    info = {"orderable": False, "restock_date": None}

    try:
        driver.get(PRODUCT_MAP[sku]["url"])
        time.sleep(0.4)
    except Exception as e:
        print(f"⚠️  read_restock_info: page load failed for {sku}: {e}")
        return info

    try:
        located = _locate_qty_input_and_context(driver, sku, waist, inseam)
    except UnorderableSizeError:
        return info   # cell exists but has no qty input — not orderable
    except Exception as e:
        print(f"⚠️  read_restock_info: locate failed for {sku} {waist}x{inseam}: {e}")
        return info
    if not located:
        return info   # size not found in grid — not orderable here

    qty_input, context = located
    # Disabled/readonly inputs are not orderable even though the cell exists.
    try:
        if qty_input.get_attribute("disabled") or qty_input.get_attribute("readonly"):
            return info
    except Exception:
        pass
    info["orderable"] = True

    # Hover the cell to reveal the group-hover:flex tooltip inside it.
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", context)
        ActionChains(driver).move_to_element(context).pause(0.25).perform()
    except Exception as e:
        print(f"⚠️  read_restock_info: hover failed for {sku} {waist}x{inseam}: {e}")
        return info   # orderable=True, restock_date=None (unknown)

    # Search the tooltip table for the 'Restock Date' row.  The tooltip lives
    # inside the same cell; scope the XPath to that cell.
    try:
        label_td = WebDriverWait(driver, 2).until(
            lambda d: context.find_element(
                By.XPATH,
                ".//td[normalize-space()='Restock Date']"
            )
        )
    except TimeoutException:
        # No Restock Date row → in-stock at this size.
        return info
    except Exception as e:
        print(f"⚠️  read_restock_info: tooltip scan failed for {sku} {waist}x{inseam}: {e}")
        return info

    try:
        value_td = label_td.find_element(By.XPATH, "following-sibling::td[1]")
        raw = (value_td.text or "").strip()
    except Exception as e:
        print(f"⚠️  read_restock_info: value cell missing for {sku} {waist}x{inseam}: {e}")
        return info

    m = _RESTOCK_DATE_RE.search(raw)
    if not m:
        return info
    token = m.group(1)
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            info["restock_date"] = datetime.strptime(token, fmt).date()
            return info
        except ValueError:
            continue
    print(f"⚠️  read_restock_info: unparsable date '{raw}' for {sku} {waist}x{inseam}")
    return info


def pick_best_sku(driver, sku, waist, inseam):
    """For PAIRABLE SKUs, return whichever of (sku, PAIRABLE[sku]) has the
    sooner/in-stock availability for this specific (waist, inseam).

    Orderability gates the decision — a swap to an UNORDERABLE sub would
    at best trigger a needless admin restock-date fix on the sub (greyed
    cell) and at worst skip the line.  So if the sub cannot accept this
    size (size missing from its grid, cell has no qty input, or input
    disabled), we never swap to it regardless of dates.

    Decision table (both orderable case):
        original in-stock, sub in-stock   → original
        original in-stock, sub dated      → original
        original dated,    sub in-stock   → sub
        both dated                         → earlier date wins; tie → original
    Orderability rules:
        alt not orderable                  → original
        original not orderable, alt ok     → sub
        neither orderable                  → original (caller will surface the failure)
    Non-PAIRABLE SKUs are returned unchanged.
    """
    if sku not in PAIRABLE:
        return sku
    alt = PAIRABLE[sku]
    if alt not in PRODUCT_MAP:
        print(f"⚠️  pick_best_sku: sub '{alt}' for '{sku}' is not in PRODUCT_MAP — using original.")
        return sku

    orig_info = read_restock_info(driver, sku, waist, inseam)
    alt_info  = read_restock_info(driver, alt, waist, inseam)

    size_str = f"{waist}{('x'+str(inseam)) if inseam is not None else ''}"

    def _fmt(info):
        if not info["orderable"]:
            return "unorderable"
        d = info["restock_date"]
        return d.strftime("%m/%d/%Y") if d else "in-stock"

    # Orderability gate — this is what prevents swapping to a sub that
    # doesn't support the requested size (e.g., MMS sizes stop at 42
    # while MWZ goes to 54).
    if not alt_info["orderable"] and orig_info["orderable"]:
        chosen = sku
    elif not orig_info["orderable"] and alt_info["orderable"]:
        chosen = alt
    elif not orig_info["orderable"] and not alt_info["orderable"]:
        chosen = sku   # nothing we can do; let try_add_line report as-is
    else:
        # Both orderable → compare dates (None means in-stock here).
        od = orig_info["restock_date"]
        ad = alt_info["restock_date"]
        if od is None and ad is None:
            chosen = sku
        elif od is None:
            chosen = sku
        elif ad is None:
            chosen = alt
        elif ad < od:
            chosen = alt
        else:
            chosen = sku   # earlier or tied

    if chosen == sku:
        print(f"   · Pick: keeping {sku} {size_str} "
              f"(orig={_fmt(orig_info)}, sub {alt}={_fmt(alt_info)})")
    else:
        print(f"   · Pick: switching {sku} → {alt} at {size_str} "
              f"(orig={_fmt(orig_info)}, sub={_fmt(alt_info)})")
    return chosen


def _coerce_size_int(v):
    """Return int for CSV size cells that look numeric, else None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "nan"):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def merge_pairable_duplicates(driver, df):
    """Pre-pass for --prefer-sooner: collapse same-size rows of a PAIRABLE
    pair into one combined line.

    Rule: for every (pair, waist, inseam) bucket that contains rows with BOTH
    SKUs of the pair present, sum their quantities and emit ONE row using the
    SKU picked by pick_best_sku (sooner restock date).  Rows where only one
    member of the pair appears at a given size are left alone — they still
    go through per-line pick_best_sku in the main loop.  Duplicate rows of
    the SAME SKU (no partner present) are untouched; preserving that avoids
    surprising the caller with unexpected row combining.

    Returns (df, prepicked_indices).  ``df`` has dropped the losing rows
    (index NOT reset, so caller can keep iterating by original index).
    ``prepicked_indices`` names rows whose SKU was decided here; the caller
    should skip its own pick_best_sku for these.
    """
    prepicked = set()
    if df.empty:
        return df, prepicked

    # Bucket rows by (pair_key, waist, inseam).  pair_key is a frozenset so
    # it canonicalises either direction of the pair into the same bucket.
    buckets = {}
    for idx, row in df.iterrows():
        sku = str(row.get("Item-Number", "")).strip()
        if sku not in PAIRABLE:
            continue
        partner = PAIRABLE[sku]
        pair_key = frozenset({sku, partner})
        waist_key = _coerce_size_int(row.get("Size-1"))
        inseam_key = _coerce_size_int(row.get("Size-2"))
        buckets.setdefault((pair_key, waist_key, inseam_key), []).append(idx)

    rows_to_drop = []
    for (pair_key, waist_key, inseam_key), idxs in buckets.items():
        skus_present = {str(df.at[i, "Item-Number"]).strip() for i in idxs}
        # Only merge when BOTH SKUs of the pair actually appear.  Solo-pair
        # rows and duplicate-same-SKU rows are handled elsewhere / untouched.
        if len(skus_present) < 2:
            continue

        # Sum quantities per SKU for the log line, and overall for the winner.
        per_sku_qty = {}
        for i in idxs:
            s = str(df.at[i, "Item-Number"]).strip()
            q_raw = df.at[i, "Qty"]
            q = int(float(q_raw)) if str(q_raw).strip() not in ("", "nan") else 0
            per_sku_qty[s] = per_sku_qty.get(s, 0) + q
        total_qty = sum(per_sku_qty.values())

        # pick_best_sku only needs one side of the pair; it compares both.
        try:
            any_sku = next(iter(pair_key))
            chosen = pick_best_sku(driver, any_sku, waist_key, inseam_key)
        except Exception as e:
            print(f"⚠️  merge pre-pass: pick_best_sku failed for pair {sorted(pair_key)} "
                  f"at {waist_key}x{inseam_key}: {e} — leaving rows untouched.")
            continue

        # Prefer a winner row that already carries `chosen`; otherwise reuse
        # the first index and rewrite its Item-Number.
        winner_idx = next((i for i in idxs if str(df.at[i, "Item-Number"]).strip() == chosen),
                         idxs[0])

        df.at[winner_idx, "Item-Number"] = chosen
        df.at[winner_idx, "Qty"]         = total_qty
        prepicked.add(winner_idx)
        for i in idxs:
            if i != winner_idx:
                rows_to_drop.append(i)

        size_str = f"{waist_key}{'x'+str(inseam_key) if inseam_key is not None else ''}"
        parts = ", ".join(f"{s} qty {q}" for s, q in per_sku_qty.items())
        print(f"   · Merge pair @ {size_str}: {parts} → {chosen} qty {total_qty}")

    if rows_to_drop:
        df = df.drop(index=rows_to_drop)   # keep original index; caller relies on it
    return df, prepicked


# ─── SUMMARY / CART HELPERS ───────────────────────────────────────────────────
def extract_sku_from_text(text):
    for code in PRODUCT_MAP.keys():
        if code in text:
            return code
    return None


def clear_cart(driver):
    """Remove all lines from the cart summary."""
    driver.get(SUMMARY_URL)
    time.sleep(0.8)
    while True:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        if not rows:
            break
        try:
            _remove_summary_row(driver, rows[0])
        except Exception:
            driver.get(SUMMARY_URL)
            time.sleep(0.8)
    print("🧹 Cart cleared.")


def _remove_summary_row(driver, tr):
    wait = WebDriverWait(driver, 10)
    btn  = tr.find_element(By.CSS_SELECTOR, "button.text-rose-600")
    driver.execute_script("arguments[0].click();", btn)
    try:
        delete_xpath = (
            "//button[normalize-space()='Delete' and contains(@class,'bg-red-')]"
            " | //div[contains(@class,'modal') or contains(@role,'dialog')]//button[normalize-space()='Delete']"
            " | //button[normalize-space()='Remove']"
            " | //button[normalize-space()='Yes, delete']"
        )
        delete_btn = wait.until(EC.element_to_be_clickable((By.XPATH, delete_xpath)))
        driver.execute_script("arguments[0].click();", delete_btn)
    except TimeoutException:
        pass
    wait.until(EC.staleness_of(tr))
    time.sleep(0.3)


def has_propper_or_wrangler_items(driver):
    """Check if the current cart contains Propper or Wrangler items."""
    try:
        driver.get(SUMMARY_URL)
        time.sleep(0.5)
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        for tr in rows:
            tds = tr.find_elements(By.CSS_SELECTOR, "td")
            if not tds:
                continue
            item_text = tds[0].text.strip()
            if item_text.startswith(("F52944", "F52594", "3W045", "3W060", "10FR13", "10FR47")):
                return True
        return False
    except Exception as e:
        print(f"⚠️  Could not check for Propper/Wrangler items: {e}")
        return False


# ─── CHECKOUT HELPERS ─────────────────────────────────────────────────────────
def fill_address_and_notes(driver, po, notes, account_email=None,
                           ship_company=None, ship_attention=None,
                           ship_street=None, ship_city=None, ship_state=None, ship_zip=None):
    """Fill PO + notes and the new-address shipping fields.

    Ported from BroberryShop.py: the address page changed — the billing
    section is left untouched (existing address default is kept), and the
    shipping fields only exist after switching to the
    'I want to use a new address' option.
    """
    wait = WebDriverWait(driver, 10)
    driver.get(ADDRESS_URL)

    po_fld  = wait.until(EC.element_to_be_clickable((By.ID, "order-purchase-order")))
    po_fld.clear()
    po_fld.send_keys(str(po))

    notes_f = wait.until(EC.element_to_be_clickable((By.NAME, "order[notes]")))
    notes_f.clear()
    if notes:
        notes_f.send_keys("\n".join(notes))

    # Switch shipping section to "I want to use a new address"
    new_addr_radio = wait.until(EC.presence_of_element_located((
        By.CSS_SELECTOR, 'input[name="user_address[shipping][select_address]"][value="newAddressShipping"]'
    )))
    driver.execute_script("arguments[0].click();", new_addr_radio)

    # Wait for the new-address shipping panel to become visible (required attrs appear on first-name)
    wait.until(EC.visibility_of_element_located((By.ID, "shipping-first-name")))

    def _norm(v):
        if v is None:
            return ""
        s = str(v).strip()
        return "" if (not s or s.lower() == "nan") else s

    # Company line: ship_company + ship_attention combined (carries the real customer name)
    company_line_parts = [p for p in (_norm(ship_company), _norm(ship_attention)) if p]
    company_line = " ".join(company_line_parts).strip()

    def _get_by_id(id_, timeout=2):
        try:
            return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, id_)))
        except Exception:
            return None

    email_el = _get_by_id("shipping-email-address", timeout=2)
    comp_el  = _get_by_id("shipping-company", timeout=2)
    fn_el    = _get_by_id("shipping-first-name", timeout=2)
    ln_el    = _get_by_id("shipping-last-name", timeout=2)
    addr_el  = _get_by_id("shipping-address-1", timeout=2)
    city_el  = _get_by_id("shipping-city", timeout=2)
    zip_el   = _get_by_id("shipping-postal-code", timeout=2)
    st_el    = _get_by_id("shipping-state", timeout=2)
    phone_el = _get_by_id("shipping-phone-no", timeout=2)

    if email_el and _norm(account_email):
        email_el.clear(); email_el.send_keys(_norm(account_email))
    if comp_el and company_line:
        comp_el.clear(); comp_el.send_keys(company_line)
    if fn_el:
        fn_el.clear(); fn_el.send_keys(str(po))
    if ln_el:
        ln_el.clear(); ln_el.send_keys(str(po))
    if addr_el and _norm(ship_street):
        addr_el.clear(); addr_el.send_keys(_norm(ship_street))
    if city_el and _norm(ship_city):
        city_el.clear(); city_el.send_keys(_norm(ship_city))
    if zip_el and _norm(ship_zip):
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", zip_el)
        # JS focus instead of .click(): the right-column order summary overlays the
        # zip field at some scroll positions and intercepts mouse clicks. CTRL+A +
        # send_keys is kept because .clear() is unreliable on type=number inputs.
        driver.execute_script("arguments[0].focus();", zip_el)
        zip_el.send_keys(Keys.CONTROL, "a")
        zip_el.send_keys(_norm(ship_zip))
    if phone_el:
        phone_el.clear(); phone_el.send_keys("407.682.1400")

    if st_el and _norm(ship_state):
        state_in = _norm(ship_state).upper()
        state_name = STATE_ABBR_TO_NAME.get(state_in, _norm(ship_state))
        try:
            from selenium.webdriver.support.ui import Select
            Select(st_el).select_by_visible_text(state_name)
        except Exception:
            pass

    # JS click: the 2-column checkout layout has a short right column, so after
    # scrolling the zip field into view this button ends up above the viewport
    # and a coordinate-based click is rejected.
    continue_btn = wait.until(EC.element_to_be_clickable((
        By.CSS_SELECTOR, "button.w-full.rounded-md.bg-rose-600"
    )))
    driver.execute_script("arguments[0].click();", continue_btn)
    time.sleep(1)


def fill_shipper_number(driver, shipper_number):
    if not shipper_number:
        return
    wait = WebDriverWait(driver, 10)
    try:
        field = wait.until(EC.element_to_be_clickable((By.ID, "order-shipper-number")))
        field.clear()
        field.send_keys(str(shipper_number))
        print(f"✓ Shipper number {shipper_number} added.")
    except Exception as e:
        print(f"⚠️  Could not fill shipper number: {e}")


def _tick_shipping_as_billing_if_present(driver):
    """
    On the summary page, tick the 'ship as billing' checkbox if it exists
    and is not already checked. This must be done before clicking Complete Checkout.
    """
    try:
        cb = WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.ID, "order-is-shipping-as-billing"))
        )
        if not cb.is_selected():
            driver.execute_script("arguments[0].click();", cb)
            print("☑  'Shipping as billing' checkbox ticked.")
            time.sleep(0.5)
    except TimeoutException:
        pass  # checkbox not present on this order — that's fine


def submit_order(driver):
    wait = WebDriverWait(driver, 20)

    # ── Tick the 'ship as billing' checkbox if it's on the page ──────────────
    _tick_shipping_as_billing_if_present(driver)

    # ── Click Complete Checkout ───────────────────────────────────────────────
    try:
        complete_btn = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//button[.//span[normalize-space()='Complete Checkout'] and not(@disabled)]"
        )))
    except TimeoutException:
        complete_btn = wait.until(EC.element_to_be_clickable((
            By.XPATH, "//button[contains(@class,'bg-green-600') and not(@disabled)]"
        )))
    try:
        driver.execute_script("arguments[0].click();", complete_btn)
    except Exception:
        complete_btn.click()

    try:
        WebDriverWait(driver, 25).until(
            EC.any_of(
                EC.url_contains("/shop/order/complete"),
                EC.url_contains("/shop/order/confirmation"),
                EC.presence_of_element_located((
                    By.XPATH, "//*[contains(., 'Thank you for your order') or contains(., 'Order Complete')]"
                ))
            )
        )
        print("✅ Order submitted.")
    except TimeoutException:
        print("⚠️  Submitted click issued, but confirmation not detected. Proceeding.")
    time.sleep(1)


# ─── CSV DISCOVERY ────────────────────────────────────────────────────────────
def _discover_csvs():
    seen, found = set(), []
    for base in CSV_DIRS:
        if not os.path.isdir(base): continue
        for f in os.listdir(base):
            if f.lower().endswith(".csv"):
                p = os.path.join(base, f)
                if p not in seen:
                    seen.add(p); found.append(p)
    return sorted(found, key=lambda p: os.path.basename(p).lower())


def _read_account_from_df(df):
    lower = {c.lower(): c for c in df.columns}
    for key in ("email", "account", "acct"):
        if key in lower:
            return str(df.iloc[0][lower[key]]).strip().lower()
    return str(df.iloc[0][df.columns[0]]).strip().lower()


def _get_col(df, *cands):
    cols_lower = {c.lower(): c for c in df.columns}
    for c in cands:
        if c in df.columns:
            return c
        if c.lower() in cols_lower:
            return cols_lower[c.lower()]
    return None


def discover_csvs_with_accounts():
    items = []
    for p in _discover_csvs():
        try:
            df = pd.read_csv(p)
            acct = _read_account_from_df(df)
            items.append((p, acct))
        except Exception as e:
            print(f"⚠️  Could not read {os.path.basename(p)} ({e}). Skipping.")
    return items


# ─── LOAD SKIPPED POs ─────────────────────────────────────────────────────────
def load_skipped_pos():
    """Return a set of PO strings from skipped_orders.xlsx."""
    if not os.path.exists(SKIPPED_ORDERS_PATH):
        print(f"✖ {SKIPPED_ORDERS_PATH} not found. Nothing to process.")
        return set()
    try:
        df = pd.read_excel(SKIPPED_ORDERS_PATH)
        # column is named 'PO'
        col = None
        for c in df.columns:
            if c.strip().upper() == "PO":
                col = c
                break
        if col is None:
            print("⚠️  skipped_orders.xlsx has no 'PO' column.")
            return set()
        return {str(v).strip() for v in df[col].dropna()}
    except Exception as e:
        print(f"⚠️  Could not read skipped_orders.xlsx: {e}")
        return set()


# ─── ORDER PROCESSING (NO SUBSTITUTION) ──────────────────────────────────────
def process_backorder_csv(driver, csv_path, account_email=None):
    """
    Place a previously-skipped (back-order) order from its CSV.
    - NO substitution / sub logic.
    - Items are added as-is; backorder items remain in the cart.
    - The 'ship as billing' checkbox is handled automatically before checkout.
    """
    print(f"\n=== Processing backorder CSV: {os.path.basename(csv_path)} ===")
    df = pd.read_csv(csv_path)

    # Normalise column names
    column_map = {
        'po': 'PO', 'productId': 'Item-Number',
        'size1': 'Size-1', 'size2': 'Size-2', 'qty': 'Qty'
    }
    df = df.rename(columns=column_map)

    po_number = str(df.iloc[0]["PO"])
    notes = []

    # ── Ship-to fields ──────────────────────────────────────────────────────
    ship_company   = df.iloc[0].get(_get_col(df, "ShipToCompany",   "shipToCompany",   "ShiptoCompany"), "")
    ship_attention = df.iloc[0].get(_get_col(df, "ShipToAttention", "shipToAttention", "ShiptoAttention"), "")
    ship_street    = df.iloc[0].get(_get_col(df, "ShipToStreet",    "shipToStreet",    "ShiptoStreet",
                                              "ShipToAddress1",     "shipToAddress1"), "")
    ship_city      = df.iloc[0].get(_get_col(df, "ShipToCity",  "shipToCity",  "ShiptoCity"), "")
    ship_state     = df.iloc[0].get(_get_col(df, "ShipToState", "shipToState", "ShiptoState"), "")
    ship_zip       = df.iloc[0].get(_get_col(df, "ShipToZip",   "shipToZip",   "ShiptoZip",
                                              "ShipToPostalCode", "shipToPostalCode"), "")

    # --prefer-sooner: collapse same-size rows of PAIRABLE pairs into one
    # combined line BEFORE the add loop so both SKUs ship as a single cart
    # line of the winner.
    prepicked_indices = set()
    if PREFER_SOONER_BO:
        df, prepicked_indices = merge_pairable_duplicates(driver, df)

    # ── Add all lines (no substitution) ────────────────────────────────────
    add_failures = []
    for idx, row in df.iterrows():
        sku     = str(row["Item-Number"]).strip()
        waist_v  = row.get("Size-1", "")
        inseam_v = row.get("Size-2", "")

        if sku not in PRODUCT_MAP:
            print(f"⚠️  Unknown SKU '{sku}' in {os.path.basename(csv_path)} — skipping line.")
            continue

        # length_grid products (F52594X250): one dim is a text label (Short/Regular/Long/Unhemmed)
        # and the other is the numeric size 2-24.  Resolve which is which automatically.
        if PRODUCT_MAP[sku].get("mode") == "length_grid":
            waist, inseam = resolve_length_grid_dims(waist_v, inseam_v)
            if inseam is None:
                print(f"⚠️  Could not determine length for {sku} "
                      f"(size1={waist_v!r}, size2={inseam_v!r}). Skipping line.")
                continue
            if waist is None:
                print(f"⚠️  Could not determine numeric size for {sku} "
                      f"(size1={waist_v!r}, size2={inseam_v!r}). Skipping line.")
                continue
        else:
            # Standard products: size1=waist (int), size2=inseam (int or blank)
            waist  = int(waist_v)  if str(waist_v).strip()  not in ("", "nan") else None
            inseam = int(inseam_v) if str(inseam_v).strip() not in ("", "nan") else None
        qty_v  = row.get("Qty", 0)
        qty    = int(qty_v)   if str(qty_v).strip()    not in ("", "nan") else 0

        # --prefer-sooner: swap to the registered sub if it restocks earlier.
        # Skip when the merge pre-pass already decided this row.
        if PREFER_SOONER_BO and sku in PAIRABLE and idx not in prepicked_indices:
            try:
                sku = pick_best_sku(driver, sku, waist, inseam)
            except Exception as e:
                print(f"⚠️  pick_best_sku failed for {sku} {waist}x{inseam}: {e} — using original.")

        status, reason = try_add_line(driver, sku, waist, inseam, qty)

        if status == 'greyed_out':
            # Sold out with no restock date → no qty input at all.  The
            # customer already approved this back order, so set the 99/99/9999
            # placeholder restock date in the admin panel and retry the add
            # instead of skipping the line.
            print(f"🛠  {reason} — setting placeholder restock date via the admin panel…")
            if fix_missing_restock_date(driver, sku, waist, inseam):
                time.sleep(1.0)   # let the shop pick up the change
                status, reason = try_add_line(driver, sku, waist, inseam, qty)
                if status == 'greyed_out':
                    status, reason = 'unavailable', (
                        f"still greyed out after restock-date fix ({reason})")
            else:
                status, reason = 'unavailable', 'restock-date fix failed'

        if status == 'unavailable':
            # On a back-order run we still skip lines that are completely
            # unavailable (no qty input / not found on page) since those
            # simply cannot be ordered.
            print(f"⚠️  Line {sku} {waist}x{inseam} qty {qty} unavailable — skipping line.")
            add_failures.append({"sku": sku, "waist": waist, "inseam": inseam, "qty": qty})
        else:
            print(f"✓ Added: {sku} {waist}{('x'+str(inseam)) if inseam is not None else ''} x{qty}")

    # ── Navigate to summary to confirm cart has items ───────────────────────
    driver.get(SUMMARY_URL)
    time.sleep(1)
    cart_rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    if not cart_rows:
        print(f"⛔ Cart is empty after adding lines for PO {po_number}. Aborting.")
        return

    # ── Shipper number for Propper / Wrangler items ─────────────────────────
    shipper_number = None
    if has_propper_or_wrangler_items(driver):
        shipper_number = "955617339"
        print(f"✓ Propper/Wrangler items detected — shipper number {shipper_number} will be used.")

    # ── Address / notes page ────────────────────────────────────────────────
    fill_address_and_notes(
        driver, po_number, notes,
        account_email=account_email,
        ship_company=ship_company, ship_attention=ship_attention,
        ship_street=ship_street, ship_city=ship_city,
        ship_state=ship_state, ship_zip=ship_zip
    )

    wait = WebDriverWait(driver, 10)
    # JS click: same 2-column layout issue as the address page — the button can
    # sit outside the viewport and a coordinate-based click is rejected.
    continue_btn = wait.until(EC.element_to_be_clickable((
        By.XPATH, "//button[contains(., 'Continue To Shipping and Payment Method')]"
    )))
    driver.execute_script("arguments[0].click();", continue_btn)

    # ── Shipping method ─────────────────────────────────────────────────────
    try:
        ship_radio = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[name="order[shipping_id]"][value="1"]'))
        )
        driver.execute_script("arguments[0].click();", ship_radio)
    except Exception:
        try:
            ship_radio = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[name="order[shipping_id]"][value="4"]'))
            )
        except Exception:
            ship_radio = WebDriverWait(driver, 2).until(
                EC.element_to_be_clickable((By.ID, "4"))
            )
        driver.execute_script("arguments[0].click();", ship_radio)

    # ── Shipper number (if needed) ──────────────────────────────────────────
    # Must come AFTER the shipping method: the field is only revealed once the
    # Wrangler - Shipper Account option is selected (ported from BroberryShop.py).
    if shipper_number:
        fill_shipper_number(driver, shipper_number)

    # ── Payment method ──────────────────────────────────────────────────────
    pay = wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, 'input[name="order[payment_id]"][value="1"]')
    ))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", pay)
    try:
        wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, 'input[name="order[payment_id]"][value="1"]')
        )).click()
    except ElementClickInterceptedException:
        driver.execute_script("arguments[0].click();", pay)

    review_btn = wait.until(EC.element_to_be_clickable((
        By.XPATH, "//button[contains(., 'Continue and Review Order')]"
    )))
    driver.execute_script("arguments[0].click();", review_btn)

    # ── submit_order handles the checkbox + Complete Checkout ───────────────
    submit_order(driver)


# ═══════════════════════════════════════════════════════════════════════════════
# ─── PM LOGGING PIPELINE (ShoptoPM logic, scoped to skipped POs only) ─────────
# ═══════════════════════════════════════════════════════════════════════════════

# ── Admin credentials / paths ─────────────────────────────────────────────────
PM_INITIALS        = os.getenv("ORDER_USER_INITIALS", "MY")
ADMIN_EMAIL        = os.getenv("BROBERRY_ADMIN_EMAIL", "internal3@broberry.com")
ADMIN_PASSWORD     = os.getenv("BROBERRY_ADMIN_PASSWORD", "MYoung454$")

DOWNLOAD_FOLDER    = os.getenv("TSG_DOWNLOAD_DIR",
                        os.path.join(os.path.expanduser("~"), "Downloads"))
TEMPLATE_XLSX      = os.path.join(SCRIPT_DIR, "Example.xlsx")
OUTPUT_XLSX        = os.path.join(SCRIPT_DIR, "Processed_orders.xlsx")

ADMIN_LOGIN_URL    = "https://admin.broberry.com/login"
ORDERS_URL         = "https://admin.broberry.com/orders"

# ── Vendor detection ──────────────────────────────────────────────────────────
WRANGLER_SKU_PREFIXES = ("3W0", "10FR")
PROPPER_SKU_PREFIXES  = ("F52944X", "F52594X")
ARIAT_EXACT_SKUS      = {"10030232"}

def _detect_vendor_from_sku(sku: str):
    if not sku:
        return None
    s = str(sku).strip().upper()
    if any(s.startswith(p) for p in WRANGLER_SKU_PREFIXES):
        return "wrangler"
    if any(s.startswith(p) for p in PROPPER_SKU_PREFIXES):
        return "propper"
    if s in ARIAT_EXACT_SKUS:
        return "ariat"
    return None

def _detect_vendors_from_df(df: pd.DataFrame):
    col_map = {c.lower().strip(): c for c in df.columns}
    for key in ("productid", "product_id", "sku", "product", "item-number", "item_number"):
        if key in col_map:
            sku_col = col_map[key]
            break
    else:
        return []
    vendors = set()
    for raw in df[sku_col].dropna().astype(str).tolist():
        v = _detect_vendor_from_sku(raw)
        if v:
            vendors.add(v)
    return sorted(vendors)

# ── Build PM records from a list of CSV paths (already matched to skipped POs) ─
def _build_pm_records(matched_csvs):
    """
    matched_csvs: list of (csv_path, acct, po)
    Returns list of dicts ready for admin processing and Excel writing.
    """
    records = []
    for csv_path, acct, po in matched_csvs:
        try:
            df = pd.read_csv(csv_path, dtype=str)
        except Exception:
            try:
                df = pd.read_csv(csv_path, dtype=str, encoding="latin-1")
            except Exception as e:
                print(f"⚠️  PM: Could not read {os.path.basename(csv_path)}: {e}")
                continue

        # Normalise column names (same rename as in the order-placing side)
        df = df.rename(columns={
            'po': 'PO', 'productId': 'Item-Number',
            'size1': 'Size-1', 'size2': 'Size-2', 'qty': 'Qty'
        })

        col_map = {c.lower(): c for c in df.columns}
        row = df.iloc[0]

        # Order cost
        order_cost = ""
        for key in ("order-cost", "order cost", "ordercost", "cost", "total"):
            if key in col_map:
                order_cost = (row.get(col_map[key]) or "").strip()
                break

        records.append({
            "email":      acct,
            "PO":         po,
            "Order-Cost": order_cost,
            "vendors":    _detect_vendors_from_df(df),
        })
        print(f"  PM record: email={acct}, PO={po}, cost={order_cost}, vendors={records[-1]['vendors']}")
    return records

# ── Admin browser setup ───────────────────────────────────────────────────────
def _setup_admin_driver():
    from selenium.webdriver.chrome.options import Options as ChromeOpts
    opts = ChromeOpts()
    os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
    prefs = {
        "download.default_directory":                        DOWNLOAD_FOLDER,
        "download.prompt_for_download":                      False,
        "download.directory_upgrade":                        True,
        "profile.default_content_settings.popups":          0,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "safebrowsing.enabled":                              True,
        "safebrowsing.disable_download_protection":          True,
    }
    opts.add_experimental_option("prefs", prefs)
    opts.add_argument("--safebrowsing-disable-download-protection")
    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()),
        options=opts
    )
    driver.maximize_window()
    return driver

def _admin_login(driver):
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    driver.get(ADMIN_LOGIN_URL)
    WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.NAME, "email")))
    driver.find_element(By.NAME, "email").clear()
    driver.find_element(By.NAME, "email").send_keys(ADMIN_EMAIL)
    driver.find_element(By.NAME, "password").clear()
    driver.find_element(By.NAME, "password").send_keys(ADMIN_PASSWORD)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    WebDriverWait(driver, 20).until(
        EC.invisibility_of_element_located((By.NAME, "email"))
    )
    print("→ Logged into admin panel.")

# ── Fetch order numbers and trigger downloads ─────────────────────────────────
def _admin_process_orders(driver, records):
    """
    For each PM record, search the admin orders page for the PO,
    grab the order number, and click the download links.
    Mirrors ShoptoPM.process_orders exactly.
    """
    driver.get(ORDERS_URL)

    for rec in records:
        po = rec["PO"]
        print(f"\n  Admin: searching for PO {po} …")

        search = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "search"))
        )
        search.clear()
        search.send_keys(po, Keys.RETURN)
        time.sleep(2)

        try:
            table_row = WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
            )
        except TimeoutException:
            print(f"  ⚠️  No result row found for PO {po} — skipping.")
            rec["order_num"] = ""
            continue

        # Extract numeric order number from first cell
        order_num = ""
        try:
            order_p = table_row.find_element(By.XPATH, ".//td[1]//p")
            text = order_p.text.strip()
            if not text.isdigit():
                candidates = [
                    e.text.strip()
                    for e in table_row.find_elements(
                        By.XPATH, ".//td[1]//*[self::p|self::span|self::div]"
                    )
                    if e.text and e.text.strip()
                ]
                text = next((t for t in candidates if t.isdigit()), "")
            order_num = text
        except Exception:
            pass

        if not order_num:
            print(f"  ⚠️  Could not extract order number for PO {po} — skipping downloads.")
            rec["order_num"] = ""
            continue

        rec["order_num"] = order_num
        print(f"  ✓ PO {po} → admin order #{order_num}")

        # Determine which download links to click
        titles  = ["Download XML"]
        vendors = set(rec.get("vendors") or [])

        # Propper + Ariat together → skip vendor XLS to avoid wrong export
        if "propper" in vendors and "ariat" in vendors:
            print(f"  ⚠️  PO {po}: both Propper and Ariat present — skipping vendor XLS.")
            vendors.discard("propper")
            vendors.discard("ariat")

        if "wrangler" in vendors:
            titles.append("Download Wrangler")
        if "propper" in vendors:
            titles.append("Download Propper")
        if "ariat" in vendors:
            titles.append("Download Ariat/Carhartt")

        for title in titles:
            for attempt in range(3):
                try:
                    row = driver.find_element(
                        By.XPATH,
                        f"//tr[.//p[normalize-space()='{order_num}'] "
                        f"or .//a[normalize-space()='{order_num}']]"
                    )
                    link = row.find_element(By.XPATH, f".//a[@title='{title}']")
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", link
                    )
                    driver.execute_script("arguments[0].click();", link)
                    time.sleep(1)
                    print(f"    ✓ Clicked '{title}'")
                    break
                except StaleElementReferenceException:
                    time.sleep(1)
                except Exception:
                    if attempt == 2:
                        print(f"    ⚠️  Could not click '{title}' for order {order_num} — skipping.")
                    time.sleep(1)

        print(f"  ✓ Downloads done for order {order_num}. Waiting 5 s …")
        time.sleep(5)

# ── Append rows to Processed_orders.xlsx ─────────────────────────────────────
def _write_pm_rows(records):
    """
    Appends one row per PM record to Processed_orders.xlsx,
    creating the file from Example.xlsx template if it doesn't exist yet.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from datetime import datetime

    now = datetime.now()
    today = f"{now.month}.{now.day}.{now:%y}"   # e.g. 8.6.26
    # Style the WHOLE row: the template's default font is the Office theme
    # font (Aptos Narrow 11), so unstyled cells would not match Arial 10.
    row_font = Font(name="Arial", size=10)

    if os.path.exists(OUTPUT_XLSX):
        wb = load_workbook(OUTPUT_XLSX)
    elif os.path.exists(TEMPLATE_XLSX):
        wb = load_workbook(TEMPLATE_XLSX)
        print(f"  Created {OUTPUT_XLSX} from template.")
    else:
        # No template — build a minimal workbook with the known headers
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Name", "Customer", "Ack", "Client PO #", "Cust Acct",
            "Who began order", "BMI Order #/ Full retailers PO",
            "Who finalized order", "Date PO finalized",
            "Notes/F/up date & who", "Vendor", "Transaction ID",
            "Order ID", "GP%", "Item Amount", "Freight (CC or N30)"
        ])
        print("  ⚠️  Example.xlsx not found — created bare workbook with default headers.")

    ws      = wb.active
    headers = [cell.value for cell in ws[1]]

    for rec in records:
        row_map = {
            "Name":                              today,
            "Customer":                          "The Sourcing Group",
            "Ack":                               rec["email"],
            "Client PO #":                       rec["PO"],
            "Cust Acct":                         "",
            "Who began order":                   PM_INITIALS,
            "BMI Order #/ Full retailers PO":    "",
            "Who finalized order":               PM_INITIALS,
            "Date PO finalized":                 today,
            "Notes/F/up date & who":             f"Order #: {rec.get('order_num', '')}",
            "Vendor":                            " / ".join(
                                                     rec.get("vendors") or ["Wrangler"]
                                                 ).title(),
            "Transaction ID":                    "Terms",
            "Order ID":                          "",
            "GP%":                               "19%",
            "Item Amount":                       rec["Order-Cost"],
            "Freight (CC or N30)":               "Cust Acct",
        }
        ws.append([row_map.get(h, "") for h in headers])
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).font = row_font
        print(f"  📝 Wrote row: PO={rec['PO']}, order#={rec.get('order_num','')}")

    wb.save(OUTPUT_XLSX)
    print(f"  ✓ Saved {len(records)} row(s) to {OUTPUT_XLSX}")

# ── Entry point called from main() ───────────────────────────────────────────
def run_pm_pipeline(matched_csvs):
    """
    matched_csvs: the same list of (csv_path, acct, po) used by the order placer.
    Logs into admin, fetches order numbers, triggers downloads,
    then appends rows to Processed_orders.xlsx.
    """
    if not matched_csvs:
        print("PM pipeline: no matched CSVs — nothing to log.")
        return

    print("\n" + "═" * 60)
    print("  PM LOGGING PIPELINE — starting …")
    print("═" * 60)

    records = _build_pm_records(matched_csvs)
    if not records:
        print("PM pipeline: no records built — skipping.")
        return

    admin_driver = None
    try:
        admin_driver = _setup_admin_driver()
        _admin_login(admin_driver)
        _admin_process_orders(admin_driver, records)
    except Exception as e:
        print(f"⚠️  PM pipeline admin step failed: {e}")
    finally:
        if admin_driver:
            try:
                admin_driver.quit()
            except Exception:
                pass

    _write_pm_rows(records)
    print("═" * 60)
    print("  PM LOGGING PIPELINE — complete.")
    print("═" * 60 + "\n")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    # ── Keep product links current with BroberryShop.py ────────────────────
    sync_product_links_from_main_script()

    # ── Load skipped POs from the Excel file ───────────────────────────────
    skipped_pos = load_skipped_pos()
    if not skipped_pos:
        print("No skipped orders to process. Exiting.")
        return
    print(f"📋 Found {len(skipped_pos)} skipped PO(s): {', '.join(sorted(skipped_pos))}")

    # ── Find CSVs and match them to skipped POs ────────────────────────────
    all_csvs = discover_csvs_with_accounts()
    if not all_csvs:
        print("No CSV files found in ./pdfs or script directory.")
        return

    # Build a list of (csv_path, account, po) only for CSVs whose PO was skipped
    creds_lower = {k.lower(): v for k, v in CREDENTIALS.items()}
    matched = []
    for csv_path, acct in all_csvs:
        try:
            df = pd.read_csv(csv_path)
            df = df.rename(columns={'po': 'PO', 'productId': 'Item-Number',
                                    'size1': 'Size-1', 'size2': 'Size-2', 'qty': 'Qty'})
            po = str(df.iloc[0]["PO"]).strip()
            if po in skipped_pos:
                matched.append((csv_path, acct, po))
        except Exception as e:
            print(f"⚠️  Could not read {os.path.basename(csv_path)}: {e}")

    if not matched:
        print("None of the CSV files match the skipped POs. Exiting.")
        return
    print(f"→ {len(matched)} CSV(s) matched to skipped POs.")

    # ── Phase 1: Place the back-order orders ──────────────────────────────
    driver = None
    current_account = None
    successfully_placed = []   # track which ones actually went through
    try:
        for csv_path, acct, po in matched:
            acct_norm = (acct or "").strip().lower()
            if acct_norm not in creds_lower:
                print(f"⚠️  {os.path.basename(csv_path)}: unknown account '{acct}'. Skipping.")
                continue

            # Switch account → restart browser
            if current_account != acct_norm:
                if driver:
                    try: driver.quit()
                    except Exception: pass
                driver = init_driver()
                ok = login(driver, acct_norm, creds_lower[acct_norm])
                if not ok:
                    print(f"✖ Login failed for {acct_norm}. Skipping.")
                    current_account = None
                    continue
                current_account = acct_norm
                print(f"→ New session for account {current_account}")

            print(f"→ Using account {current_account} for PO {po} ({os.path.basename(csv_path)})")
            process_backorder_csv(driver, csv_path, account_email=current_account)
            successfully_placed.append((csv_path, acct, po))

    finally:
        if driver:
            try: driver.quit()
            except Exception: pass
        shutdown_restock_admin_driver()

    # ── Phase 2: PM logging for every order we just placed ────────────────
    if successfully_placed:
        run_pm_pipeline(successfully_placed)
    else:
        print("No orders were successfully placed — skipping PM logging.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Place previously-skipped back-order orders from CSVs."
    )
    parser.add_argument(
        "--prefer-sooner",
        action="store_true",
        help=(
            "For SKUs that have a registered substitute in PAIRABLE, compare "
            "the restock dates of the ordered item and its sub and place "
            "whichever has the sooner (earlier) availability.  Without this "
            "flag the script places items exactly as ordered."
        ),
    )
    args = parser.parse_args()
    PREFER_SOONER_BO = bool(args.prefer_sooner)
    if PREFER_SOONER_BO:
        print("🔀 Mode: --prefer-sooner (compare sub restock dates for PAIRABLE items)")
    main()
